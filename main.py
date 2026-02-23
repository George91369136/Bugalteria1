# === ОФИСНЫЙ ДОМ — WEB (PostgreSQL) ===
# -*- coding: utf-8 -*-

import os
import re
import io
import csv
import uuid
import time
import json
import traceback
import threading
import imaplib
import email as email_lib
from email.header import decode_header
import email.utils
from datetime import datetime, date, timedelta
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from typing import Dict, List, Optional, Tuple, Any
from html import escape as h

from wsgiref.simple_server import make_server, WSGIRequestHandler
from urllib.parse import parse_qs, urlencode, quote as urlencode_component
import cgi

import psycopg2
from psycopg2.extras import RealDictCursor
from psycopg2 import pool

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    Workbook = None
    load_workbook = None
    Font = Alignment = PatternFill = Border = Side = get_column_letter = None

try:
    from bs4 import BeautifulSoup
except ImportError:
    BeautifulSoup = None

try:
    from reportlab.pdfgen import canvas
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.lib import colors
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
except ImportError:
    canvas = None
    A4 = None
    mm = None
    colors = None
    pdfmetrics = None
    TTFont = None

try:
    import pdfplumber
except ImportError:
    pdfplumber = None

APP_DIR = os.path.dirname(os.path.abspath(__file__))
DOWNLOAD_DIR = os.path.join(APP_DIR, "_downloads_tmp")
os.makedirs(DOWNLOAD_DIR, exist_ok=True)

OUR_COMPANY_DEFAULT_NAME = "ИП Селецкий"

NAV = [
    ("Операции", [
        ("Банк-клиент", "/bank"),
        ("Касса", "/cash"),
        ("Маркетплейс", "/marketplace"),
    ]),
    ("Документы", [
        ("Акты / накладные", "/acts"),
        ("Платёжные поручения", "/payments"),
        ("Зарплата", "/salary"),
        ("УПД (HTML)", "/upd"),
        ("Передача на реализацию", "/realization"),
    ]),
    ("Отчёты", [
        ("Операц. прибыль", "/reports/op-profit"),
        ("Акт сверки", "/reports/recon"),
        ("КУДиР (PDF)", "/reports/kudir"),
    ]),
    ("Налоги", [
        ("УСН", "/tax/usn"),
    ]),
    ("Справочники", [
        ("Контрагенты", "/counterparties"),
    ]),
]


DB_POOL = None

def init_db_pool():
    global DB_POOL
    if DB_POOL is None:
        DB_POOL = pool.ThreadedConnectionPool(
            minconn=2,
            maxconn=10,
            dsn=os.environ.get("DATABASE_URL")
        )

def get_db_connection():
    global DB_POOL
    if DB_POOL is None:
        init_db_pool()
    return DB_POOL.getconn()

def return_db_connection(conn):
    global DB_POOL
    if DB_POOL is not None:
        try:
            DB_POOL.putconn(conn)
        except:
            pass

from contextlib import contextmanager

@contextmanager
def db_connection():
    """Context manager for safe database connection handling"""
    conn = get_db_connection()
    try:
        yield conn
    finally:
        return_db_connection(conn)


def init_database():
    conn = get_db_connection()
    cur = conn.cursor()
    
    cur.execute("""
        CREATE TABLE IF NOT EXISTS settings (
            id SERIAL PRIMARY KEY,
            key VARCHAR(255) UNIQUE NOT NULL,
            value TEXT
        )
    """)
    
    cur.execute("""
        CREATE TABLE IF NOT EXISTS cp_category_map (
            id SERIAL PRIMARY KEY,
            counterparty_key VARCHAR(500) UNIQUE NOT NULL,
            category VARCHAR(500)
        )
    """)
    
    cur.execute("""
        CREATE TABLE IF NOT EXISTS user_category_map (
            id SERIAL PRIMARY KEY,
            counterparty_key VARCHAR(500) NOT NULL,
            category VARCHAR(500)
        )
    """)
    cur.execute("""
        CREATE INDEX IF NOT EXISTS idx_user_category_map_key ON user_category_map (counterparty_key)
    """)
    
    cur.execute("""
        CREATE TABLE IF NOT EXISTS bank_rows (
            id VARCHAR(64) PRIMARY KEY,
            date_str VARCHAR(20),
            month VARCHAR(10),
            incoming DECIMAL(18,2) DEFAULT 0,
            outgoing DECIMAL(18,2) DEFAULT 0,
            purpose TEXT,
            counterparty TEXT,
            doctype VARCHAR(100),
            skip_outgoing BOOLEAN DEFAULT FALSE,
            category VARCHAR(255),
            cp_inn VARCHAR(20),
            cp_kpp VARCHAR(20),
            cp_account VARCHAR(50),
            cp_bank TEXT,
            cp_bik VARCHAR(20),
            cp_corr VARCHAR(50),
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)
    
    cur.execute("""
        CREATE TABLE IF NOT EXISTS counterparties (
            id VARCHAR(64) PRIMARY KEY,
            kind VARCHAR(100),
            name TEXT NOT NULL,
            inn VARCHAR(20),
            kpp VARCHAR(20),
            bank TEXT,
            bik VARCHAR(20),
            corr VARCHAR(50),
            account VARCHAR(50),
            legal_address TEXT,
            phone VARCHAR(50),
            is_our_company BOOLEAN DEFAULT FALSE,
            full_name TEXT,
            inspection_code VARCHAR(50),
            oktmo VARCHAR(20),
            okato VARCHAR(20),
            signatory TEXT,
            sfr_reg_number VARCHAR(50),
            pfr_reg_self VARCHAR(50),
            pfr_reg_employees VARCHAR(50),
            pfr_terr_code VARCHAR(20),
            pfr_terr_organ TEXT,
            payment_details TEXT,
            okpo VARCHAR(20),
            okopf VARCHAR(20),
            okfs VARCHAR(20),
            okved1 VARCHAR(50),
            okved2 VARCHAR(50),
            okpo_rosstat VARCHAR(20),
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)
    cur.execute("ALTER TABLE counterparties ADD COLUMN IF NOT EXISTS is_our_company BOOLEAN DEFAULT FALSE")
    cur.execute("ALTER TABLE counterparties ADD COLUMN IF NOT EXISTS full_name TEXT")
    cur.execute("ALTER TABLE counterparties ADD COLUMN IF NOT EXISTS inspection_code VARCHAR(50)")
    cur.execute("ALTER TABLE counterparties ADD COLUMN IF NOT EXISTS oktmo VARCHAR(20)")
    cur.execute("ALTER TABLE counterparties ADD COLUMN IF NOT EXISTS okato VARCHAR(20)")
    cur.execute("ALTER TABLE counterparties ADD COLUMN IF NOT EXISTS signatory TEXT")
    cur.execute("ALTER TABLE counterparties ADD COLUMN IF NOT EXISTS sfr_reg_number VARCHAR(50)")
    cur.execute("ALTER TABLE counterparties ADD COLUMN IF NOT EXISTS pfr_reg_self VARCHAR(50)")
    cur.execute("ALTER TABLE counterparties ADD COLUMN IF NOT EXISTS pfr_reg_employees VARCHAR(50)")
    cur.execute("ALTER TABLE counterparties ADD COLUMN IF NOT EXISTS pfr_terr_code VARCHAR(20)")
    cur.execute("ALTER TABLE counterparties ADD COLUMN IF NOT EXISTS pfr_terr_organ TEXT")
    cur.execute("ALTER TABLE counterparties ADD COLUMN IF NOT EXISTS payment_details TEXT")
    cur.execute("ALTER TABLE counterparties ADD COLUMN IF NOT EXISTS okpo VARCHAR(20)")
    cur.execute("ALTER TABLE counterparties ADD COLUMN IF NOT EXISTS okopf VARCHAR(20)")
    cur.execute("ALTER TABLE counterparties ADD COLUMN IF NOT EXISTS okfs VARCHAR(20)")
    cur.execute("ALTER TABLE counterparties ADD COLUMN IF NOT EXISTS okved1 VARCHAR(50)")
    cur.execute("ALTER TABLE counterparties ADD COLUMN IF NOT EXISTS okved2 VARCHAR(50)")
    cur.execute("ALTER TABLE counterparties ADD COLUMN IF NOT EXISTS okpo_rosstat VARCHAR(20)")
    
    cur.execute("""
        CREATE TABLE IF NOT EXISTS acts (
            id VARCHAR(64) PRIMARY KEY,
            doc_no VARCHAR(50),
            doc_date VARCHAR(20),
            executor_id VARCHAR(64),
            customer_id VARCHAR(64),
            direction VARCHAR(50) DEFAULT 'provide',
            executor_json TEXT,
            customer_json TEXT,
            basis TEXT,
            vat_mode VARCHAR(100),
            lines_json TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)
    
    cur.execute("ALTER TABLE acts ADD COLUMN IF NOT EXISTS direction VARCHAR(50) DEFAULT 'provide'")
    cur.execute("ALTER TABLE acts ADD COLUMN IF NOT EXISTS executor_json TEXT")
    cur.execute("ALTER TABLE acts ADD COLUMN IF NOT EXISTS customer_json TEXT")
    
    cur.execute("""
        CREATE TABLE IF NOT EXISTS payment_orders (
            id VARCHAR(64) PRIMARY KEY,
            number VARCHAR(50),
            date_str VARCHAR(20),
            amount DECIMAL(18,2),
            amount_words TEXT,
            payer_json TEXT,
            receiver_json TEXT,
            purpose TEXT,
            pay_type VARCHAR(50),
            vid_op VARCHAR(10),
            ocher VARCHAR(10),
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)
    
    cur.execute("""
        CREATE TABLE IF NOT EXISTS payment_templates (
            id SERIAL PRIMARY KEY,
            name VARCHAR(255) UNIQUE NOT NULL,
            template_json TEXT
        )
    """)
    
    cur.execute("""
        CREATE TABLE IF NOT EXISTS employees (
            id VARCHAR(64) PRIMARY KEY,
            name TEXT NOT NULL,
            inn VARCHAR(20),
            passport VARCHAR(100),
            passport_issued TEXT,
            bank TEXT,
            bik VARCHAR(20),
            corr VARCHAR(50),
            account VARCHAR(50),
            salary DECIMAL(18,2),
            advance DECIMAL(18,2),
            main_part DECIMAL(18,2),
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)
    cur.execute("ALTER TABLE employees ADD COLUMN IF NOT EXISTS inn VARCHAR(20)")
    cur.execute("ALTER TABLE employees ADD COLUMN IF NOT EXISTS passport VARCHAR(100)")
    cur.execute("ALTER TABLE employees ADD COLUMN IF NOT EXISTS passport_issued TEXT")
    cur.execute("ALTER TABLE employees ADD COLUMN IF NOT EXISTS bank TEXT")
    cur.execute("ALTER TABLE employees ADD COLUMN IF NOT EXISTS bik VARCHAR(20)")
    cur.execute("ALTER TABLE employees ADD COLUMN IF NOT EXISTS corr VARCHAR(50)")
    cur.execute("ALTER TABLE employees ADD COLUMN IF NOT EXISTS account VARCHAR(50)")
    cur.execute("ALTER TABLE employees ADD COLUMN IF NOT EXISTS advance DECIMAL(18,2)")
    cur.execute("ALTER TABLE employees ADD COLUMN IF NOT EXISTS main_part DECIMAL(18,2)")
    
    cur.execute("""
        CREATE TABLE IF NOT EXISTS salary_payments (
            id VARCHAR(64) PRIMARY KEY,
            employee_id VARCHAR(64),
            month VARCHAR(10),
            pay_type VARCHAR(20),
            amount DECIMAL(18,2),
            payment_order_id VARCHAR(64),
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)
    cur.execute("ALTER TABLE salary_payments ADD COLUMN IF NOT EXISTS pay_type VARCHAR(20)")
    cur.execute("ALTER TABLE salary_payments ADD COLUMN IF NOT EXISTS payment_order_id VARCHAR(64)")
    cur.execute("ALTER TABLE payment_orders ADD COLUMN IF NOT EXISTS source VARCHAR(50)")
    
    cur.execute("""
        CREATE TABLE IF NOT EXISTS upd_rows (
            id VARCHAR(64) PRIMARY KEY,
            doc_no VARCHAR(50),
            doc_date VARCHAR(20),
            counterparty TEXT,
            inn VARCHAR(20),
            amount DECIMAL(18,2),
            vat DECIMAL(18,2),
            description TEXT,
            source_file TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)
    
    cur.execute("""
        CREATE TABLE IF NOT EXISTS real_rows (
            id VARCHAR(64) PRIMARY KEY,
            doc_no VARCHAR(50),
            doc_date VARCHAR(20),
            counterparty TEXT,
            amount DECIMAL(18,2),
            description TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)
    
    cur.execute("""
        CREATE TABLE IF NOT EXISTS basis_history (
            id SERIAL PRIMARY KEY,
            basis TEXT UNIQUE NOT NULL
        )
    """)
    
    cur.execute("""
        CREATE TABLE IF NOT EXISTS cash_payments (
            id VARCHAR(64) PRIMARY KEY,
            date_str VARCHAR(20),
            nomenclature TEXT,
            amount DECIMAL(18,2) DEFAULT 0,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)
    
    cur.execute("""
        CREATE TABLE IF NOT EXISTS marketplace_rows (
            id VARCHAR(64) PRIMARY KEY,
            platform VARCHAR(50) NOT NULL,
            period_type VARCHAR(20),
            period_label VARCHAR(100),
            year INTEGER,
            quarter INTEGER,
            month INTEGER,
            date_start VARCHAR(20),
            date_end VARCHAR(20),
            amount DECIMAL(18,2) DEFAULT 0,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)
    
    cur.execute("CREATE INDEX IF NOT EXISTS idx_marketplace_rows_platform ON marketplace_rows(platform)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_bank_rows_date ON bank_rows(date_str)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_bank_rows_created ON bank_rows(created_at DESC)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_cash_payments_date ON cash_payments(date_str)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_counterparties_name ON counterparties(name)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_counterparties_inn ON counterparties(inn)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_acts_date ON acts(doc_date)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_payment_orders_date ON payment_orders(date_str)")
    
    conn.commit()
    cur.close()
    return_db_connection(conn)


def new_id() -> str:
    return uuid.uuid4().hex


def norm_text(s: str) -> str:
    s = (s or "").upper().strip()
    s = re.sub(r"\s+", " ", s)
    return s


def norm_spaces(s: str) -> str:
    s = (s or "").strip()
    s = re.sub(r"\s+", " ", s)
    return s


def norm_category(s: str) -> str:
    """Normalize category name: trim, collapse spaces, title case"""
    s = (s or "").strip()
    s = re.sub(r"\s+", " ", s)
    if not s:
        return "Прочее"
    return s.title()


def parse_date_ddmmyyyy(s: str) -> Optional[date]:
    if not s:
        return None
    m = re.search(r"^(\d{1,2})\.(\d{1,2})\.(\d{4})$", s.strip())
    if not m:
        return None
    dd, mm_, yyyy = m.groups()
    try:
        return date(int(yyyy), int(mm_), int(dd))
    except ValueError:
        return None


def format_ddmmyyyy(d: Optional[date]) -> str:
    return d.strftime("%d.%m.%Y") if isinstance(d, date) else ""


def month_from_date_str(date_str: str) -> str:
    d = parse_date_ddmmyyyy(date_str)
    if not d:
        return ""
    return f"{d.month:02d}.{d.year}"


def decimal_from_str(s) -> Decimal:
    if s is None:
        return Decimal("0")
    s = str(s).replace("\xa0", "").replace(" ", "").replace(",", ".")
    if not s:
        return Decimal("0")
    try:
        return Decimal(s)
    except InvalidOperation:
        return Decimal("0")


def money2(d: Decimal) -> Decimal:
    return d.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


def _group_thousands(num: str) -> str:
    if not num:
        return "0"
    out = []
    i = len(num)
    while i > 0:
        j = max(0, i - 3)
        out.append(num[j:i])
        i = j
    return ".".join(reversed(out))


def fmt_num(value: Any, decimals: int = 2, strip_trailing_zeros: bool = False) -> str:
    if value is None:
        return ""
    try:
        if isinstance(value, Decimal):
            d = value
        elif isinstance(value, (int, float)):
            d = Decimal(str(value))
        else:
            d = decimal_from_str(value)
    except Exception:
        return ""

    sign = "-" if d < 0 else ""
    d = abs(d)

    q = Decimal("1") if decimals <= 0 else Decimal("1." + ("0" * decimals))
    d = d.quantize(q, rounding=ROUND_HALF_UP)

    s = format(d, "f")
    if "." in s:
        int_part, frac_part = s.split(".", 1)
    else:
        int_part, frac_part = s, ""

    int_part = _group_thousands(int_part)

    if decimals <= 0:
        return sign + int_part

    if strip_trailing_zeros:
        frac_part = frac_part.rstrip("0")
        if not frac_part:
            return sign + int_part

    return sign + int_part + "," + frac_part[:decimals]


def fmt_money(value: Any) -> str:
    return fmt_num(value, decimals=2, strip_trailing_zeros=False)


def fmt_int(value: Any) -> str:
    return fmt_num(value, decimals=0, strip_trailing_zeros=True)


def require_openpyxl():
    if Workbook is None:
        raise RuntimeError("Нужен openpyxl: pip install openpyxl")


def excel_autosize(ws, max_width: int = 60):
    if get_column_letter is None:
        return
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            v = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(v))
        ws.column_dimensions[col_letter].width = min(max_len + 2, max_width)


def require_reportlab():
    if canvas is None:
        raise RuntimeError("Нужен reportlab: pip install reportlab")


def register_ru_font() -> str:
    require_reportlab()
    assert pdfmetrics is not None and TTFont is not None

    for name in ("RU", "DejaVuSans", "Arial"):
        try:
            pdfmetrics.getFont(name)
            return name
        except Exception:
            pass

    candidates = [
        "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
        "/usr/share/fonts/TTF/DejaVuSans.ttf",
        os.path.join(APP_DIR, "DejaVuSans.ttf"),
    ]
    for p in candidates:
        if os.path.exists(p):
            try:
                pdfmetrics.registerFont(TTFont("RU", p))
                return "RU"
            except Exception:
                continue

    return "Helvetica"


def safe_filename(name: str) -> str:
    name = re.sub(r"[^0-9A-Za-zА-Яа-я_\-\. ]+", "_", name or "")
    name = name.strip().replace(" ", "_")
    return name or "file"


def redirect(location: str, start_response):
    start_response("302 Found", [("Location", location)])
    return [b""]


def read_body(environ) -> bytes:
    try:
        length = int(environ.get("CONTENT_LENGTH") or "0")
    except Exception:
        length = 0
    return environ["wsgi.input"].read(length) if length > 0 else b""


def parse_post_form(environ) -> Dict[str, Any]:
    ctype = (environ.get("CONTENT_TYPE") or "").lower()
    if "multipart/form-data" in ctype:
        fs = cgi.FieldStorage(fp=environ["wsgi.input"], environ=environ, keep_blank_values=True)
        out: Dict[str, Any] = {}
        for k in fs.keys():
            out[k] = fs[k]
        return out

    body = read_body(environ).decode("utf-8", errors="replace")
    q = parse_qs(body, keep_blank_values=True)
    return {k: (v[0] if isinstance(v, list) and v else "") for k, v in q.items()}


def first_fs(v: Any) -> Any:
    if isinstance(v, list):
        return v[0] if v else None
    return v


def get_upload(form: Dict[str, Any], key: str = "file") -> Any:
    v = first_fs(form.get(key))
    if v is None:
        return None
    if getattr(v, "file", None) is None:
        return None
    return v


def qs(environ) -> Dict[str, str]:
    q = parse_qs(environ.get("QUERY_STRING", ""), keep_blank_values=True)
    return {k: (v[0] if v else "") for k, v in q.items()}


def not_found(start_response):
    start_response("404 Not Found", [("Content-Type", "text/html; charset=utf-8")])
    return [b"<h1>404</h1>"]


def serve_text(start_response, html: str, status: str = "200 OK"):
    data = html.encode("utf-8")
    start_response(status, [
        ("Content-Type", "text/html; charset=utf-8"),
        ("Content-Length", str(len(data))),
        ("Cache-Control", "no-cache, no-store, must-revalidate"),
    ])
    return [data]


def serve_file_download(start_response, file_path: str, content_type: str):
    if not os.path.exists(file_path):
        return not_found(start_response)
    with open(file_path, "rb") as f:
        data = f.read()
    filename = os.path.basename(file_path)
    headers = [("Content-Type", content_type), ("Content-Length", str(len(data)))]
    headers.append(("Content-Disposition", f'attachment; filename="{filename}"'))
    start_response("200 OK", headers)
    return [data]


_INN_RE = re.compile(r"(?:\bИНН\b[:\s]*)(\d{10}|\d{12})", re.IGNORECASE)

def split_inn_from_name(name: str) -> Tuple[str, str]:
    s = (name or "").strip()
    if not s:
        return ("", "")

    inn = ""

    m = _INN_RE.search(s)
    if m:
        inn = m.group(1)
        s = _INN_RE.sub(" ", s).strip()

    if not inn:
        m2 = re.match(r"^\s*(\d{10}|\d{12})\s*([,;:\-–—]|\s)+\s*(.+)$", s)
        if m2:
            inn = m2.group(1)
            s = (m2.group(3) or "").strip()

    if not inn:
        m3 = re.match(r"^\s*(\d{10}|\d{12})([A-Za-zА-Яа-я\"«].+)$", s)
        if m3:
            inn = m3.group(1)
            s = (m3.group(2) or "").strip()

    if not inn and re.search(r"[A-Za-zА-Яа-я]", s):
        m4 = re.search(r"\b(\d{10}|\d{12})\b", s)
        if m4:
            inn = m4.group(1)
            s = re.sub(r"\b" + re.escape(inn) + r"\b", " ", s).strip()

    s = re.sub(r"\bИНН\b[:\s]*", " ", s, flags=re.IGNORECASE)
    s = re.sub(r"\s+", " ", s).strip()
    s = s.strip(" ,;:-–—")

    return (s, inn)


def load_counterparty_category_map(path: str) -> Dict[str, str]:
    ext = os.path.splitext(path)[1].lower().strip()
    mapping: Dict[str, str] = {}

    def add_pair(cat: str, cp: str):
        cat = norm_spaces(cat)
        cp = norm_spaces(cp)
        if not cat or not cp:
            return
        mapping[norm_text(cp)] = cat

    if ext == ".xlsx":
        require_openpyxl()
        wb = load_workbook(path, data_only=True)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            cat = row[0] if len(row) > 0 else ""
            cp = row[1] if len(row) > 1 else ""
            add_pair(str(cat or ""), str(cp or ""))
        return mapping

    if ext == ".csv":
        with open(path, "r", encoding="cp1251", errors="replace") as f:
            sample = f.read(4096)
            f.seek(0)
            delim = ";" if sample.count(";") >= sample.count(",") else ","
            reader = csv.reader(f, delimiter=delim)
            try:
                next(reader)
            except StopIteration:
                return mapping
            for row in reader:
                cat = row[0] if len(row) > 0 else ""
                cp = row[1] if len(row) > 1 else ""
                add_pair(cat, cp)
        return mapping

    raise ValueError("Поддерживаются только .xlsx или .csv")


class AccountingStateDB:
    def __init__(self):
        self._last_saved_at = ""
        self._cache = {}
        self._cache_ttl = 30
        self._cache_time = {}
    
    def _get_cache(self, key):
        if key in self._cache and key in self._cache_time:
            if time.time() - self._cache_time[key] < self._cache_ttl:
                return self._cache[key]
        return None
    
    def _set_cache(self, key, value):
        self._cache[key] = value
        self._cache_time[key] = time.time()
    
    def _clear_cache(self, key=None):
        if key:
            self._cache.pop(key, None)
            self._cache_time.pop(key, None)
        else:
            self._cache.clear()
            self._cache_time.clear()
    
    @property
    def last_saved_at(self):
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute("SELECT value FROM settings WHERE key = 'last_saved_at'")
        row = cur.fetchone()
        cur.close()
        return_db_connection(conn)
        return row[0] if row else "—"
    
    def save(self):
        now = datetime.now().strftime("%d.%m.%Y %H:%M:%S")
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO settings (key, value) VALUES ('last_saved_at', %s)
            ON CONFLICT (key) DO UPDATE SET value = %s
        """, (now, now))
        conn.commit()
        cur.close()
        return_db_connection(conn)
    
    @property
    def settings(self) -> Dict[str, Any]:
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute("SELECT key, value FROM settings")
        rows = cur.fetchall()
        cur.close()
        return_db_connection(conn)
        result = {"our_company_id": None}
        for key, value in rows:
            result[key] = value
        return result
    
    def set_setting(self, key: str, value: str):
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO settings (key, value) VALUES (%s, %s)
            ON CONFLICT (key) DO UPDATE SET value = %s
        """, (key, value, value))
        conn.commit()
        cur.close()
        return_db_connection(conn)
    
    @property
    def cp_category_map(self) -> Dict[str, str]:
        cached = self._get_cache("cp_category_map")
        if cached is not None:
            return cached
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute("SELECT counterparty_key, category FROM cp_category_map")
        rows = cur.fetchall()
        cur.close()
        return_db_connection(conn)
        result = {row[0]: row[1] for row in rows}
        self._set_cache("cp_category_map", result)
        return result
    
    @cp_category_map.setter
    def cp_category_map(self, value: Dict[str, str]):
        self._clear_cache("cp_category_map")
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute("DELETE FROM cp_category_map")
        for k, v in value.items():
            cur.execute("""
                INSERT INTO cp_category_map (counterparty_key, category) VALUES (%s, %s)
            """, (k, v))
        conn.commit()
        cur.close()
        return_db_connection(conn)
    
    @property
    def cp_map_source(self) -> str:
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute("SELECT value FROM settings WHERE key = 'cp_map_source'")
        row = cur.fetchone()
        cur.close()
        return_db_connection(conn)
        return row[0] if row else ""
    
    @cp_map_source.setter
    def cp_map_source(self, value: str):
        self.set_setting("cp_map_source", value)
    
    @property
    def user_category_map(self) -> Dict[str, List[str]]:
        """Returns dict mapping counterparty_key -> list of categories"""
        cached = self._get_cache("user_category_map")
        if cached is not None:
            return cached
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute("SELECT counterparty_key, category FROM user_category_map ORDER BY id")
        rows = cur.fetchall()
        cur.close()
        return_db_connection(conn)
        result: Dict[str, List[str]] = {}
        for row in rows:
            key, cat = row[0], row[1]
            if key not in result:
                result[key] = []
            if cat and cat not in result[key]:
                result[key].append(cat)
        self._set_cache("user_category_map", result)
        return result
    
    @user_category_map.setter
    def user_category_map(self, value: Dict[str, Any]):
        self._clear_cache("user_category_map")
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute("DELETE FROM user_category_map")
        for k, v in value.items():
            cats = v if isinstance(v, list) else [v]
            for cat in cats:
                cur.execute("""
                    INSERT INTO user_category_map (counterparty_key, category) VALUES (%s, %s)
                """, (k, cat))
        conn.commit()
        cur.close()
        return_db_connection(conn)
    
    def add_user_category(self, cp_key: str, category: str):
        """Add a category for counterparty (allows multiple categories per counterparty)"""
        self._clear_cache("user_category_map")
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute("SELECT id FROM user_category_map WHERE counterparty_key = %s AND category = %s", (cp_key, category))
        if cur.fetchone():
            cur.close()
            return_db_connection(conn)
            return
        cur.execute("""
            INSERT INTO user_category_map (counterparty_key, category) VALUES (%s, %s)
        """, (cp_key, category))
        conn.commit()
        cur.close()
        return_db_connection(conn)
    
    def get_user_categories(self, cp_key: str) -> List[str]:
        """Get all categories for a counterparty"""
        return self.user_category_map.get(cp_key, [])
    
    def remove_user_category(self, cp_key: str, category: str):
        """Remove a category from counterparty"""
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute("DELETE FROM user_category_map WHERE counterparty_key = %s AND category = %s", (cp_key, category))
        conn.commit()
        cur.close()
        return_db_connection(conn)
        if cp_key in self.user_category_map:
            self.user_category_map[cp_key] = [c for c in self.user_category_map[cp_key] if c != category]
            if not self.user_category_map[cp_key]:
                del self.user_category_map[cp_key]
    
    @property
    def bank_rows(self) -> List[Dict[str, Any]]:
        cached = self._get_cache("bank_rows")
        if cached is not None:
            return cached
        conn = get_db_connection()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        cur.execute("""
            SELECT id, date_str as date, month, incoming, outgoing, purpose, counterparty,
                   doctype, skip_outgoing, category, cp_inn, cp_kpp, cp_account, cp_bank, cp_bik, cp_corr
            FROM bank_rows ORDER BY created_at DESC
        """)
        rows = cur.fetchall()
        cur.close()
        return_db_connection(conn)
        result = [dict(row) for row in rows]
        self._set_cache("bank_rows", result)
        return result
    
    def add_bank_row(self, r: Dict[str, Any]):
        self._clear_cache("bank_rows")
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO bank_rows (id, date_str, month, incoming, outgoing, purpose, counterparty,
                                   doctype, skip_outgoing, category, cp_inn, cp_kpp, cp_account, cp_bank, cp_bik, cp_corr)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            ON CONFLICT (id) DO NOTHING
        """, (
            r.get("id"), r.get("date"), r.get("month"),
            str(decimal_from_str(r.get("incoming") or 0)), str(decimal_from_str(r.get("outgoing") or 0)),
            r.get("purpose"), r.get("counterparty"), r.get("doctype"),
            bool(r.get("skip_outgoing")), r.get("category"),
            r.get("cp_inn"), r.get("cp_kpp"), r.get("cp_account"),
            r.get("cp_bank"), r.get("cp_bik"), r.get("cp_corr")
        ))
        conn.commit()
        cur.close()
        return_db_connection(conn)
    
    def update_bank_row(self, rid: str, updates: Dict[str, Any]):
        if not updates:
            return
        self._clear_cache("bank_rows")
        conn = get_db_connection()
        cur = conn.cursor()
        
        set_parts = []
        values = []
        for k, v in updates.items():
            if k == "date":
                set_parts.append("date_str = %s")
            elif k == "skip_outgoing":
                set_parts.append("skip_outgoing = %s")
                v = bool(v)
            else:
                set_parts.append(f"{k} = %s")
            values.append(v)
        
        values.append(rid)
        cur.execute(f"UPDATE bank_rows SET {', '.join(set_parts)} WHERE id = %s", values)
        conn.commit()
        cur.close()
        return_db_connection(conn)
    
    def get_bank_row(self, rid: str) -> Optional[Dict[str, Any]]:
        conn = get_db_connection()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        cur.execute("""
            SELECT id, date_str as date, month, incoming, outgoing, purpose, counterparty,
                   doctype, skip_outgoing, category, cp_inn, cp_kpp, cp_account, cp_bank, cp_bik, cp_corr
            FROM bank_rows WHERE id = %s
        """, (rid,))
        row = cur.fetchone()
        cur.close()
        return_db_connection(conn)
        return dict(row) if row else None
    
    @property
    def cash_rows(self) -> List[Dict[str, Any]]:
        conn = get_db_connection()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        cur.execute("""
            SELECT id, date_str as date, nomenclature, amount
            FROM cash_payments ORDER BY date_str DESC, created_at DESC
        """)
        rows = cur.fetchall()
        cur.close()
        return_db_connection(conn)
        return [dict(row) for row in rows]
    
    def add_cash_row(self, r: Dict[str, Any]):
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO cash_payments (id, date_str, nomenclature, amount)
            VALUES (%s, %s, %s, %s)
            ON CONFLICT (id) DO NOTHING
        """, (
            r.get("id"), r.get("date"), r.get("nomenclature"),
            str(decimal_from_str(r.get("amount") or 0))
        ))
        conn.commit()
        cur.close()
        return_db_connection(conn)
    
    def update_cash_row(self, rid: str, updates: Dict[str, Any]):
        if not updates:
            return
        conn = get_db_connection()
        cur = conn.cursor()
        set_parts = []
        values = []
        for k, v in updates.items():
            if k == "date":
                set_parts.append("date_str = %s")
            else:
                set_parts.append(f"{k} = %s")
            values.append(v)
        values.append(rid)
        cur.execute(f"UPDATE cash_payments SET {', '.join(set_parts)} WHERE id = %s", values)
        conn.commit()
        cur.close()
        return_db_connection(conn)
    
    def delete_cash_row(self, rid: str):
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute("DELETE FROM cash_payments WHERE id = %s", (rid,))
        conn.commit()
        cur.close()
        return_db_connection(conn)
    
    def get_cash_row(self, rid: str) -> Optional[Dict[str, Any]]:
        conn = get_db_connection()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        cur.execute("""
            SELECT id, date_str as date, nomenclature, amount
            FROM cash_payments WHERE id = %s
        """, (rid,))
        row = cur.fetchone()
        cur.close()
        return_db_connection(conn)
        return dict(row) if row else None
    
    @property
    def counterparties(self) -> List[Dict[str, Any]]:
        cached = self._get_cache("counterparties")
        if cached is not None:
            return cached
        conn = get_db_connection()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        cur.execute("SELECT * FROM counterparties ORDER BY name")
        rows = cur.fetchall()
        cur.close()
        return_db_connection(conn)
        result = [dict(row) for row in rows]
        self._set_cache("counterparties", result)
        return result
    
    def add_counterparty(self, c: Dict[str, Any]):
        self._clear_cache("counterparties")
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO counterparties (id, kind, name, inn, kpp, bank, bik, corr, account, legal_address, phone,
                is_our_company, full_name, inspection_code, oktmo, okato, signatory, sfr_reg_number,
                pfr_reg_self, pfr_reg_employees, pfr_terr_code, pfr_terr_organ, payment_details,
                okpo, okopf, okfs, okved1, okved2, okpo_rosstat)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            ON CONFLICT (id) DO NOTHING
        """, (
            c.get("id"), c.get("kind"), c.get("name"), c.get("inn"), c.get("kpp"),
            c.get("bank"), c.get("bik"), c.get("corr"), c.get("account"),
            c.get("legal_address"), c.get("phone"),
            c.get("is_our_company", False), c.get("full_name"), c.get("inspection_code"),
            c.get("oktmo"), c.get("okato"), c.get("signatory"), c.get("sfr_reg_number"),
            c.get("pfr_reg_self"), c.get("pfr_reg_employees"), c.get("pfr_terr_code"),
            c.get("pfr_terr_organ"), c.get("payment_details"), c.get("okpo"), c.get("okopf"),
            c.get("okfs"), c.get("okved1"), c.get("okved2"), c.get("okpo_rosstat")
        ))
        conn.commit()
        cur.close()
        return_db_connection(conn)
    
    def update_counterparty(self, cid: str, c: Dict[str, Any]):
        self._clear_cache("counterparties")
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute("""
            UPDATE counterparties SET kind=%s, name=%s, inn=%s, kpp=%s, bank=%s, bik=%s,
                   corr=%s, account=%s, legal_address=%s, phone=%s,
                   is_our_company=%s, full_name=%s, inspection_code=%s, oktmo=%s, okato=%s,
                   signatory=%s, sfr_reg_number=%s, pfr_reg_self=%s, pfr_reg_employees=%s,
                   pfr_terr_code=%s, pfr_terr_organ=%s, payment_details=%s,
                   okpo=%s, okopf=%s, okfs=%s, okved1=%s, okved2=%s, okpo_rosstat=%s
            WHERE id = %s
        """, (
            c.get("kind"), c.get("name"), c.get("inn"), c.get("kpp"),
            c.get("bank"), c.get("bik"), c.get("corr"), c.get("account"),
            c.get("legal_address"), c.get("phone"),
            c.get("is_our_company", False), c.get("full_name"), c.get("inspection_code"),
            c.get("oktmo"), c.get("okato"), c.get("signatory"), c.get("sfr_reg_number"),
            c.get("pfr_reg_self"), c.get("pfr_reg_employees"), c.get("pfr_terr_code"),
            c.get("pfr_terr_organ"), c.get("payment_details"), c.get("okpo"), c.get("okopf"),
            c.get("okfs"), c.get("okved1"), c.get("okved2"), c.get("okpo_rosstat"), cid
        ))
        conn.commit()
        cur.close()
        return_db_connection(conn)
    
    def delete_counterparty(self, cid: str):
        self._clear_cache("counterparties")
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute("DELETE FROM counterparties WHERE id = %s", (cid,))
        conn.commit()
        cur.close()
        return_db_connection(conn)
    
    def get_counterparty_by_id(self, cid: str) -> Optional[Dict[str, Any]]:
        conn = get_db_connection()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        cur.execute("SELECT * FROM counterparties WHERE id = %s", (cid,))
        row = cur.fetchone()
        cur.close()
        return_db_connection(conn)
        return dict(row) if row else None
    
    def find_counterparty_by_name(self, name: str) -> Optional[Dict[str, Any]]:
        n = norm_text(name)
        for c in self.counterparties:
            if norm_text(c.get("name", "")) == n:
                return c
        for c in self.counterparties:
            cn = norm_text(c.get("name", ""))
            if cn and (cn in n or n in cn):
                return c
        return None
    
    def get_our_company_card(self) -> Optional[Dict[str, Any]]:
        cid = self.settings.get("our_company_id")
        if cid:
            c = self.get_counterparty_by_id(cid)
            if c:
                return c
        for c in self.counterparties:
            if c.get("is_our_company"):
                return c
        for c in self.counterparties:
            if OUR_COMPANY_DEFAULT_NAME.upper() in norm_text(c.get("name", "")):
                return c
        return None
    
    def auto_upsert_counterparty_from_bank_row(self, r: Dict[str, Any]):
        name = (r.get("counterparty") or "").strip()
        if not name:
            return

        clean, inn2 = split_inn_from_name(name)
        if clean:
            name = clean
            r["counterparty"] = clean
        if inn2 and not (r.get("cp_inn") or "").strip():
            r["cp_inn"] = inn2

        c = self.find_counterparty_by_name(name)
        if not c:
            c = {
                "id": new_id(),
                "kind": "Юридическое лицо",
                "name": name,
                "inn": r.get("cp_inn", "") or "",
                "kpp": r.get("cp_kpp", "") or "",
                "bank": r.get("cp_bank", "") or "",
                "bik": r.get("cp_bik", "") or "",
                "corr": r.get("cp_corr", "") or "",
                "account": r.get("cp_account", "") or "",
                "legal_address": "",
                "phone": "",
            }
            self.add_counterparty(c)
    
    @property
    def acts(self) -> List[Dict[str, Any]]:
        conn = get_db_connection()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        cur.execute("SELECT * FROM acts ORDER BY created_at DESC")
        rows = cur.fetchall()
        cur.close()
        return_db_connection(conn)
        result = []
        for row in rows:
            d = dict(row)
            if d.get("lines_json"):
                try:
                    d["lines"] = json.loads(d["lines_json"])
                except:
                    d["lines"] = []
            else:
                d["lines"] = []
            if d.get("executor_json"):
                try:
                    d["executor"] = json.loads(d["executor_json"])
                except:
                    d["executor"] = {}
            else:
                d["executor"] = {}
            if d.get("customer_json"):
                try:
                    d["customer"] = json.loads(d["customer_json"])
                except:
                    d["customer"] = {}
            else:
                d["customer"] = {}
            result.append(d)
        return result
    
    def add_act(self, act: Dict[str, Any]):
        conn = get_db_connection()
        cur = conn.cursor()
        lines_json = json.dumps(act.get("lines", []), ensure_ascii=False)
        executor_json = json.dumps(act.get("executor", {}), ensure_ascii=False)
        customer_json = json.dumps(act.get("customer", {}), ensure_ascii=False)
        cur.execute("""
            INSERT INTO acts (id, doc_no, doc_date, direction, executor_json, customer_json, basis, vat_mode, lines_json)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
        """, (
            act.get("id"), act.get("doc_no"), act.get("doc_date"),
            act.get("direction", "provide"), executor_json, customer_json,
            act.get("basis"), act.get("vat_mode"), lines_json
        ))
        conn.commit()
        cur.close()
        return_db_connection(conn)
    
    def update_act(self, aid: str, act: Dict[str, Any]):
        conn = get_db_connection()
        cur = conn.cursor()
        lines_json = json.dumps(act.get("lines", []), ensure_ascii=False)
        executor_json = json.dumps(act.get("executor", {}), ensure_ascii=False)
        customer_json = json.dumps(act.get("customer", {}), ensure_ascii=False)
        cur.execute("""
            UPDATE acts SET doc_no=%s, doc_date=%s, direction=%s, executor_json=%s, customer_json=%s,
                   basis=%s, vat_mode=%s, lines_json=%s
            WHERE id = %s
        """, (
            act.get("doc_no"), act.get("doc_date"),
            act.get("direction", "provide"), executor_json, customer_json,
            act.get("basis"), act.get("vat_mode"), lines_json, aid
        ))
        conn.commit()
        cur.close()
        return_db_connection(conn)
    
    def delete_act(self, aid: str):
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute("DELETE FROM acts WHERE id = %s", (aid,))
        conn.commit()
        cur.close()
        return_db_connection(conn)
    
    def get_act_by_id(self, aid: str) -> Optional[Dict[str, Any]]:
        conn = get_db_connection()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        cur.execute("SELECT * FROM acts WHERE id = %s", (aid,))
        row = cur.fetchone()
        cur.close()
        return_db_connection(conn)
        if not row:
            return None
        d = dict(row)
        if d.get("lines_json"):
            try:
                d["lines"] = json.loads(d["lines_json"])
            except:
                d["lines"] = []
        else:
            d["lines"] = []
        if d.get("executor_json"):
            try:
                d["executor"] = json.loads(d["executor_json"])
            except:
                d["executor"] = {}
        else:
            d["executor"] = {}
        if d.get("customer_json"):
            try:
                d["customer"] = json.loads(d["customer_json"])
            except:
                d["customer"] = {}
        else:
            d["customer"] = {}
        return d
    
    @property
    def payment_orders(self) -> List[Dict[str, Any]]:
        conn = get_db_connection()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        cur.execute("SELECT * FROM payment_orders ORDER BY created_at DESC")
        rows = cur.fetchall()
        cur.close()
        return_db_connection(conn)
        result = []
        for row in rows:
            d = dict(row)
            d["date"] = d.pop("date_str", "")
            if d.get("payer_json"):
                try:
                    d["payer"] = json.loads(d["payer_json"])
                except:
                    d["payer"] = {}
            if d.get("receiver_json"):
                try:
                    d["receiver"] = json.loads(d["receiver_json"])
                except:
                    d["receiver"] = {}
            result.append(d)
        return result
    
    def add_payment_order(self, po: Dict[str, Any]):
        conn = get_db_connection()
        cur = conn.cursor()
        payer_json = json.dumps(po.get("payer", {}), ensure_ascii=False)
        receiver_json = json.dumps(po.get("receiver", {}), ensure_ascii=False)
        cur.execute("""
            INSERT INTO payment_orders (id, number, date_str, amount, amount_words, payer_json, receiver_json, purpose, pay_type, vid_op, ocher)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """, (
            po.get("id"), po.get("number"), po.get("date"),
            str(decimal_from_str(po.get("amount") or 0)), po.get("amount_words"),
            payer_json, receiver_json, po.get("purpose"),
            po.get("pay_type"), po.get("vid_op"), po.get("ocher")
        ))
        conn.commit()
        cur.close()
        return_db_connection(conn)
    
    def delete_payment_order(self, poid: str):
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute("DELETE FROM payment_orders WHERE id = %s", (poid,))
        conn.commit()
        cur.close()
        return_db_connection(conn)
    
    @property
    def employees(self) -> List[Dict[str, Any]]:
        conn = get_db_connection()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        cur.execute("SELECT * FROM employees ORDER BY name")
        rows = cur.fetchall()
        cur.close()
        return_db_connection(conn)
        return [dict(row) for row in rows]
    
    def add_employee(self, emp: Dict[str, Any]):
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO employees (id, name, inn, passport, passport_issued, bank, bik, corr, account, salary, advance, main_part)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """, (
            emp.get("id"), emp.get("name"), emp.get("inn"), emp.get("passport"), emp.get("passport_issued"),
            emp.get("bank"), emp.get("bik"), emp.get("corr"), emp.get("account"),
            str(decimal_from_str(emp.get("salary") or 0)),
            str(decimal_from_str(emp.get("advance") or 0)),
            str(decimal_from_str(emp.get("main") or 0))
        ))
        conn.commit()
        cur.close()
        return_db_connection(conn)
    
    def update_employee(self, eid: str, emp: Dict[str, Any]):
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute("""
            UPDATE employees SET name=%s, inn=%s, passport=%s, passport_issued=%s, bank=%s, bik=%s, corr=%s, account=%s, salary=%s, advance=%s, main_part=%s WHERE id = %s
        """, (
            emp.get("name"), emp.get("inn"), emp.get("passport"), emp.get("passport_issued"),
            emp.get("bank"), emp.get("bik"), emp.get("corr"), emp.get("account"),
            str(decimal_from_str(emp.get("salary") or 0)),
            str(decimal_from_str(emp.get("advance") or 0)),
            str(decimal_from_str(emp.get("main") or 0)), eid
        ))
        conn.commit()
        cur.close()
        return_db_connection(conn)
    
    def get_employee_by_id(self, eid: str) -> Optional[Dict[str, Any]]:
        conn = get_db_connection()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        cur.execute("SELECT * FROM employees WHERE id = %s", (eid,))
        row = cur.fetchone()
        cur.close()
        return_db_connection(conn)
        return dict(row) if row else None
    
    def delete_employee(self, eid: str):
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute("DELETE FROM employees WHERE id = %s", (eid,))
        conn.commit()
        cur.close()
        return_db_connection(conn)
    
    @property
    def salary_payments(self) -> List[Dict[str, Any]]:
        conn = get_db_connection()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        cur.execute("SELECT * FROM salary_payments ORDER BY created_at DESC")
        rows = cur.fetchall()
        cur.close()
        return_db_connection(conn)
        result = []
        for row in rows:
            d = dict(row)
            if d.get("created_at"):
                d["created_at_str"] = d["created_at"].strftime("%d.%m.%Y %H:%M:%S")
            else:
                d["created_at_str"] = ""
            result.append(d)
        return result
    
    def add_salary_payment(self, sp: Dict[str, Any]):
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO salary_payments (id, employee_id, month, pay_type, amount, payment_order_id)
            VALUES (%s, %s, %s, %s, %s, %s)
        """, (
            sp.get("id"), sp.get("employee_id"), sp.get("month"), sp.get("type"),
            str(decimal_from_str(sp.get("amount") or 0)), sp.get("payment_order_id")
        ))
        conn.commit()
        cur.close()
        return_db_connection(conn)
    
    def delete_salary_payment(self, sid: str):
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute("DELETE FROM salary_payments WHERE id = %s", (sid,))
        conn.commit()
        cur.close()
        return_db_connection(conn)
    
    def add_payment_order_with_source(self, po: Dict[str, Any]):
        conn = get_db_connection()
        cur = conn.cursor()
        payer_json = json.dumps(po.get("payer", {}), ensure_ascii=False)
        receiver_json = json.dumps(po.get("receiver", {}), ensure_ascii=False)
        cur.execute("""
            INSERT INTO payment_orders (id, number, date_str, amount, amount_words, payer_json, receiver_json, purpose, pay_type, vid_op, ocher, source)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """, (
            po.get("id"), po.get("number"), po.get("date"),
            str(decimal_from_str(po.get("amount") or 0)), po.get("amount_words"),
            payer_json, receiver_json, po.get("purpose"),
            po.get("pay_type"), po.get("vid_op"), po.get("ocher"), po.get("source")
        ))
        conn.commit()
        cur.close()
        return_db_connection(conn)
    
    @property
    def upd_rows(self) -> List[Dict[str, Any]]:
        conn = get_db_connection()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        cur.execute("SELECT * FROM upd_rows ORDER BY created_at DESC")
        rows = cur.fetchall()
        cur.close()
        return_db_connection(conn)
        return [dict(row) for row in rows]
    
    def add_upd_row(self, r: Dict[str, Any]):
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO upd_rows (id, doc_no, doc_date, counterparty, inn, amount, vat, description, source_file)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
        """, (
            r.get("id"), r.get("doc_no"), r.get("doc_date"), r.get("counterparty"),
            r.get("inn"), str(decimal_from_str(r.get("amount") or 0)), str(decimal_from_str(r.get("vat") or 0)),
            r.get("description"), r.get("source_file")
        ))
        conn.commit()
        cur.close()
        return_db_connection(conn)
    
    def delete_upd_row(self, rid: str):
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute("DELETE FROM upd_rows WHERE id = %s", (rid,))
        conn.commit()
        cur.close()
        return_db_connection(conn)
    
    @property
    def real_rows(self) -> List[Dict[str, Any]]:
        conn = get_db_connection()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        cur.execute("SELECT * FROM real_rows ORDER BY created_at DESC")
        rows = cur.fetchall()
        cur.close()
        return_db_connection(conn)
        return [dict(row) for row in rows]
    
    def add_real_row(self, r: Dict[str, Any]):
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO real_rows (id, doc_no, doc_date, counterparty, amount, description)
            VALUES (%s, %s, %s, %s, %s, %s)
        """, (
            r.get("id"), r.get("doc_no"), r.get("doc_date"), r.get("counterparty"),
            str(decimal_from_str(r.get("amount") or 0)), r.get("description")
        ))
        conn.commit()
        cur.close()
        return_db_connection(conn)
    
    def delete_real_row(self, rid: str):
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute("DELETE FROM real_rows WHERE id = %s", (rid,))
        conn.commit()
        cur.close()
        return_db_connection(conn)
    
    @property
    def basis_history(self) -> List[str]:
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute("SELECT basis FROM basis_history ORDER BY id DESC")
        rows = cur.fetchall()
        cur.close()
        return_db_connection(conn)
        return [row[0] for row in rows]
    
    def add_basis(self, basis: str):
        if not basis.strip():
            return
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO basis_history (basis) VALUES (%s)
            ON CONFLICT (basis) DO NOTHING
        """, (basis,))
        conn.commit()
        cur.close()
        return_db_connection(conn)
    
    def recalc_bank_categories(self):
        cp_map = self.cp_category_map
        user_map = self.user_category_map
        
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute("SELECT id, counterparty, purpose FROM bank_rows")
        rows = cur.fetchall()
        
        for rid, counterparty, purpose in rows:
            category = detect_category(counterparty or "", purpose or "", cp_map, user_map)
            cur.execute("UPDATE bank_rows SET category = %s WHERE id = %s", (category, rid))
        
        conn.commit()
        cur.close()
        return_db_connection(conn)
    
    def sanitize_names_and_inn(self):
        conn = get_db_connection()
        cur = conn.cursor()
        
        cur.execute("SELECT id, name, inn FROM counterparties")
        for cid, name, inn in cur.fetchall():
            clean, inn2 = split_inn_from_name(name or "")
            if clean and clean != name:
                cur.execute("UPDATE counterparties SET name = %s WHERE id = %s", (clean, cid))
            if inn2 and not (inn or "").strip():
                cur.execute("UPDATE counterparties SET inn = %s WHERE id = %s", (inn2, cid))
        
        cur.execute("SELECT id, counterparty, cp_inn FROM bank_rows")
        for rid, counterparty, cp_inn in cur.fetchall():
            clean, inn2 = split_inn_from_name(counterparty or "")
            if clean and clean != counterparty:
                cur.execute("UPDATE bank_rows SET counterparty = %s WHERE id = %s", (clean, rid))
            if inn2 and not (cp_inn or "").strip():
                cur.execute("UPDATE bank_rows SET cp_inn = %s WHERE id = %s", (inn2, rid))
        
        conn.commit()
        cur.close()
        return_db_connection(conn)


STATE = AccountingStateDB()


def is_fns_treasury(counterparty: str) -> bool:
    """Check if counterparty is FNS Treasury (case-insensitive, ignores brackets etc)"""
    cp_clean = re.sub(r'[^\w\s]', '', counterparty.lower())
    cp_clean = ' '.join(cp_clean.split())
    return "казначейство" in cp_clean and "фнс" in cp_clean


def detect_tax_type(purpose: str) -> str:
    """Detect specific tax type from payment purpose"""
    pur = purpose.upper()
    if "УСН" in pur or "УПРОЩЕН" in pur:
        return "Налоги - УСН"
    if "НДС" in pur:
        return "Налоги - НДС"
    if "НДФЛ" in pur:
        return "Налоги - НДФЛ"
    if "СТРАХОВ" in pur:
        return "Налоги - Страховые"
    if "ЕНП" in pur:
        return "Налоги - ЕНП"
    return "Налоги"


def detect_category(counterparty: str, purpose: str, cp_map: Dict[str, str], user_map: Dict[str, List[str]]) -> str:
    cp = norm_text(counterparty)
    pur = norm_text(purpose)

    if user_map:
        keys = sorted(user_map.keys(), key=len, reverse=True)
        for k in keys:
            if k and (k in cp):
                cats = user_map.get(k, [])
                if len(cats) == 1:
                    return norm_spaces(cats[0].strip() or "Прочее")
                elif len(cats) > 1:
                    return "СПОРНАЯ"

    if cp_map:
        keys = sorted(cp_map.keys(), key=len, reverse=True)
        for k in keys:
            if k and (k in cp):
                return norm_spaces((cp_map.get(k) or "").strip() or "Прочее")

    if is_fns_treasury(counterparty):
        return detect_tax_type(purpose)

    if "ИНТЕРНЕТ РЕШЕНИЯ" in cp or "ИНТЕРНЕТРЕШЕНИЯ" in cp.replace(" ", ""):
        if "0216761417" in purpose:
            return "Озон производство"
        return "Озон"

    FIXED_MAP = {
        "ООО РВБ": "РВБ",
        "ООО \"РВБ\"": "РВБ",
        "ФНС": "Налоги",
        "СКБ КОНТУР": "Консультационные услуги",
        "КОНТУР": "Консультационные услуги",
    }
    for k, v in FIXED_MAP.items():
        if norm_text(k) in cp:
            return v

    if "КОМИСС" in pur or "ОБСЛУЖ" in pur or "ТАРИФ" in pur:
        return "Комиссия банка"
    if "АРЕНД" in pur:
        return "Аренда"
    if "НАЛОГ" in pur or "ЕНП" in pur or "КБК" in pur:
        return "Налоги"
    if "ЗАРПЛАТ" in pur or "ЗАРАБОТН" in pur:
        return "Зарплата"
    if "ИНТЕРНЕТ" in pur or "СВЯЗ" in pur:
        return "Связь, интернет"

    return "Прочее"


def _get_any(doc: Dict[str, str], keys: List[str]) -> str:
    for k in keys:
        v = (doc.get(k) or "").strip()
        if v:
            return v
    return ""


def extract_counterparty_details(doc: Dict[str, str], side: str) -> Dict[str, str]:
    if side == "PAYER":
        name = _get_any(doc, ["Плательщик", "Плательщик1"])
        inn = _get_any(doc, ["ПлательщикИНН", "ПлательщикИНН1"])
        kpp = _get_any(doc, ["ПлательщикКПП", "ПлательщикКПП1"])
        acc = _get_any(doc, ["ПлательщикСчет", "ПлательщикРасчСчет", "ПлательщикСчет1"])
        bank = _get_any(doc, ["ПлательщикБанк1", "ПлательщикБанк", "ПлательщикБанк2"])
        bik = _get_any(doc, ["ПлательщикБИК", "ПлательщикБИК1"])
        corr = _get_any(doc, ["ПлательщикКорСчет", "ПлательщикКорСчет1", "ПлательщикКС"])
    else:
        name = _get_any(doc, ["Получатель", "Получатель1"])
        inn = _get_any(doc, ["ПолучательИНН", "ПолучательИНН1"])
        kpp = _get_any(doc, ["ПолучательКПП", "ПолучательКПП1"])
        acc = _get_any(doc, ["ПолучательСчет", "ПолучательРасчСчет", "ПолучательСчет1"])
        bank = _get_any(doc, ["ПолучательБанк1", "ПолучательБанк", "ПолучательБанк2"])
        bik = _get_any(doc, ["ПолучательБИК", "ПолучательБИК1"])
        corr = _get_any(doc, ["ПолучательКорСчет", "ПолучательКорСчет1", "ПолучательКС"])

    return {"name": name, "inn": inn, "kpp": kpp, "account": acc, "bank": bank, "bik": bik, "corr": corr}


def parse_client_bank_file(path: str) -> List[Dict[str, Any]]:
    with open(path, "r", encoding="cp1251", errors="replace") as f:
        lines = [line.rstrip("\n\r") for line in f]

    main_account = ""
    documents: List[Dict[str, str]] = []
    current_doc: Optional[Dict[str, str]] = None

    for raw in lines:
        line = raw.strip()
        if not line:
            continue

        if line.startswith("РасчСчет=") and not main_account:
            main_account = line.split("=", 1)[1].strip()

        if line.startswith("СекцияДокумент="):
            section_type = line.split("=", 1)[1].strip()
            current_doc = {"_SectionType": section_type}
            continue

        if line == "КонецДокумента":
            if current_doc is not None:
                documents.append(current_doc)
                current_doc = None
            continue

        if current_doc is not None and "=" in line:
            key, val = line.split("=", 1)
            current_doc[key.strip()] = val.strip()

    rows: List[Dict[str, Any]] = []
    for doc in documents:
        row = convert_document_to_row(doc, main_account)
        row["id"] = new_id()
        rows.append(row)

    return rows


def convert_document_to_row(doc: Dict[str, str], main_account: str) -> Dict[str, Any]:
    sum_str = (doc.get("Сумма", "0") or "0").replace(" ", "").replace(",", ".")
    amount = decimal_from_str(sum_str)

    payer_acc = doc.get("ПлательщикСчет", "") or doc.get("ПлательщикРасчСчет", "") or ""
    receiver_acc = doc.get("ПолучательСчет", "") or doc.get("ПолучательРасчСчет", "") or ""

    incoming = Decimal("0")
    outgoing = Decimal("0")

    dt = doc.get("Дата", "") or ""
    purpose = doc.get("НазначениеПлатежа", "") or ""
    doctype = doc.get("_SectionType", "") or doc.get("СекцияДокумент", "") or ""
    counterparty = doc.get("Получатель") or doc.get("Плательщик") or ""

    cp_details = {"name": "", "inn": "", "kpp": "", "account": "", "bank": "", "bik": "", "corr": ""}

    if receiver_acc.strip() and receiver_acc.strip() == (main_account or "").strip():
        incoming = amount
        dt = doc.get("ДатаПоступило") or doc.get("Дата") or ""
        cp_details = extract_counterparty_details(doc, "PAYER")
        counterparty = cp_details.get("name") or (doc.get("Плательщик", "") or doc.get("Плательщик1", "") or counterparty)

    elif payer_acc.strip() and payer_acc.strip() == (main_account or "").strip():
        outgoing = amount
        dt = doc.get("ДатаСписано") or doc.get("Дата") or ""
        cp_details = extract_counterparty_details(doc, "RECEIVER")
        counterparty = cp_details.get("name") or (doc.get("Получатель", "") or doc.get("Получатель1", "") or counterparty)

    clean_name, inn2 = split_inn_from_name(counterparty)
    if inn2 and not (cp_details.get("inn") or "").strip():
        cp_details["inn"] = inn2
    counterparty = clean_name or counterparty

    return {
        "date": dt,
        "month": month_from_date_str(dt),
        "incoming": incoming,
        "outgoing": outgoing,
        "purpose": purpose,
        "counterparty": counterparty,
        "doctype": doctype,
        "skip_outgoing": False,
        "category": "",
        "cp_inn": cp_details.get("inn", ""),
        "cp_kpp": cp_details.get("kpp", ""),
        "cp_account": cp_details.get("account", ""),
        "cp_bank": cp_details.get("bank", ""),
        "cp_bik": cp_details.get("bik", ""),
        "cp_corr": cp_details.get("corr", ""),
    }


def bank_row_fingerprint(r: Dict[str, Any]) -> str:
    dt = (r.get("date") or "").strip()
    inc = Decimal(str(r.get("incoming") or 0)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    out = Decimal(str(r.get("outgoing") or 0)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    amt_part = f"IN:{inc}" if inc > 0 else f"OUT:{out}"
    cp = norm_text(r.get("counterparty") or "")
    pur = norm_text(r.get("purpose") or "")
    doc = norm_text(r.get("doctype") or "")
    return "|".join([dt, amt_part, cp, pur, doc])


BASE_CSS = r"""
:root{
  --bg:#FFF7D6;
  --panel:#FFFFFF;
  --line:#D7D7D7;
  --text:#2B2B2B;
  --muted:#6B6B6B;
  --accent:#FFC94A;
  --accent2:#FFB300;
  --green:#0F8F4A;
  --red:#B73A2E;
  --shadow: 0 2px 10px rgba(0,0,0,.06);
  --radius:14px;
  --font: ui-sans-serif, system-ui, -apple-system, "Segoe UI", Roboto, Arial, "Noto Sans", "Liberation Sans", sans-serif;
}
*{box-sizing:border-box}
body{
  margin:0;
  font-family:var(--font);
  color:var(--text);
  background:linear-gradient(180deg,#FFF7D6 0%, #FFF2B8 100%);
}
a{color:inherit;text-decoration:none}
.header{
  position:sticky; top:0; z-index:10;
  display:flex; align-items:center; justify-content:space-between;
  padding:10px 16px;
  background:rgba(255,255,255,.65);
  backdrop-filter: blur(10px);
  border-bottom:1px solid var(--line);
}
.brand{display:flex; align-items:baseline; gap:10px;}
.brand .title{font-weight:800; letter-spacing:.2px;}
.badge{
  font-size:12px; padding:3px 8px; border-radius:999px;
  border:1px solid var(--line); background:#d4edda; color:#155724;
}
.search{flex:1; max-width:680px; margin:0 16px;}
.search input{
  width:100%;
  padding:10px 12px;
  border:1px solid var(--line);
  border-radius:12px;
  background:#fff;
}
.hactions{display:flex; align-items:center; gap:10px; font-size:13px; color:var(--muted);}
.btn{
  display:inline-flex; align-items:center; justify-content:center;
  gap:8px;
  padding:9px 12px;
  border-radius:12px;
  border:1px solid var(--line);
  background:#fff;
  cursor:pointer;
}
.btn.primary{background:var(--accent); border-color:#F2B733; font-weight:700;}
.btn.primary:hover{background:var(--accent2)}
.btn.danger{background:#fff;border-color:#E7B0A9;color:#8F1D14;}
.btn.small{padding:6px 10px; border-radius:10px; font-size:13px}
.layout{display:grid; grid-template-columns: 280px 1fr; gap:14px; padding:14px;}
.sidebar{
  position:sticky; top:64px; align-self:start;
  background:rgba(255,255,255,.75);
  border:1px solid var(--line);
  border-radius:var(--radius);
  box-shadow:var(--shadow);
  padding:10px;
}
.navgroup{margin:8px 0 14px}
.navtitle{font-size:12px;color:var(--muted);font-weight:800;text-transform:uppercase;letter-spacing:.08em;padding:6px 10px}
.navitem{display:flex;align-items:center;justify-content:space-between;padding:10px 10px;border-radius:12px;border:1px solid transparent;margin:4px 0;}
.navitem:hover{background:#fff}
.navitem.active{background:#fff;border-color:#F2B733;box-shadow:0 0 0 3px rgba(255,201,74,.25);}
.main{
  background:rgba(255,255,255,.72);
  border:1px solid var(--line);
  border-radius:var(--radius);
  box-shadow:var(--shadow);
  padding:14px;
  min-height: calc(100vh - 100px);
}
.breadcrumbs{font-size:13px;color:var(--muted);margin-bottom:8px}
.pagehead{display:flex;align-items:flex-end;justify-content:space-between;gap:12px;margin-bottom:10px;}
.pagehead h1{margin:0;font-size:18px;letter-spacing:.2px;}
.toolbar{display:flex;gap:8px;flex-wrap:wrap;align-items:center;}
.panel{background:#fff;border:1px solid var(--line);border-radius:14px;padding:12px;}
.table{width:100%;border-collapse:separate;border-spacing:0;overflow:hidden;border-radius:14px;border:1px solid var(--line);background:#fff;}
.table th,.table td{padding:9px 10px;border-bottom:1px solid #EFEFEF;vertical-align:top;font-size:13px;}
.table th{background:#FFF9E6;font-weight:800;color:#3A3A3A;position:sticky;top:0;z-index:1;}
.table tr:hover td{background:#FFFDF6}
.right{text-align:right}
.center{text-align:center}
.muted{color:var(--muted)}
.badge2{display:inline-flex;align-items:center;padding:3px 8px;border-radius:999px;border:1px solid var(--line);background:#fff;font-size:12px;}
.income{color:var(--green); font-weight:700}
.outgoing{color:var(--red); font-weight:700}
.smallnote{font-size:12px;color:var(--muted)}
.flash{background:#fff;border:1px solid #F2B733;border-left:6px solid #F2B733;padding:10px 12px;border-radius:14px;margin-bottom:12px;}
.hr{height:1px;background:#EFEFEF;margin:12px 0}
.file-hidden{position:absolute; left:-9999px; width:1px; height:1px; opacity:0;}
.purpose-wrap{white-space:pre-line;}
.purpose-clamp{
  display:-webkit-box;
  -webkit-box-orient: vertical;
  -webkit-line-clamp: 3;
  overflow:hidden;
}
.purpose-expanded{
  display:block;
}
.purpose-more{
  margin-top:6px;
  font-size:12px;
  color:#6B6B6B;
  cursor:pointer;
  text-decoration: underline;
}
"""


def flash_box(text: str) -> str:
    if not text:
        return ""
    return f'<div class="flash">{h(text)}</div>'


def render_layout(path: str, title: str, crumbs: str, body_html: str, flash: str = "") -> str:
    nav_html = []
    for group, items in NAV:
        nav_html.append(f'<div class="navgroup"><div class="navtitle">{h(group)}</div>')
        for name, url in items:
            active = "active" if url == path else ""
            nav_html.append(f'<a class="navitem {active}" href="{h(url)}"><span>{h(name)}</span></a>')
        nav_html.append("</div>")
    nav_html = "\n".join(nav_html)

    last_saved = STATE.last_saved_at or "—"

    return f"""<!doctype html>
<html lang="ru">
<head>
  <meta charset="utf-8"/>
  <meta name="viewport" content="width=device-width, initial-scale=1"/>
  <title>{h(title)} — {OUR_COMPANY_DEFAULT_NAME}</title>
  <style>{BASE_CSS}</style>
</head>
<body>
  <div class="header">
    <div class="brand">
      <div class="title">Бухгалтерия — {h(OUR_COMPANY_DEFAULT_NAME)}</div>
      <div class="badge">ONLINE</div>
    </div>
    <div class="search">
      <form method="GET" action="/search">
        <input name="q" placeholder="Поиск: контрагент, документ, назначение, номер…" />
      </form>
    </div>
    <div class="hactions">
      <span class="muted">последнее сохранение: {h(last_saved)}</span>
      <form method="POST" action="/action/save" style="margin:0">
        <button class="btn primary" type="submit">💾 Сохранить</button>
      </form>
    </div>
  </div>

  <div class="layout">
    <aside class="sidebar">
      {nav_html}
      <div class="hr"></div>
      <div class="smallnote">Данные хранятся в базе данных PostgreSQL.</div>
    </aside>

    <main class="main">
      <div class="breadcrumbs">{h(crumbs)}</div>
      {flash}
      {body_html}
    </main>
  </div>
</body>
</html>"""


def page_bank(path: str, qd: Dict[str, str] = None, flash: str = "") -> str:
    if qd is None:
        qd = {}
    sort_mode = qd.get("sort", "upload_desc")
    date_from = qd.get("from", "").strip()
    date_to = qd.get("to", "").strip()
    search_q = qd.get("q", "").strip()
    try:
        current_page = max(1, int(qd.get("page", "1")))
    except:
        current_page = 1
    per_page = 100
    
    bank_rows = list(STATE.bank_rows)
    
    if date_from or date_to:
        from_dt = parse_date_ddmmyyyy(date_from)
        to_dt = parse_date_ddmmyyyy(date_to)
        filtered = []
        for r in bank_rows:
            rd = parse_date_ddmmyyyy(r.get("date", "") or "")
            if rd:
                if from_dt and rd < from_dt:
                    continue
                if to_dt and rd > to_dt:
                    continue
                filtered.append(r)
            else:
                filtered.append(r)
        bank_rows = filtered
    
    if search_q:
        sq = search_q.upper()
        filtered = []
        for r in bank_rows:
            searchable = " ".join([
                r.get("counterparty", "") or "",
                r.get("purpose", "") or "",
                r.get("category", "") or "",
                str(r.get("incoming") or ""),
                str(r.get("outgoing") or ""),
            ]).upper()
            if sq in searchable:
                filtered.append(r)
        bank_rows = filtered
    
    total_ops = len(bank_rows)
    
    def date_sort_key(r):
        d = parse_date_ddmmyyyy(r.get("date", "") or "")
        return d if d else date(1900, 1, 1)
    
    if sort_mode == "date_asc":
        bank_rows.sort(key=date_sort_key)
    elif sort_mode == "date_desc":
        bank_rows.sort(key=date_sort_key, reverse=True)
    elif sort_mode == "amount_desc":
        bank_rows.sort(key=lambda r: Decimal(str(r.get("incoming") or 0)) + Decimal(str(r.get("outgoing") or 0)), reverse=True)
    elif sort_mode == "amount_asc":
        bank_rows.sort(key=lambda r: Decimal(str(r.get("incoming") or 0)) + Decimal(str(r.get("outgoing") or 0)))
    
    cp_map = STATE.cp_category_map
    cp_map_source = STATE.cp_map_source
    user_map = STATE.user_category_map
    
    map_info = "—"
    if cp_map_source and cp_map:
        map_info = f"{os.path.basename(cp_map_source)} ({fmt_int(len(cp_map))} строк)"
    
    email_status = "ожидание (12:00)"
    if EMAIL_LAST_DOWNLOAD.get("time"):
        t = EMAIL_LAST_DOWNLOAD["time"].strftime("%d.%m %H:%M")
        s = EMAIL_LAST_DOWNLOAD.get("status", "")
        a = EMAIL_LAST_DOWNLOAD.get("added", 0)
        email_status = f"{t} — {s} (+{a})"

    js = r"""
<script>
function pickAndSubmit(inputId, formId){
  const inp = document.getElementById(inputId);
  const form = document.getElementById(formId);
  inp.onchange = ()=>{ if(inp.files && inp.files.length){ form.submit(); } };
  inp.click();
}
function submitOnCheck(formId){
  document.getElementById(formId).submit();
}
function togglePurpose(id){
  const box = document.getElementById("pur_"+id);
  const link = document.getElementById("pur_lnk_"+id);
  if(!box || !link) return;
  const expanded = box.classList.contains("purpose-expanded");
  if(expanded){
    box.classList.remove("purpose-expanded");
    box.classList.add("purpose-clamp");
    link.textContent = "Показать полностью";
  }else{
    box.classList.remove("purpose-clamp");
    box.classList.add("purpose-expanded");
    link.textContent = "Свернуть";
  }
}
</script>
"""

    sort_options = [
        ("upload_desc", "По загрузке (новые сверху)"),
        ("date_desc", "По дате операции (новые сверху)"),
        ("date_asc", "По дате операции (старые сверху)"),
        ("amount_desc", "По сумме (большие сверху)"),
        ("amount_asc", "По сумме (малые сверху)"),
    ]
    sort_select = "".join(
        f'<option value="{v}" {"selected" if v == sort_mode else ""}>{h(lbl)}</option>'
        for v, lbl in sort_options
    )

    toolbar = f"""
    <div class="pagehead">
      <div>
        <h1>Банк-клиент</h1>
        <div class="smallnote">Загрузка выписок 1CClientBankExchange (.txt), авто-статьи, «не учитывать» для списаний.</div>
      </div>
      <div class="toolbar">

        <form id="bankUploadForm" method="POST" action="/action/bank/upload" enctype="multipart/form-data" style="margin:0">
          <input class="file-hidden" id="bankFile" type="file" name="file" accept=".txt" required/>
          <button class="btn primary" type="button" onclick="pickAndSubmit('bankFile','bankUploadForm')">Загрузить выписку</button>
        </form>

        <form id="cpmapForm" method="POST" action="/action/bank/cpmap" enctype="multipart/form-data" style="margin:0">
          <input class="file-hidden" id="cpmapFile" type="file" name="file" accept=".xlsx,.csv" required/>
          <button class="btn" type="button" onclick="pickAndSubmit('cpmapFile','cpmapForm')">Справочник контрагент→статья</button>
        </form>

        <form method="POST" action="/action/bank/fetch-email" style="margin:0">
          <button class="btn" type="submit">Скачать с почты</button>
        </form>

        <a class="btn" href="/action/bank/export-csv">Экспорт CSV</a>
      </div>
    </div>

    <div class="panel" style="margin-bottom:12px;display:flex;gap:14px;align-items:center;flex-wrap:wrap">
      <span class="badge2"><b>Операций:</b>&nbsp;{fmt_int(total_ops)}</span>
      <span class="badge2"><b>Справочник:</b>&nbsp;{h(map_info)}</span>
      <span class="badge2"><b>Ручные «подписи»:</b>&nbsp;{fmt_int(len(user_map))}</span>
      <span class="badge2"><b>Авто-загрузка с почты:</b>&nbsp;{h(email_status)}</span>
    </div>

    <div class="panel" style="margin-bottom:12px;display:flex;gap:14px;align-items:center;flex-wrap:wrap">
      <form id="filterForm" method="GET" action="/bank" style="display:flex;gap:10px;align-items:center;flex-wrap:wrap;margin:0">
        <input type="hidden" name="sort" value="{h(sort_mode)}"/>
        <span style="display:flex;align-items:center;gap:4px">
          <b>С:</b>
          <input type="text" name="from" value="{h(date_from)}" placeholder="ДД.ММ.ГГГГ" 
                 style="width:100px;padding:6px 8px;border:1px solid #ccc;border-radius:8px"/>
        </span>
        <span style="display:flex;align-items:center;gap:4px">
          <b>По:</b>
          <input type="text" name="to" value="{h(date_to)}" placeholder="ДД.ММ.ГГГГ" 
                 style="width:100px;padding:6px 8px;border:1px solid #ccc;border-radius:8px"/>
        </span>
        <span style="display:flex;align-items:center;gap:4px">
          <b>Поиск:</b>
          <input type="text" name="q" value="{h(search_q)}" placeholder="контрагент, назначение, сумма..." 
                 style="width:220px;padding:6px 8px;border:1px solid #ccc;border-radius:8px"/>
        </span>
        <button class="btn" type="submit">Применить</button>
        <a class="btn" href="/bank">Сбросить</a>
      </form>
      <span class="badge2" style="margin-left:auto">
        <b>Сортировка:</b>&nbsp;
        <select id="sortSelect" style="padding:4px 8px;border-radius:8px;border:1px solid #ccc">
          {sort_select}
        </select>
      </span>
    </div>
    <script>
      document.getElementById('sortSelect').onchange = function() {{
        document.querySelector('input[name="sort"]').value = this.value;
        document.getElementById('filterForm').submit();
      }};
    </script>
    {js}
    """

    current_params = []
    if sort_mode and sort_mode != "upload_desc":
        current_params.append(f"sort={sort_mode}")
    if date_from:
        current_params.append(f"from={urlencode_component(date_from)}")
    if date_to:
        current_params.append(f"to={urlencode_component(date_to)}")
    if search_q:
        current_params.append(f"q={urlencode_component(search_q)}")
    return_url = "/bank" + ("?" + "&".join(current_params) if current_params else "")

    disputed_rows = [r for r in bank_rows if r.get("category") == "СПОРНАЯ"]
    normal_rows = [r for r in bank_rows if r.get("category") != "СПОРНАЯ"]

    def build_row_html(r, show_disputed_form=False):
        rid = r.get("id", "")
        inc = Decimal(str(r.get("incoming") or 0))
        out = Decimal(str(r.get("outgoing") or 0))
        skip = bool(r.get("skip_outgoing", False))

        skip_cell = '<span class="muted">—</span>'
        if out > 0:
            form_id = f"sk_{rid}"
            checked = "checked" if skip else ""
            skip_cell = f"""
            <form id="{h(form_id)}" method="POST" action="/action/bank/set-skip" style="margin:0">
              <input type="hidden" name="id" value="{h(rid)}"/>
              <input type="hidden" name="value" value="0"/>
              <input type="checkbox" name="value" value="1" {checked}
                     onchange="submitOnCheck('{h(form_id)}')"/>
            </form>
            """

        purpose_full = (r.get("purpose", "") or "")
        purpose_box = f"""
          <div class="purpose-wrap">
            <div id="pur_{h(rid)}" class="purpose-clamp" title="{h(purpose_full)}">{h(purpose_full)}</div>
            <span id="pur_lnk_{h(rid)}" class="purpose-more" onclick="togglePurpose('{h(rid)}')">Показать полностью</span>
          </div>
        """
        
        cat_display = h(r.get("category",""))
        action_cell = f'<a class="btn small" href="/bank/assign?id={h(rid)}&ret={urlencode_component(return_url)}">Подписать</a>'
        
        if show_disputed_form:
            cp_name = (r.get("counterparty") or "").strip()
            cp_key = norm_text(cp_name)
            cp_cats = STATE.get_user_categories(cp_key)
            if cp_cats:
                opts = "".join(f'<option value="{h(c)}">{h(c)}</option>' for c in cp_cats)
                cat_display = '<span style="color:#c00;font-weight:bold">СПОРНАЯ</span>'
                action_cell = f"""
                <form method="POST" action="/action/bank/resolve-disputed" style="margin:0;display:flex;flex-direction:column;gap:4px">
                  <input type="hidden" name="id" value="{h(rid)}"/>
                  <input type="hidden" name="return_url" value="{h(return_url)}"/>
                  <select name="category" style="padding:4px;border-radius:6px;border:1px solid #ccc" required>
                    <option value="">Выберите...</option>
                    {opts}
                  </select>
                  <button class="btn small primary" type="submit">Назначить</button>
                </form>
                """

        cp_name = (r.get("counterparty", "") or "").strip()
        checkbox = ""
        if not show_disputed_form and cp_name:
            checkbox = f'<input type="checkbox" class="bulk-select" data-cp="{h(cp_name)}" />'
        
        return f"""
        <tr>
          <td class="center">{checkbox}</td>
          <td class="center">{skip_cell}</td>
          <td class="center">{h(r.get("date",""))}</td>
          <td class="center">{h(r.get("month",""))}</td>
          <td class="right">{f"<span class='income'>{h(fmt_money(inc))}</span>" if inc else ""}</td>
          <td class="right">{f"<span class='outgoing'>{h(fmt_money(out))}</span>" if out else ""}</td>
          <td>{cat_display}</td>
          <td>{purpose_box}</td>
          <td>{h(r.get("counterparty",""))}</td>
          <td class="center">{h(r.get("doctype",""))}</td>
          <td class="center">
            {action_cell}
          </td>
        </tr>
        """

    total_normal = len(normal_rows)
    total_pages = max(1, (total_normal + per_page - 1) // per_page)
    current_page = min(current_page, total_pages)
    start_idx = (current_page - 1) * per_page
    end_idx = start_idx + per_page
    page_rows = normal_rows[start_idx:end_idx]
    
    rows_html = [build_row_html(r, show_disputed_form=False) for r in page_rows]
    disputed_rows_html = [build_row_html(r, show_disputed_form=True) for r in disputed_rows]

    disputed_section = ""
    if disputed_rows:
        disputed_section = f"""
    <div class="panel" style="margin-bottom:16px;border:2px solid #c00;background:#fff5f5">
      <div style="display:flex;align-items:center;gap:10px;margin-bottom:10px">
        <span style="color:#c00;font-weight:bold;font-size:1.1em">Спорные операции ({len(disputed_rows)})</span>
        <span class="smallnote">У этих контрагентов заведено несколько статей — выберите нужную для каждой операции</span>
      </div>
      <table class="table">
        <thead>
          <tr>
            <th style="width:40px"></th>
            <th style="width:100px" class="center">Не учитывать</th>
            <th style="width:95px" class="center">Дата</th>
            <th style="width:85px" class="center">Месяц</th>
            <th style="width:140px" class="right">Поступление</th>
            <th style="width:140px" class="right">Списание</th>
            <th style="width:190px">Статья</th>
            <th>Назначение</th>
            <th style="width:260px">Контрагент</th>
            <th style="width:140px" class="center">Вид</th>
            <th style="width:150px" class="center">Выбрать статью</th>
          </tr>
        </thead>
        <tbody>
          {''.join(disputed_rows_html)}
        </tbody>
      </table>
    </div>
    """

    used_categories = set()
    for row in STATE.bank_rows:
        cat = (row.get("category") or "").strip()
        if cat and cat not in ("Прочее", "СПОРНАЯ"):
            used_categories.add(cat)
    predefined = ["Прочее", "Налоги", "Налоги - УСН", "Налоги - НДС", "Налоги - НДФЛ", "Налоги - Страховые",
                  "Комиссия банка", "Аренда", "Зарплата", "Связь, интернет", "Озон", "Wildberries"]
    for p in predefined:
        used_categories.add(p)
    categories_sorted = sorted(used_categories)
    bulk_datalist = "\n".join(f'<option value="{h(c)}">' for c in categories_sorted)

    bulk_form = f"""
    <div id="bulkPanel" class="panel" style="margin-bottom:12px;display:none;background:#f0f8ff;border:1px solid #7AA2F7">
      <form method="POST" action="/action/bank/bulk-assign" style="display:flex;gap:12px;align-items:center;flex-wrap:wrap">
        <input type="hidden" name="return_url" value="{h(return_url)}"/>
        <span><b>Выбрано контрагентов:</b> <span id="bulkCount">0</span></span>
        <input type="hidden" name="counterparties" id="bulkCounterparties" value=""/>
        <input type="text" name="category" list="bulk_cat_list" placeholder="Введите статью..."
               style="width:200px;padding:8px;border:1px solid #ccc;border-radius:8px" required/>
        <datalist id="bulk_cat_list">{bulk_datalist}</datalist>
        <button class="btn primary" type="submit">Назначить статью выбранным</button>
        <button class="btn" type="button" onclick="clearBulkSelection()">Отменить выбор</button>
      </form>
    </div>
    <script>
      function updateBulkPanel() {{
        const checked = document.querySelectorAll('.bulk-select:checked');
        const panel = document.getElementById('bulkPanel');
        const countEl = document.getElementById('bulkCount');
        const inputEl = document.getElementById('bulkCounterparties');
        const cpSet = new Set();
        checked.forEach(cb => cpSet.add(cb.dataset.cp));
        countEl.textContent = cpSet.size;
        inputEl.value = Array.from(cpSet).join('|||');
        panel.style.display = cpSet.size > 0 ? 'block' : 'none';
      }}
      function clearBulkSelection() {{
        document.querySelectorAll('.bulk-select:checked').forEach(cb => cb.checked = false);
        updateBulkPanel();
      }}
      document.addEventListener('DOMContentLoaded', function() {{
        document.querySelectorAll('.bulk-select').forEach(cb => cb.addEventListener('change', updateBulkPanel));
      }});
    </script>
    """

    table_html = bulk_form + f"""
    <table class="table">
      <thead>
        <tr>
          <th style="width:40px" class="center"><input type="checkbox" id="selectAll" onclick="toggleSelectAll(this)"/></th>
          <th style="width:100px" class="center">Не учитывать</th>
          <th style="width:95px" class="center">Дата</th>
          <th style="width:85px" class="center">Месяц</th>
          <th style="width:140px" class="right">Поступление</th>
          <th style="width:140px" class="right">Списание</th>
          <th style="width:190px">Статья</th>
          <th>Назначение</th>
          <th style="width:260px">Контрагент</th>
          <th style="width:140px" class="center">Вид</th>
          <th style="width:100px" class="center">Действие</th>
        </tr>
      </thead>
      <tbody>
        {''.join(rows_html) if rows_html else '<tr><td colspan="11" class="muted">Пока нет операций. Загрузите выписку.</td></tr>'}
      </tbody>
    </table>
    <script>
      function toggleSelectAll(el) {{
        document.querySelectorAll('.bulk-select').forEach(cb => {{ cb.checked = el.checked; }});
        updateBulkPanel();
      }}
    </script>
    """
    
    base_params = []
    if sort_mode and sort_mode != "upload_desc":
        base_params.append(f"sort={sort_mode}")
    if date_from:
        base_params.append(f"from={urlencode_component(date_from)}")
    if date_to:
        base_params.append(f"to={urlencode_component(date_to)}")
    if search_q:
        base_params.append(f"q={urlencode_component(search_q)}")
    base_url = "/bank" + ("?" + "&".join(base_params) if base_params else "")
    
    pagination_html = ""
    if total_pages > 1:
        pages = []
        sep = "&" if base_params else "?"
        if current_page > 1:
            pages.append(f'<a class="btn" href="{base_url}{sep}page={current_page - 1}">← Назад</a>')
        for p in range(1, total_pages + 1):
            if p == current_page:
                pages.append(f'<span class="badge2" style="background:#7AA2F7;color:white"><b>{p}</b></span>')
            elif abs(p - current_page) <= 2 or p == 1 or p == total_pages:
                pages.append(f'<a class="btn" href="{base_url}{sep}page={p}">{p}</a>')
            elif abs(p - current_page) == 3:
                pages.append('<span>...</span>')
        if current_page < total_pages:
            pages.append(f'<a class="btn" href="{base_url}{sep}page={current_page + 1}">Вперёд →</a>')
        pagination_html = f'<div style="margin-top:12px;display:flex;gap:8px;align-items:center;flex-wrap:wrap">{" ".join(pages)}</div>'
    
    info_html = f'<div class="smallnote" style="margin-top:8px">Показаны {fmt_int(start_idx + 1)}–{fmt_int(min(end_idx, total_normal))} из {fmt_int(total_normal)} операций (страница {current_page}/{total_pages})</div>'
    table_html = table_html + info_html + pagination_html

    body = toolbar + disputed_section + table_html
    return render_layout(path="/bank", title="Банк-клиент", crumbs="Операции → Банк-клиент", body_html=body, flash=flash_box(flash))


def page_bank_assign(path: str, qd: Dict[str, str], flash: str = "") -> str:
    rid = qd.get("id", "")
    return_url = qd.get("ret", "/bank")
    rr = STATE.get_bank_row(rid)
    if not rr:
        return render_layout("/bank", "Подписать", "Операции → Банк-клиент", "<h1>Строка не найдена</h1>", flash=flash_box(flash))

    cp_name = (rr.get("counterparty") or "").strip()
    cur = (rr.get("category") or "").strip()
    cp_key = norm_text(cp_name)
    
    existing_cats = STATE.get_user_categories(cp_key)
    existing_html = ""
    if existing_cats:
        cats_items = []
        for c in existing_cats:
            cats_items.append(f'''
              <span class="badge2" style="display:inline-flex;align-items:center;gap:6px">
                {h(c)}
                <form method="POST" action="/action/bank/remove-category" style="display:inline;margin:0">
                  <input type="hidden" name="id" value="{h(rid)}"/>
                  <input type="hidden" name="cp_key" value="{h(cp_key)}"/>
                  <input type="hidden" name="category" value="{h(c)}"/>
                  <input type="hidden" name="return_url" value="{h(return_url)}&id={h(rid)}"/>
                  <button type="submit" style="background:none;border:none;cursor:pointer;color:#C00;font-weight:bold;padding:0 4px" title="Удалить статью">×</button>
                </form>
              </span>
            ''')
        existing_html = f"""
        <div style="margin-bottom:12px">
          <div class="smallnote"><b>Заведённые статьи для этого контрагента:</b> (нажмите × чтобы удалить)</div>
          <div style="display:flex;gap:8px;flex-wrap:wrap;margin-top:4px">
            {"".join(cats_items)}
          </div>
        </div>
        """

    used_categories = set()
    for row in STATE.bank_rows:
        cat = (row.get("category") or "").strip()
        if cat and cat not in ("Прочее", "СПОРНАЯ"):
            used_categories.add(cat)
    for cats in STATE.user_category_map.values():
        for cat in cats:
            if cat and cat.strip():
                used_categories.add(cat.strip())
    
    predefined = ["Прочее", "Налоги", "Налоги - УСН", "Налоги - НДС", "Налоги - НДФЛ", "Налоги - Страховые",
                  "Комиссия банка", "Аренда", "Зарплата", "Связь, интернет", "Озон", "Wildberries"]
    for p in predefined:
        used_categories.add(p)
    
    categories_sorted = sorted(used_categories)
    datalist_options = "\n".join(f'<option value="{h(c)}">' for c in categories_sorted)

    body = f"""
    <div class="pagehead">
      <div>
        <h1>Подписать контрагента</h1>
        <div class="smallnote">Можно добавить несколько статей — если их больше одной, операции станут «спорными» до выбора.</div>
      </div>
      <div class="toolbar">
        <a class="btn" href="{h(return_url)}">← Назад</a>
      </div>
    </div>

    <div class="panel">
      <div style="display:grid;grid-template-columns:1fr 1fr;gap:12px">
        <div>
          <div class="smallnote"><b>Контрагент</b></div>
          <div>{h(cp_name)}</div>
        </div>
        <div>
          <div class="smallnote"><b>Текущая статья операции</b></div>
          <div>{h(cur)}</div>
        </div>
      </div>
      {existing_html}
      <div class="hr"></div>
      <form method="POST" action="/action/bank/assign-category">
        <input type="hidden" name="id" value="{h(rid)}"/>
        <input type="hidden" name="return_url" value="{h(return_url)}"/>
        <div>
          <div class="smallnote"><b>Добавить статью</b> (введите или выберите из списка)</div>
          <input name="category" value="" list="category_list" placeholder="Введите новую статью..."
                 style="width:100%;padding:10px;border:1px solid #D7D7D7;border-radius:12px" required />
          <datalist id="category_list">
            {datalist_options}
          </datalist>
        </div>
        <div style="margin-top:12px;display:flex;gap:10px">
          <button class="btn primary" type="submit">Добавить статью</button>
          <a class="btn" href="{h(return_url)}">Отмена</a>
        </div>
      </form>
    </div>
    """
    return render_layout(path="/bank", title="Подписать контрагента", crumbs="Операции → Банк-клиент → Подписать", body_html=body, flash=flash_box(flash))


def page_cash(path: str, flash: str = "") -> str:
    cash_rows = STATE.cash_rows
    total_ops = len(cash_rows)
    total_sum = sum(Decimal(str(r.get("amount") or 0)) for r in cash_rows)

    body = f"""
    <div class="pagehead">
      <div>
        <h1>Касса</h1>
        <div class="smallnote">Учёт наличных платежей. Суммы учитываются при расчёте УСН как поступления.</div>
      </div>
      <div class="toolbar">
        <a class="btn primary" href="/cash/new">+ Добавить</a>
      </div>
    </div>

    <div class="panel" style="margin-bottom:12px;display:flex;gap:14px;align-items:center;flex-wrap:wrap">
      <span class="badge2"><b>Операций:</b>&nbsp;{fmt_int(total_ops)}</span>
      <span class="badge2"><b>Сумма:</b>&nbsp;{h(fmt_money(total_sum))}</span>
    </div>
    """

    rows_html = []
    for r in cash_rows:
        rid = r.get("id", "")
        amt = Decimal(str(r.get("amount") or 0))
        rows_html.append(f"""
        <tr>
          <td class="center">{h(r.get("date",""))}</td>
          <td>{h(r.get("nomenclature",""))}</td>
          <td class="right"><span class="income">{h(fmt_money(amt))}</span></td>
          <td class="center">
            <a class="btn small" href="/cash/edit?id={h(rid)}">Изменить</a>
            <form method="POST" action="/action/cash/delete" style="display:inline;margin:0">
              <input type="hidden" name="id" value="{h(rid)}"/>
              <button class="btn small" type="submit" onclick="return confirm('Удалить запись?')">Удалить</button>
            </form>
          </td>
        </tr>
        """)

    table_html = f"""
    <table class="table">
      <thead>
        <tr>
          <th style="width:120px" class="center">Дата</th>
          <th>Номенклатура</th>
          <th style="width:160px" class="right">Сумма</th>
          <th style="width:200px" class="center">Действие</th>
        </tr>
      </thead>
      <tbody>
        {''.join(rows_html) if rows_html else '<tr><td colspan="4" class="muted">Пока нет записей. Добавьте наличный платёж.</td></tr>'}
      </tbody>
    </table>
    """

    body += table_html
    return render_layout(path="/cash", title="Касса", crumbs="Операции → Касса", body_html=body, flash=flash_box(flash))


def page_cash_form(path: str, qd: Dict[str, str], flash: str = "") -> str:
    rid = qd.get("id", "")
    is_edit = bool(rid)
    row = STATE.get_cash_row(rid) if is_edit else {}
    
    today = date.today().strftime("%d.%m.%Y")
    d = row.get("date", today) if row else today
    nom = row.get("nomenclature", "") if row else ""
    amt = row.get("amount", "") if row else ""

    title = "Изменить запись" if is_edit else "Добавить наличный платёж"
    action = "/action/cash/update" if is_edit else "/action/cash/add"

    body = f"""
    <div class="pagehead">
      <div>
        <h1>{title}</h1>
        <div class="smallnote">Введите данные о наличном платеже.</div>
      </div>
      <div class="toolbar">
        <a class="btn" href="/cash">← Назад</a>
      </div>
    </div>

    <div class="panel">
      <form method="POST" action="{action}">
        <input type="hidden" name="id" value="{h(rid)}"/>
        <div style="display:grid;grid-template-columns:1fr 1fr 1fr;gap:16px">
          <div>
            <div class="smallnote"><b>Дата</b></div>
            <input name="date" value="{h(d)}" placeholder="ДД.ММ.ГГГГ" style="width:100%;padding:10px;border:1px solid #D7D7D7;border-radius:12px" required />
          </div>
          <div>
            <div class="smallnote"><b>Номенклатура</b></div>
            <input name="nomenclature" value="{h(nom)}" placeholder="Описание товара/услуги" style="width:100%;padding:10px;border:1px solid #D7D7D7;border-radius:12px" required />
          </div>
          <div>
            <div class="smallnote"><b>Сумма</b></div>
            <input name="amount" value="{h(fmt_num(amt) if amt else '')}" placeholder="0,00" style="width:100%;padding:10px;border:1px solid #D7D7D7;border-radius:12px" required />
          </div>
        </div>
        <div style="margin-top:16px;display:flex;gap:10px">
          <button class="btn primary" type="submit">Сохранить</button>
          <a class="btn" href="/cash">Отмена</a>
        </div>
      </form>
    </div>
    """
    return render_layout(path="/cash", title=title, crumbs=f"Операции → Касса → {title}", body_html=body, flash=flash_box(flash))


def parse_ozon_excel_total(file_path: str) -> Decimal:
    if not Workbook or not load_workbook:
        return Decimal("0")
    try:
        wb = load_workbook(file_path, data_only=True)
        ws = wb.active
        for row in ws.iter_rows(values_only=True):
            for i, cell in enumerate(row):
                if cell and "итого реализовано" in str(cell).lower() and "вычетом возвратов" in str(cell).lower():
                    for j in range(i+1, len(row)):
                        if row[j] is not None:
                            return decimal_from_str(str(row[j]))
        for row in ws.iter_rows(values_only=True):
            for i, cell in enumerate(row):
                if cell and "итого" in str(cell).lower():
                    for j in range(i+1, len(row)):
                        if row[j] is not None:
                            val = decimal_from_str(str(row[j]))
                            if val > 0:
                                return val
    except Exception:
        pass
    return Decimal("0")


def parse_wb_pdf_total(file_path: str) -> Decimal:
    if pdfplumber is None:
        return Decimal("0")
    try:
        with pdfplumber.open(file_path) as pdf:
            for page in pdf.pages:
                text = page.extract_text() or ""
                for line in text.split("\n"):
                    low = line.lower()
                    if "итого к перечислению продавцу" in low or "к перечислению продавцу" in low:
                        nums = re.findall(r"[\d\s]+[,.]?\d*", line)
                        for n in reversed(nums):
                            val = decimal_from_str(n)
                            if val > 0:
                                return val
    except Exception:
        pass
    return Decimal("0")


def get_period_dates(period_type: str, year: int, quarter: int, month: int) -> Tuple[str, str]:
    if period_type == "quarter":
        start_month = (quarter - 1) * 3 + 1
        end_month = start_month + 2
        start_date = date(year, start_month, 1)
        if end_month == 12:
            end_date = date(year, 12, 31)
        else:
            end_date = date(year, end_month + 1, 1) - timedelta(days=1)
    else:
        start_date = date(year, month, 1)
        if month == 12:
            end_date = date(year, 12, 31)
        else:
            end_date = date(year, month + 1, 1) - timedelta(days=1)
    return start_date.strftime("%d.%m.%Y"), end_date.strftime("%d.%m.%Y")


def get_period_label(period_type: str, year: int, quarter: int, month: int) -> str:
    if period_type == "quarter":
        return f"{quarter} квартал {year}"
    else:
        months_ru = ["", "январь", "февраль", "март", "апрель", "май", "июнь", 
                     "июль", "август", "сентябрь", "октябрь", "ноябрь", "декабрь"]
        return f"{months_ru[month]} {year}"


def page_marketplace(path: str, flash: str = "") -> str:
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    cur.execute("SELECT * FROM marketplace_rows ORDER BY year DESC, quarter DESC, month DESC, created_at DESC")
    rows = cur.fetchall()
    cur.close()
    return_db_connection(conn)

    total_sum = sum(Decimal(str(r.get("amount") or 0)) for r in rows)

    body = f"""
    <div class="pagehead">
      <div>
        <h1>Маркетплейс</h1>
        <div class="smallnote">Учёт поступлений с Wildberries и Ozon. Суммы могут участвовать в расчёте налогов.</div>
      </div>
      <div class="toolbar">
        <a class="btn primary" href="/marketplace/add?platform=Wildberries">Поступления Wildberries...</a>
        <a class="btn primary" href="/marketplace/add?platform=Ozon">Поступления Ozon...</a>
      </div>
    </div>

    <div class="panel" style="margin-bottom:12px;display:flex;gap:14px;align-items:center;flex-wrap:wrap">
      <span class="badge2"><b>Записей:</b>&nbsp;{len(rows)}</span>
      <span class="badge2"><b>Сумма поступлений:</b>&nbsp;{h(fmt_money(total_sum))}</span>
    </div>
    """

    rows_html = []
    for r in rows:
        rid = r.get("id", "")
        amt = Decimal(str(r.get("amount") or 0))
        rows_html.append(f'''
        <tr>
          <td class="center">{h(r.get("platform",""))}</td>
          <td class="center">{h(r.get("period_label",""))}</td>
          <td class="center">{h(r.get("date_start",""))}</td>
          <td class="center">{h(r.get("date_end",""))}</td>
          <td class="right"><span class="income">{h(fmt_money(amt))}</span></td>
          <td class="center">
            <form method="POST" action="/action/marketplace/delete" style="display:inline;margin:0">
              <input type="hidden" name="id" value="{h(rid)}"/>
              <button class="btn small" type="submit" onclick="return confirm('Удалить запись?')">Удалить</button>
            </form>
          </td>
        </tr>
        ''')

    table_html = f"""
    <table class="table">
      <thead>
        <tr>
          <th style="width:140px" class="center">Площадка</th>
          <th style="width:180px" class="center">Период</th>
          <th style="width:110px" class="center">С</th>
          <th style="width:110px" class="center">По</th>
          <th style="width:160px" class="right">Поступления</th>
          <th style="width:120px" class="center">Действие</th>
        </tr>
      </thead>
      <tbody>
        {''.join(rows_html) if rows_html else '<tr><td colspan="6" class="muted">Пока нет записей. Добавьте поступление с маркетплейса.</td></tr>'}
      </tbody>
    </table>
    """

    body += table_html
    return render_layout(path="/marketplace", title="Маркетплейс", crumbs="Операции → Маркетплейс", body_html=body, flash=flash_box(flash))


def page_marketplace_add(path: str, qd: Dict[str, str], flash: str = "") -> str:
    platform = qd.get("platform", "Wildberries")
    current_year = date.today().year

    file_type = "PDF" if platform == "Wildberries" else "Excel"
    file_accept = ".pdf" if platform == "Wildberries" else ".xlsx,.xlsm"

    body = f"""
    <div class="pagehead">
      <div>
        <h1>Поступления {h(platform)}</h1>
        <div class="smallnote">Добавьте запись о поступлении с {h(platform)}.</div>
      </div>
    </div>

    <div class="panel">
      <form method="POST" action="/action/marketplace/add" enctype="multipart/form-data">
        <input type="hidden" name="platform" value="{h(platform)}"/>
        
        <div style="margin-bottom:16px">
          <label style="font-weight:600;display:block;margin-bottom:6px">Тип периода</label>
          <select name="period_type" id="period_type" style="width:200px;padding:10px;border:1px solid #D7D7D7;border-radius:12px" onchange="togglePeriod()">
            <option value="quarter">Квартал</option>
            <option value="month">Месяц</option>
          </select>
        </div>

        <div style="margin-bottom:16px">
          <label style="font-weight:600;display:block;margin-bottom:6px">Год</label>
          <input type="number" name="year" value="{current_year}" min="2020" max="2030" style="width:120px;padding:10px;border:1px solid #D7D7D7;border-radius:12px"/>
        </div>

        <div id="quarter_block" style="margin-bottom:16px">
          <label style="font-weight:600;display:block;margin-bottom:6px">Квартал</label>
          <select name="quarter" style="width:200px;padding:10px;border:1px solid #D7D7D7;border-radius:12px">
            <option value="1">1 квартал</option>
            <option value="2">2 квартал</option>
            <option value="3">3 квартал</option>
            <option value="4">4 квартал</option>
          </select>
        </div>

        <div id="month_block" style="margin-bottom:16px;display:none">
          <label style="font-weight:600;display:block;margin-bottom:6px">Месяц</label>
          <select name="month" style="width:200px;padding:10px;border:1px solid #D7D7D7;border-radius:12px">
            <option value="1">Январь</option>
            <option value="2">Февраль</option>
            <option value="3">Март</option>
            <option value="4">Апрель</option>
            <option value="5">Май</option>
            <option value="6">Июнь</option>
            <option value="7">Июль</option>
            <option value="8">Август</option>
            <option value="9">Сентябрь</option>
            <option value="10">Октябрь</option>
            <option value="11">Ноябрь</option>
            <option value="12">Декабрь</option>
          </select>
        </div>

        <div style="margin-bottom:16px">
          <label style="font-weight:600;display:block;margin-bottom:6px">Способ ввода суммы</label>
          <select name="input_mode" id="input_mode" style="width:200px;padding:10px;border:1px solid #D7D7D7;border-radius:12px" onchange="toggleInputMode()">
            <option value="manual">Вручную</option>
            <option value="file">Загрузить файл ({file_type})</option>
          </select>
        </div>

        <div id="manual_block" style="margin-bottom:16px">
          <label style="font-weight:600;display:block;margin-bottom:6px">Сумма поступления</label>
          <input type="text" name="amount" placeholder="0.00" style="width:200px;padding:10px;border:1px solid #D7D7D7;border-radius:12px"/>
        </div>

        <div id="file_block" style="margin-bottom:16px;display:none">
          <label style="font-weight:600;display:block;margin-bottom:6px">Файл ({file_type})</label>
          <input type="file" name="file" accept="{file_accept}" style="padding:10px"/>
          <div class="smallnote" style="margin-top:6px">
            {"Программа найдёт строку «Итого к перечислению продавцу» и подставит сумму." if platform == "Wildberries" else "Программа найдёт строку «Итого реализовано (за вычетом возвратов)» и подставит сумму."}
          </div>
        </div>

        <div style="margin-top:20px;display:flex;gap:10px">
          <button class="btn primary" type="submit">Добавить</button>
          <a class="btn" href="/marketplace">Отмена</a>
        </div>
      </form>
    </div>

    <script>
    function togglePeriod() {{
      var pt = document.getElementById('period_type').value;
      document.getElementById('quarter_block').style.display = pt === 'quarter' ? 'block' : 'none';
      document.getElementById('month_block').style.display = pt === 'month' ? 'block' : 'none';
    }}
    function toggleInputMode() {{
      var im = document.getElementById('input_mode').value;
      document.getElementById('manual_block').style.display = im === 'manual' ? 'block' : 'none';
      document.getElementById('file_block').style.display = im === 'file' ? 'block' : 'none';
    }}
    </script>
    """

    return render_layout(path="/marketplace", title=f"Поступления {platform}", crumbs=f"Операции → Маркетплейс → {platform}", body_html=body, flash=flash_box(flash))


def page_counterparties(path: str, flash: str = "") -> str:
    counterparties = STATE.counterparties
    our_id = STATE.settings.get("our_company_id")

    body = f"""
    <div class="pagehead">
      <div>
        <h1>Контрагенты</h1>
        <div class="smallnote">Справочник контрагентов. Автоматически пополняется из банковских выписок.</div>
      </div>
      <div class="toolbar">
        <a class="btn primary" href="/counterparties/new">+ Добавить</a>
      </div>
    </div>
    """

    rows = []
    for c in counterparties:
        cid = c.get("id", "")
        is_our = (cid == our_id)
        our_badge = '<span class="badge2" style="background:#d4edda;color:#155724">Наша компания</span>' if is_our else ""
        rows.append(f"""
        <tr>
          <td>{h(c.get("name",""))} {our_badge}</td>
          <td>{h(c.get("inn",""))}</td>
          <td>{h(c.get("kpp",""))}</td>
          <td>{h(c.get("bank",""))}</td>
          <td>{h(c.get("account",""))}</td>
          <td class="center">
            <a class="btn small" href="/counterparties/edit?id={h(cid)}">Изменить</a>
          </td>
        </tr>
        """)

    body += f"""
    <table class="table">
      <thead>
        <tr>
          <th>Наименование</th>
          <th style="width:130px">ИНН</th>
          <th style="width:100px">КПП</th>
          <th style="width:200px">Банк</th>
          <th style="width:200px">Р/счет</th>
          <th style="width:100px" class="center">Действие</th>
        </tr>
      </thead>
      <tbody>
        {''.join(rows) if rows else '<tr><td colspan="6" class="muted">Нет контрагентов.</td></tr>'}
      </tbody>
    </table>
    """
    return render_layout("/counterparties", "Контрагенты", "Справочники → Контрагенты", body, flash=flash_box(flash))


def get_our_company_fields_html(c: Dict[str, Any], is_our: bool) -> str:
    display = "block" if is_our else "none"
    return f"""
    <div id="ourCompanyFields" style="display:{display}">
      <div class="hr"></div>
      <b>Данные организации (для отчётности)</b>
      <div style="display:grid;grid-template-columns:1fr 1fr;gap:12px;margin-top:10px">
        <div><div class="smallnote"><b>Код инспекции</b></div><input name="inspection_code" value="{h(c.get('inspection_code') or '')}" style="width:100%;padding:10px;border:1px solid #D7D7D7;border-radius:12px"/></div>
        <div><div class="smallnote"><b>Полное наименование</b></div><input name="full_name" value="{h(c.get('full_name') or '')}" style="width:100%;padding:10px;border:1px solid #D7D7D7;border-radius:12px"/></div>
        <div><div class="smallnote"><b>ОКТМО</b></div><input name="oktmo" value="{h(c.get('oktmo') or '')}" style="width:100%;padding:10px;border:1px solid #D7D7D7;border-radius:12px"/></div>
        <div><div class="smallnote"><b>ОКАТО</b></div><input name="okato" value="{h(c.get('okato') or '')}" style="width:100%;padding:10px;border:1px solid #D7D7D7;border-radius:12px"/></div>
        <div style="grid-column:span 2"><div class="smallnote"><b>Отчётность подписывает</b></div><input name="signatory" value="{h(c.get('signatory') or '')}" style="width:100%;padding:10px;border:1px solid #D7D7D7;border-radius:12px"/></div>
      </div>
      <div class="hr"></div>
      <b>Социальный фонд</b>
      <div style="margin-top:10px"><div class="smallnote"><b>Регистрационный номер</b></div><input name="sfr_reg_number" value="{h(c.get('sfr_reg_number') or '')}" style="width:100%;padding:10px;border:1px solid #D7D7D7;border-radius:12px"/></div>
      <div class="hr"></div>
      <b>Пенсионный фонд</b>
      <div style="display:grid;grid-template-columns:1fr 1fr;gap:12px;margin-top:10px">
        <div><div class="smallnote"><b>Рег. номер (за себя)</b></div><input name="pfr_reg_self" value="{h(c.get('pfr_reg_self') or '')}" style="width:100%;padding:10px;border:1px solid #D7D7D7;border-radius:12px"/></div>
        <div><div class="smallnote"><b>Рег. номер (за сотрудников)</b></div><input name="pfr_reg_employees" value="{h(c.get('pfr_reg_employees') or '')}" style="width:100%;padding:10px;border:1px solid #D7D7D7;border-radius:12px"/></div>
        <div><div class="smallnote"><b>Код терр. органа</b></div><input name="pfr_terr_code" value="{h(c.get('pfr_terr_code') or '')}" style="width:100%;padding:10px;border:1px solid #D7D7D7;border-radius:12px"/></div>
        <div><div class="smallnote"><b>Территориальный орган</b></div><input name="pfr_terr_organ" value="{h(c.get('pfr_terr_organ') or '')}" style="width:100%;padding:10px;border:1px solid #D7D7D7;border-radius:12px"/></div>
      </div>
      <div class="hr"></div>
      <b>Платёжные реквизиты (для отчётов)</b>
      <div style="margin-top:10px"><textarea name="payment_details" style="width:100%;min-height:60px;padding:10px;border:1px solid #D7D7D7;border-radius:12px">{h(c.get('payment_details') or '')}</textarea></div>
      <div class="hr"></div>
      <b>Коды статистики</b>
      <div style="display:grid;grid-template-columns:1fr 1fr 1fr;gap:12px;margin-top:10px">
        <div><div class="smallnote"><b>ОКПО</b></div><input name="okpo" value="{h(c.get('okpo') or '')}" style="width:100%;padding:10px;border:1px solid #D7D7D7;border-radius:12px"/></div>
        <div><div class="smallnote"><b>ОКОПФ</b></div><input name="okopf" value="{h(c.get('okopf') or '')}" style="width:100%;padding:10px;border:1px solid #D7D7D7;border-radius:12px"/></div>
        <div><div class="smallnote"><b>ОКФС</b></div><input name="okfs" value="{h(c.get('okfs') or '')}" style="width:100%;padding:10px;border:1px solid #D7D7D7;border-radius:12px"/></div>
        <div><div class="smallnote"><b>ОКВЭД ред.1</b></div><input name="okved1" value="{h(c.get('okved1') or '')}" style="width:100%;padding:10px;border:1px solid #D7D7D7;border-radius:12px"/></div>
        <div><div class="smallnote"><b>ОКВЭД ред.2</b></div><input name="okved2" value="{h(c.get('okved2') or '')}" style="width:100%;padding:10px;border:1px solid #D7D7D7;border-radius:12px"/></div>
        <div><div class="smallnote"><b>ОКПО Росстата</b></div><input name="okpo_rosstat" value="{h(c.get('okpo_rosstat') or '')}" style="width:100%;padding:10px;border:1px solid #D7D7D7;border-radius:12px"/></div>
      </div>
    </div>
    """

def page_counterparty_new(flash: str = "") -> str:
    our_fields = get_our_company_fields_html({}, False)
    body = f"""
    <div class="pagehead">
      <div>
        <h1>Новый контрагент</h1>
      </div>
      <div class="toolbar">
        <a class="btn" href="/counterparties">← Назад</a>
      </div>
    </div>

    <form method="POST" action="/action/cp/save" class="panel">
      <input type="hidden" name="id" value=""/>
      <div style="display:grid;grid-template-columns:1fr 1fr;gap:12px">
        <div>
          <div class="smallnote"><b>Тип</b></div>
          <select name="kind" style="width:100%;padding:10px;border:1px solid #D7D7D7;border-radius:12px">
            <option value="Юридическое лицо">Юридическое лицо</option>
            <option value="ИП">ИП</option>
            <option value="Физическое лицо">Физическое лицо</option>
          </select>
        </div>
        <div>
          <div class="smallnote"><b>Наименование</b></div>
          <input name="name" style="width:100%;padding:10px;border:1px solid #D7D7D7;border-radius:12px" required/>
        </div>
        <div>
          <div class="smallnote"><b>ИНН</b></div>
          <input name="inn" style="width:100%;padding:10px;border:1px solid #D7D7D7;border-radius:12px"/>
        </div>
        <div>
          <div class="smallnote"><b>КПП</b></div>
          <input name="kpp" style="width:100%;padding:10px;border:1px solid #D7D7D7;border-radius:12px"/>
        </div>
        <div>
          <div class="smallnote"><b>Банк</b></div>
          <input name="bank" style="width:100%;padding:10px;border:1px solid #D7D7D7;border-radius:12px"/>
        </div>
        <div>
          <div class="smallnote"><b>БИК</b></div>
          <input name="bik" style="width:100%;padding:10px;border:1px solid #D7D7D7;border-radius:12px"/>
        </div>
        <div>
          <div class="smallnote"><b>Корр. счет</b></div>
          <input name="corr" style="width:100%;padding:10px;border:1px solid #D7D7D7;border-radius:12px"/>
        </div>
        <div>
          <div class="smallnote"><b>Р/счет</b></div>
          <input name="account" style="width:100%;padding:10px;border:1px solid #D7D7D7;border-radius:12px"/>
        </div>
        <div style="grid-column:span 2">
          <div class="smallnote"><b>Юр. адрес</b></div>
          <input name="legal_address" style="width:100%;padding:10px;border:1px solid #D7D7D7;border-radius:12px"/>
        </div>
        <div>
          <div class="smallnote"><b>Телефон</b></div>
          <input name="phone" style="width:100%;padding:10px;border:1px solid #D7D7D7;border-radius:12px"/>
        </div>
        <div>
          <label style="display:flex;align-items:center;gap:8px;margin-top:20px">
            <input type="checkbox" name="is_our" value="1" id="isOurCheck" onchange="toggleOurFields()"/>
            <span><b>Это наша организация</b></span>
          </label>
        </div>
      </div>
      {our_fields}
      <div class="hr"></div>
      <div style="display:flex;gap:10px">
        <button class="btn primary" type="submit">Сохранить</button>
        <a class="btn" href="/counterparties">Отмена</a>
      </div>
    </form>
    <script>
      function toggleOurFields() {{
        var cb = document.getElementById('isOurCheck');
        var fields = document.getElementById('ourCompanyFields');
        fields.style.display = cb.checked ? 'block' : 'none';
      }}
    </script>
    """
    return render_layout("/counterparties", "Новый контрагент", "Справочники → Контрагенты → Новый", body, flash=flash_box(flash))


def page_counterparty_edit(qd: Dict[str, str], flash: str = "") -> str:
    cid = qd.get("id", "")
    c = STATE.get_counterparty_by_id(cid)
    if not c:
        return render_layout("/counterparties", "Изменить", "Справочники → Контрагенты", "<h1>Контрагент не найден</h1>", flash=flash_box(flash))

    our_id = STATE.settings.get("our_company_id")
    is_our = (cid == our_id) or c.get("is_our_company", False)
    our_fields = get_our_company_fields_html(c, is_our)

    body = f"""
    <div class="pagehead">
      <div>
        <h1>Изменить контрагента</h1>
      </div>
      <div class="toolbar">
        <a class="btn" href="/counterparties">← Назад</a>
      </div>
    </div>

    <form method="POST" action="/action/cp/save" class="panel">
      <input type="hidden" name="id" value="{h(cid)}"/>
      <div style="display:grid;grid-template-columns:1fr 1fr;gap:12px">
        <div>
          <div class="smallnote"><b>Тип</b></div>
          <select name="kind" style="width:100%;padding:10px;border:1px solid #D7D7D7;border-radius:12px">
            <option value="Юридическое лицо" {'selected' if c.get('kind')=='Юридическое лицо' else ''}>Юридическое лицо</option>
            <option value="ИП" {'selected' if c.get('kind')=='ИП' else ''}>ИП</option>
            <option value="Физическое лицо" {'selected' if c.get('kind')=='Физическое лицо' else ''}>Физическое лицо</option>
          </select>
        </div>
        <div>
          <div class="smallnote"><b>Наименование</b></div>
          <input name="name" value="{h(c.get('name',''))}" style="width:100%;padding:10px;border:1px solid #D7D7D7;border-radius:12px" required/>
        </div>
        <div>
          <div class="smallnote"><b>ИНН</b></div>
          <input name="inn" value="{h(c.get('inn',''))}" style="width:100%;padding:10px;border:1px solid #D7D7D7;border-radius:12px"/>
        </div>
        <div>
          <div class="smallnote"><b>КПП</b></div>
          <input name="kpp" value="{h(c.get('kpp',''))}" style="width:100%;padding:10px;border:1px solid #D7D7D7;border-radius:12px"/>
        </div>
        <div>
          <div class="smallnote"><b>Банк</b></div>
          <input name="bank" value="{h(c.get('bank',''))}" style="width:100%;padding:10px;border:1px solid #D7D7D7;border-radius:12px"/>
        </div>
        <div>
          <div class="smallnote"><b>БИК</b></div>
          <input name="bik" value="{h(c.get('bik',''))}" style="width:100%;padding:10px;border:1px solid #D7D7D7;border-radius:12px"/>
        </div>
        <div>
          <div class="smallnote"><b>Корр. счет</b></div>
          <input name="corr" value="{h(c.get('corr',''))}" style="width:100%;padding:10px;border:1px solid #D7D7D7;border-radius:12px"/>
        </div>
        <div>
          <div class="smallnote"><b>Р/счет</b></div>
          <input name="account" value="{h(c.get('account',''))}" style="width:100%;padding:10px;border:1px solid #D7D7D7;border-radius:12px"/>
        </div>
        <div style="grid-column:span 2">
          <div class="smallnote"><b>Юр. адрес</b></div>
          <input name="legal_address" value="{h(c.get('legal_address',''))}" style="width:100%;padding:10px;border:1px solid #D7D7D7;border-radius:12px"/>
        </div>
        <div>
          <div class="smallnote"><b>Телефон</b></div>
          <input name="phone" value="{h(c.get('phone',''))}" style="width:100%;padding:10px;border:1px solid #D7D7D7;border-radius:12px"/>
        </div>
        <div>
          <label style="display:flex;align-items:center;gap:8px;margin-top:20px">
            <input type="checkbox" name="is_our" value="1" id="isOurCheck" {'checked' if is_our else ''} onchange="toggleOurFields()"/>
            <span><b>Это наша организация</b></span>
          </label>
        </div>
      </div>
      {our_fields}
      <div class="hr"></div>
      <div style="display:flex;gap:10px">
        <button class="btn primary" type="submit">Сохранить</button>
        <a class="btn" href="/counterparties">Отмена</a>
      </div>
    </form>
    <div class="hr"></div>
    <form method="POST" action="/action/cp/delete" style="margin-top:12px">
      <input type="hidden" name="id" value="{h(cid)}"/>
      <button class="btn danger" type="submit" onclick="return confirm('Удалить контрагента?')">Удалить</button>
    </form>
    <script>
      function toggleOurFields() {{
        var cb = document.getElementById('isOurCheck');
        var fields = document.getElementById('ourCompanyFields');
        fields.style.display = cb.checked ? 'block' : 'none';
      }}
    </script>
    """
    return render_layout("/counterparties", "Изменить контрагента", "Справочники → Контрагенты → Изменить", body, flash=flash_box(flash))


def page_search(qd: Dict[str, str], flash: str = "") -> str:
    q = (qd.get("q") or "").strip()
    items = []

    if q:
        q_upper = norm_text(q)
        bank_rows = STATE.bank_rows
        for r in bank_rows:
            if q_upper in norm_text(r.get("counterparty", "")) or q_upper in norm_text(r.get("purpose", "")):
                items.append((
                    "Банк",
                    f"{r.get('date','')} — {r.get('counterparty','')}",
                    f"Сумма: {fmt_money(r.get('incoming') or r.get('outgoing') or 0)}",
                    "/bank"
                ))
                if len(items) >= 120:
                    break
        
        counterparties = STATE.counterparties
        for c in counterparties:
            if q_upper in norm_text(c.get("name", "")) or q_upper in norm_text(c.get("inn", "")):
                items.append((
                    "Контрагент",
                    c.get("name", ""),
                    f"ИНН: {c.get('inn','')} | Банк: {c.get('bank','')}",
                    f"/counterparties/edit?id={c.get('id','')}"
                ))
                if len(items) >= 120:
                    break

    rows = []
    for kind, title, meta, link in items:
        rows.append(f"""
        <tr>
          <td style="width:120px"><span class="badge2">{h(kind)}</span></td>
          <td><a href="{h(link)}"><b>{h(title)}</b></a><div class="smallnote">{h(meta)}</div></td>
        </tr>
        """)

    body = f"""
    <div class="pagehead">
      <div>
        <h1>Поиск</h1>
        <div class="smallnote">По банку и контрагентам.</div>
      </div>
      <div class="toolbar">
        <a class="btn" href="/bank">← Банк</a>
        <a class="btn" href="/counterparties">Контрагенты</a>
      </div>
    </div>

    <div class="panel" style="margin-bottom:12px">
      <form method="GET" action="/search" style="display:flex;gap:10px;align-items:center;margin:0">
        <div style="flex:1">
          <div class="smallnote"><b>Запрос</b></div>
          <input name="q" value="{h(q)}" style="width:100%;padding:10px;border:1px solid #D7D7D7;border-radius:12px"
                 placeholder="Контрагент / ИНН / назначение платежа …"/>
        </div>
        <div style="padding-top:18px">
          <button class="btn primary" type="submit">Найти</button>
        </div>
      </form>
    </div>

    <table class="table">
      <thead><tr><th>Тип</th><th>Результат</th></tr></thead>
      <tbody>
        {''.join(rows) if rows else '<tr><td colspan="2" class="muted">Введите запрос.</td></tr>'}
      </tbody>
    </table>
    """
    return render_layout("/search", "Поиск", "Поиск", body, flash=flash_box(flash))


RU_MONTH_NAMES = {
    "01": "январь", "02": "февраль", "03": "март", "04": "апрель",
    "05": "май", "06": "июнь", "07": "июль", "08": "август",
    "09": "сентябрь", "10": "октябрь", "11": "ноябрь", "12": "декабрь",
}


def in_range(d: Optional[date], start: Optional[date], end: Optional[date]) -> bool:
    if d is None:
        return False if (start and end) else True
    if start and end:
        return start <= d <= end
    return True


def compute_date_range(mode: str, dfrom: str, dto: str) -> Tuple[Optional[date], Optional[date], str]:
    if mode == "custom":
        d1 = parse_date_ddmmyyyy(dfrom)
        d2 = parse_date_ddmmyyyy(dto)
        if not d1 or not d2:
            raise ValueError("Введите даты в формате ДД.ММ.ГГГГ")
        if d2 < d1:
            raise ValueError("'По' должна быть >= 'С'")
        return d1, d2, f"{format_ddmmyyyy(d1)} – {format_ddmmyyyy(d2)}"
    if mode == "year":
        year = date.today().year
        return date(year, 1, 1), date(year, 12, 31), f"{year} год"
    if mode == "month":
        today = date.today()
        start = date(today.year, today.month, 1)
        if today.month == 12:
            end = date(today.year, 12, 31)
        else:
            end = date(today.year, today.month + 1, 1) - timedelta(days=1)
        return start, end, f"{RU_MONTH_NAMES.get(str(today.month).zfill(2), '')} {today.year}"
    return None, None, "За всё время"


def _period_form_html(action: str, mode: str, dfrom: str, dto: str) -> str:
    return f"""
    <div style="display:flex;gap:12px;flex-wrap:wrap;align-items:flex-end">
      <div style="max-width:200px">
        <div class="smallnote"><b>Период</b></div>
        <select name="mode" style="width:100%;padding:10px;border:1px solid #D7D7D7;border-radius:12px">
          <option value="all" {"selected" if mode=="all" else ""}>За всё время</option>
          <option value="year" {"selected" if mode=="year" else ""}>Текущий год</option>
          <option value="month" {"selected" if mode=="month" else ""}>Текущий месяц</option>
          <option value="custom" {"selected" if mode=="custom" else ""}>Произвольный</option>
        </select>
      </div>
      <div style="max-width:160px">
        <div class="smallnote"><b>с</b></div>
        <input name="from" value="{h(dfrom)}" placeholder="ДД.ММ.ГГГГ" style="width:100%;padding:10px;border:1px solid #D7D7D7;border-radius:12px"/>
      </div>
      <div style="max-width:160px">
        <div class="smallnote"><b>по</b></div>
        <input name="to" value="{h(dto)}" placeholder="ДД.ММ.ГГГГ" style="width:100%;padding:10px;border:1px solid #D7D7D7;border-radius:12px"/>
      </div>
      <button class="btn primary" type="submit">Сформировать</button>
    </div>
    """


def counterparty_names() -> List[str]:
    return [c.get("name", "") for c in STATE.counterparties if c.get("name")]


def party_from_counterparty(c: Dict[str, Any]) -> Dict[str, Any]:
    return {
        "name": c.get("name", ""),
        "inn": c.get("inn", ""),
        "kpp": c.get("kpp", ""),
        "account": c.get("account", ""),
        "bank": c.get("bank", ""),
        "bik": c.get("bik", ""),
        "corr": c.get("corr", ""),
        "address": c.get("legal_address", ""),
    }


def build_act_excel(out_path: str, act: Dict[str, Any]):
    require_openpyxl()
    wb = Workbook()
    ws = wb.active
    ws.title = "Акт"
    thin = Side(style="thin", color="9E9E9E")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def set_cell(a1: str, value, bold=False, size=11, align="left"):
        c = ws[a1]
        c.value = value
        c.font = Font(bold=bold, size=size)
        c.alignment = Alignment(horizontal=align, vertical="top", wrap_text=True)
        return c

    doc_no = act.get("doc_no", "")
    doc_date = act.get("doc_date", "")
    set_cell("A1", f"Акт № {doc_no} от {doc_date}", bold=True, size=14)
    ws.merge_cells("A1:F1")

    executor = act.get("executor", {}) or {}
    customer = act.get("customer", {}) or {}

    set_cell("A3", "Исполнитель:", bold=True)
    set_cell("B3", executor.get("name", ""))
    ws.merge_cells("B3:F3")
    set_cell("A4", "ИНН:", bold=True)
    set_cell("B4", executor.get("inn", ""))
    set_cell("C4", "КПП:", bold=True)
    set_cell("D4", executor.get("kpp", ""))
    set_cell("A5", "р/с:", bold=True)
    set_cell("B5", executor.get("account", ""))
    set_cell("C5", "в банке:", bold=True)
    set_cell("D5", executor.get("bank", ""))
    ws.merge_cells("D5:F5")
    set_cell("A6", "БИК:", bold=True)
    set_cell("B6", executor.get("bik", ""))
    set_cell("C6", "к/с:", bold=True)
    set_cell("D6", executor.get("corr", ""))
    ws.merge_cells("D6:F6")

    set_cell("A8", "Заказчик:", bold=True)
    set_cell("B8", customer.get("name", ""))
    ws.merge_cells("B8:F8")
    set_cell("A9", "ИНН:", bold=True)
    set_cell("B9", customer.get("inn", ""))
    set_cell("C9", "КПП:", bold=True)
    set_cell("D9", customer.get("kpp", ""))
    set_cell("A10", "Адрес:", bold=True)
    set_cell("B10", customer.get("address", ""))
    ws.merge_cells("B10:F10")

    set_cell("A12", "Основание:", bold=True)
    set_cell("B12", act.get("basis", "") or "")
    ws.merge_cells("B12:F12")

    start_row = 14
    headers = ["№", "Наименование работ, услуг", "Кол-во", "Ед.", "Цена", "Сумма"]
    for i, htxt in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=i)
        cell.value = htxt
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
        cell.fill = PatternFill("solid", fgColor="E2EFDA")

    lines = act.get("lines", []) or []
    row = start_row + 1
    total = Decimal("0")

    for idx, ln in enumerate(lines, start=1):
        qty = decimal_from_str(ln.get("qty", 0))
        price = decimal_from_str(ln.get("price", 0))
        amount = decimal_from_str(ln.get("amount", 0))
        if amount == 0 and qty and price:
            amount = money2(qty * price)
        total += amount

        values = [
            str(idx),
            ln.get("name", ""),
            fmt_num(qty, decimals=2, strip_trailing_zeros=True),
            ln.get("unit", "шт"),
            fmt_money(price),
            fmt_money(amount),
        ]
        for c, v in enumerate(values, start=1):
            cell = ws.cell(row=row, column=c, value=v)
            cell.border = border
            if c in (3, 5, 6):
                cell.alignment = Alignment(horizontal="right", vertical="top", wrap_text=True)
            elif c == 4:
                cell.alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)
            else:
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        row += 1

    row += 1
    set_cell(f"E{row}", "Итого:", bold=True, align="right")
    set_cell(f"F{row}", fmt_money(money2(total)), bold=True, align="right")
    row += 1
    set_cell(f"E{row}", act.get("vat_mode", "Без налога (НДС)"), bold=True, align="right")

    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 55
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 8
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 14

    wb.save(out_path)


def build_payment_order_excel(out_path: str, po: Dict[str, Any]):
    require_openpyxl()
    wb = Workbook()
    ws = wb.active
    ws.title = "Платежка"
    thin = Side(style="thin", color="9E9E9E")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    def box(r1, c1, r2, c2):
        for rr in range(r1, r2 + 1):
            for cc in range(c1, c2 + 1):
                ws.cell(row=rr, column=cc).border = border

    ws["A1"] = "ПЛАТЕЖНОЕ ПОРУЧЕНИЕ"
    ws["A1"].font = Font(bold=True, size=13)
    ws.merge_cells("A1:H1")

    ws["A3"] = f"№ {po.get('number','')}"
    ws["A3"].font = Font(bold=True)
    ws["E3"] = po.get("date", "")
    ws["E3"].font = Font(bold=True)
    ws["G3"] = po.get("pay_type", "Электронно")

    amt = decimal_from_str(po.get("amount", 0))
    ws["A5"] = "Сумма"
    ws["B5"] = fmt_money(amt)
    ws["C5"] = "руб."
    ws["A6"] = "Сумма прописью"
    ws.merge_cells("B6:H6")
    ws["B6"] = po.get("amount_words", "")

    payer = po.get("payer", {}) or {}
    recv = po.get("receiver", {}) or {}

    ws["A8"] = "Плательщик"
    ws.merge_cells("B8:H8")
    ws["B8"] = payer.get("name", "")
    ws["A9"] = "ИНН"
    ws["B9"] = payer.get("inn", "")
    ws["D9"] = "КПП"
    ws["E9"] = payer.get("kpp", "")
    ws["A10"] = "Сч. №"
    ws.merge_cells("B10:H10")
    ws["B10"] = payer.get("account", "")
    ws["A12"] = "Банк плательщика"
    ws.merge_cells("B12:H12")
    ws["B12"] = payer.get("bank", "")
    ws["A13"] = "БИК"
    ws["B13"] = payer.get("bik", "")
    ws["D13"] = "Сч. №"
    ws.merge_cells("E13:H13")
    ws["E13"] = payer.get("corr", "")

    ws["A16"] = "Получатель"
    ws.merge_cells("B16:H16")
    ws["B16"] = recv.get("name", "")
    ws["A17"] = "ИНН"
    ws["B17"] = recv.get("inn", "")
    ws["D17"] = "КПП"
    ws["E17"] = recv.get("kpp", "")
    ws["A18"] = "Сч. №"
    ws.merge_cells("B18:H18")
    ws["B18"] = recv.get("account", "")
    ws["A20"] = "Банк получателя"
    ws.merge_cells("B20:H20")
    ws["B20"] = recv.get("bank", "")
    ws["A21"] = "БИК"
    ws["B21"] = recv.get("bik", "")
    ws["D21"] = "Сч. №"
    ws.merge_cells("E21:H21")
    ws["E21"] = recv.get("corr", "")

    ws["A24"] = "Назначение платежа"
    ws.merge_cells("B24:H26")
    ws["B24"] = po.get("purpose", "")

    ws["A28"] = "Вид оп."
    ws["B28"] = po.get("vid_op", "01")
    ws["D28"] = "Очер. плат."
    ws["E28"] = po.get("ocher", "5")

    box(3, 1, 28, 8)
    for r in range(3, 29):
        for c in range(1, 9):
            cell = ws.cell(row=r, column=c)
            cell.alignment = center if r == 3 else left

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 8
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 22
    ws.column_dimensions["F"].width = 16
    ws.column_dimensions["G"].width = 14
    ws.column_dimensions["H"].width = 18
    wb.save(out_path)


def _required_party_missing(p: Dict[str, Any]) -> List[str]:
    req = ["name", "inn", "account", "bank", "bik"]
    return [k for k in req if not (p.get(k) or "").strip()]


def _act_lines_table(lines: List[Dict[str, Any]]) -> str:
    rows = []
    for i, ln in enumerate(lines):
        rows.append(f"""
        <tr>
          <td class="center">{i+1}</td>
          <td><input name="ln_name_{i}" value="{h(str(ln.get('name','')))}" style="width:100%"/></td>
          <td><input name="ln_qty_{i}" value="{h(str(ln.get('qty','1')))}" style="width:100%;text-align:right"/></td>
          <td><input name="ln_unit_{i}" value="{h(str(ln.get('unit','шт')))}" style="width:100%;text-align:center"/></td>
          <td><input name="ln_price_{i}" value="{h(str(ln.get('price','0')))}" style="width:100%;text-align:right"/></td>
          <td><input name="ln_amount_{i}" value="{h(str(ln.get('amount','')))}" style="width:100%;text-align:right" placeholder="авто"/></td>
          <td class="center"><button class="btn small danger" type="button" onclick="removeLine(this)">×</button></td>
        </tr>
        """)

    js = r"""
<script>
function addLine(){
  const tbody = document.getElementById("linesBody");
  const idx = tbody.children.length;
  const tr = document.createElement("tr");
  tr.innerHTML = `
    <td class="center">${idx+1}</td>
    <td><input name="ln_name_${idx}" value="" style="width:100%"/></td>
    <td><input name="ln_qty_${idx}" value="1" style="width:100%;text-align:right"/></td>
    <td><input name="ln_unit_${idx}" value="шт" style="width:100%;text-align:center"/></td>
    <td><input name="ln_price_${idx}" value="0" style="width:100%;text-align:right"/></td>
    <td><input name="ln_amount_${idx}" value="" style="width:100%;text-align:right" placeholder="авто"/></td>
    <td class="center"><button class="btn small danger" type="button" onclick="removeLine(this)">×</button></td>
  `;
  tbody.appendChild(tr);
  renumber();
}
function removeLine(btn){
  const tr = btn.closest("tr");
  tr.remove();
  renumber();
}
function renumber(){
  const tbody = document.getElementById("linesBody");
  [...tbody.children].forEach((tr, i)=> {
    tr.children[0].textContent = i+1;
    const inputs = tr.querySelectorAll("input");
    inputs.forEach(inp=>{
      const n = inp.getAttribute("name") || "";
      if(n.includes("ln_name_")) inp.setAttribute("name", "ln_name_"+i);
      if(n.includes("ln_qty_")) inp.setAttribute("name", "ln_qty_"+i);
      if(n.includes("ln_unit_")) inp.setAttribute("name", "ln_unit_"+i);
      if(n.includes("ln_price_")) inp.setAttribute("name", "ln_price_"+i);
      if(n.includes("ln_amount_")) inp.setAttribute("name", "ln_amount_"+i);
    });
  });
  document.getElementById("linesCount").value = tbody.children.length;
}
window.addEventListener("load", ()=>{ renumber(); });
</script>
"""
    return f"""
    <input type="hidden" id="linesCount" name="lines_count" value="{len(lines)}"/>
    <table class="table">
      <thead>
        <tr>
          <th style="width:55px" class="center">№</th>
          <th>Наименование*</th>
          <th style="width:110px" class="right">Кол-во*</th>
          <th style="width:90px" class="center">Ед.</th>
          <th style="width:130px" class="right">Цена*</th>
          <th style="width:140px" class="right">Сумма</th>
          <th style="width:60px" class="center">—</th>
        </tr>
      </thead>
      <tbody id="linesBody">{''.join(rows)}</tbody>
    </table>
    <div style="margin-top:10px;display:flex;gap:10px;align-items:center">
      <button class="btn small" type="button" onclick="addLine()">+ Добавить строку</button>
      <span class="smallnote">Если «Сумма» пустая — будет посчитана как Кол-во × Цена.</span>
    </div>
    {js}
    """


def page_acts(path: str, flash: str = "") -> str:
    our = STATE.get_our_company_card()
    our_name = norm_text(our.get("name", "")) if our else OUR_COMPANY_DEFAULT_NAME.upper()
    rows = []
    for a in STATE.acts:
        aid = a.get("id", "")
        lines = a.get("lines", []) or []
        total = Decimal("0")
        for ln in lines:
            amount = decimal_from_str(ln.get("amount", 0))
            if amount == 0:
                qty = decimal_from_str(ln.get("qty", 0))
                price = decimal_from_str(ln.get("price", 0))
                amount = money2(qty * price)
            total += amount
        direction = a.get("direction", "provide")
        executor = a.get("executor", {}) or {}
        customer = a.get("customer", {}) or {}
        other = ""
        if norm_text(executor.get("name", "")) == our_name:
            other = customer.get("name", "")
        elif norm_text(customer.get("name", "")) == our_name:
            other = executor.get("name", "")
        else:
            other = customer.get("name", "") or executor.get("name", "")
        rows.append(f"""
        <tr>
          <td class="center">{h(a.get("doc_no",""))}</td>
          <td class="center">{h(a.get("doc_date",""))}</td>
          <td>{"Оказание услуг" if direction=="provide" else "Получение услуг"}</td>
          <td>{h(other)}</td>
          <td class="right">{h(fmt_money(total)) if total else ""}</td>
          <td class="center" style="white-space:nowrap">
            <a class="btn small" href="/acts/edit?id={h(aid)}">Открыть</a>
            <a class="btn small" href="/action/acts/export?id={h(aid)}">Excel</a>
            <form method="POST" action="/action/acts/delete" style="display:inline;margin:0">
              <input type="hidden" name="id" value="{h(aid)}"/>
              <button class="btn small danger" type="submit">Удалить</button>
            </form>
          </td>
        </tr>
        """)
    body = f"""
    <div class="pagehead">
      <div>
        <h1>Акты / накладные</h1>
        <div class="smallnote">Создание/хранение актов, экспорт Excel.</div>
      </div>
      <div class="toolbar">
        <a class="btn primary" href="/acts/new">+ Создать акт</a>
      </div>
    </div>
    <table class="table">
      <thead>
        <tr>
          <th style="width:90px" class="center">№</th>
          <th style="width:110px" class="center">Дата</th>
          <th style="width:170px">Тип</th>
          <th>Контрагент</th>
          <th style="width:160px" class="right">Сумма</th>
          <th style="width:260px" class="center">Действия</th>
        </tr>
      </thead>
      <tbody>
        {''.join(rows) if rows else '<tr><td colspan="6" class="muted">Документов нет. Создайте акт.</td></tr>'}
      </tbody>
    </table>
    """
    return render_layout("/acts", "Акты", "Документы → Акты / накладные", body, flash=flash_box(flash))


def page_act_editor(mode: str, qd: Dict[str, str], flash: str = "") -> str:
    if mode == "edit":
        aid = (qd.get("id") or "").strip()
        act = STATE.get_act_by_id(aid)
        if not act:
            return render_layout("/acts", "Акт", "Документы → Акты", "<h1>Документ не найден</h1><a class='btn' href='/acts'>Назад</a>", flash=flash_box(flash))
        init = act.copy()
    else:
        init = {
            "id": new_id(),
            "doc_no": "",
            "doc_date": format_ddmmyyyy(date.today()),
            "direction": "provide",
            "executor": {},
            "customer": {},
            "basis": "",
            "vat_mode": "Без налога (НДС)",
            "lines": [{"name": "", "qty": "1", "unit": "шт", "price": "0", "amount": ""}],
        }
        our = STATE.get_our_company_card()
        if our:
            init["executor"] = party_from_counterparty(our)

    direction = init.get("direction", "provide")
    basis_cur = init.get("basis", "") or ""
    basis_opts = ['<option value=""></option>'] + [f'<option value="{h(x)}">{h(x)}</option>' for x in STATE.basis_history[-60:]]

    def party_block(prefix: str, title: str, data: Dict[str, Any]) -> str:
        names = counterparty_names()
        options = ['<option value=""></option>'] + [f'<option value="{h(n)}">{h(n)}</option>' for n in names]
        return f"""
        <div class="panel">
          <div style="display:flex;justify-content:space-between;align-items:center;gap:10px">
            <div><b>{h(title)}</b></div>
            <div style="display:flex;gap:8px;align-items:center">
              <select name="{prefix}_card" style="max-width:260px">{''.join(options)}</select>
              <button class="btn small" name="_apply_card" value="{prefix}" type="submit">Подставить</button>
            </div>
          </div>
          <div class="hr"></div>
          <div class="smallnote"><b>Наименование*</b></div>
          <input name="{prefix}_name" value="{h(data.get('name',''))}" style="width:100%;padding:10px;border:1px solid #D7D7D7;border-radius:12px"/>
          <div style="display:grid;grid-template-columns:1fr 1fr;gap:12px;margin-top:10px">
            <div>
              <div class="smallnote"><b>ИНН*</b></div>
              <input name="{prefix}_inn" value="{h(data.get('inn',''))}" style="width:100%;padding:10px;border:1px solid #D7D7D7;border-radius:12px"/>
            </div>
            <div>
              <div class="smallnote"><b>КПП</b></div>
              <input name="{prefix}_kpp" value="{h(data.get('kpp',''))}" style="width:100%;padding:10px;border:1px solid #D7D7D7;border-radius:12px"/>
            </div>
          </div>
          <div style="margin-top:10px">
            <div class="smallnote"><b>р/с*</b></div>
            <input name="{prefix}_account" value="{h(data.get('account',''))}" style="width:100%;padding:10px;border:1px solid #D7D7D7;border-radius:12px"/>
          </div>
          <div style="margin-top:10px">
            <div class="smallnote"><b>Банк*</b></div>
            <input name="{prefix}_bank" value="{h(data.get('bank',''))}" style="width:100%;padding:10px;border:1px solid #D7D7D7;border-radius:12px"/>
          </div>
          <div style="display:grid;grid-template-columns:1fr 1fr;gap:12px;margin-top:10px">
            <div>
              <div class="smallnote"><b>БИК*</b></div>
              <input name="{prefix}_bik" value="{h(data.get('bik',''))}" style="width:100%;padding:10px;border:1px solid #D7D7D7;border-radius:12px"/>
            </div>
            <div>
              <div class="smallnote"><b>к/с</b></div>
              <input name="{prefix}_corr" value="{h(data.get('corr',''))}" style="width:100%;padding:10px;border:1px solid #D7D7D7;border-radius:12px"/>
            </div>
          </div>
          <div style="margin-top:10px">
            <div class="smallnote"><b>Адрес</b></div>
            <input name="{prefix}_address" value="{h(data.get('address',''))}" style="width:100%;padding:10px;border:1px solid #D7D7D7;border-radius:12px"/>
          </div>
        </div>
        """

    body = f"""
    <div class="pagehead">
      <div>
        <h1>{'Новый акт' if mode=='new' else 'Редактирование акта'}</h1>
        <div class="smallnote">* — обязательные поля.</div>
      </div>
      <div class="toolbar">
        <a class="btn" href="/acts">← Назад</a>
      </div>
    </div>
    <form method="POST" action="/action/acts/save">
      <input type="hidden" name="id" value="{h(init.get("id",""))}"/>
      <input type="hidden" name="mode" value="{h(mode)}"/>
      <div class="panel" style="margin-bottom:12px">
        <div style="display:flex;gap:10px;flex-wrap:wrap;align-items:flex-end">
          <div style="max-width:160px;flex:0 0 auto">
            <div class="smallnote"><b>№*</b></div>
            <input name="doc_no" value="{h(init.get("doc_no",""))}" style="width:100%;padding:10px;border:1px solid #D7D7D7;border-radius:12px" required/>
          </div>
          <div style="max-width:160px;flex:0 0 auto">
            <div class="smallnote"><b>Дата*</b></div>
            <input name="doc_date" value="{h(init.get("doc_date",""))}" placeholder="ДД.ММ.ГГГГ"
                   style="width:100%;padding:10px;border:1px solid #D7D7D7;border-radius:12px" required/>
          </div>
          <div style="flex:1;min-width:320px">
            <div class="smallnote"><b>Тип</b></div>
            <div style="display:flex;gap:14px;align-items:center;padding:10px 0">
              <label style="display:flex;gap:6px;align-items:center;margin:0">
                <input type="radio" name="direction" value="provide" {"checked" if direction=="provide" else ""}/>
                Оказание (мы исполнитель)
              </label>
              <label style="display:flex;gap:6px;align-items:center;margin:0">
                <input type="radio" name="direction" value="receive" {"checked" if direction=="receive" else ""}/>
                Получение (мы заказчик)
              </label>
            </div>
          </div>
          <div style="margin-left:auto;display:flex;gap:10px">
            <button class="btn primary" type="submit">Сохранить</button>
            <a class="btn" href="/acts">Отмена</a>
          </div>
        </div>
        <div class="hr"></div>
        <div>
          <div class="smallnote"><b>Основание</b></div>
          <input list="basisList" name="basis" value="{h(basis_cur)}"
                 style="width:100%;padding:10px;border:1px solid #D7D7D7;border-radius:12px"
                 placeholder="Например: Договор №… от …"/>
          <datalist id="basisList">{''.join(basis_opts)}</datalist>
        </div>
        <div style="margin-top:10px">
          <div class="smallnote"><b>НДС</b></div>
          <select name="vat_mode" style="width:260px;padding:10px;border:1px solid #D7D7D7;border-radius:12px">
            <option value="Без налога (НДС)" selected>Без налога (НДС)</option>
          </select>
        </div>
      </div>
      <div style="display:grid;grid-template-columns:1fr 1fr;gap:12px;margin-bottom:12px">
        {party_block("executor", "Исполнитель", init.get("executor", {}) or {})}
        {party_block("customer", "Заказчик", init.get("customer", {}) or {})}
      </div>
      <div class="panel">
        <b>Услуги</b>
        <div class="hr"></div>
        {_act_lines_table(init.get("lines", []) or [])}
        <div style="margin-top:12px;display:flex;gap:10px;justify-content:flex-end">
          <button class="btn primary" type="submit">Сохранить</button>
          <a class="btn" href="/acts">Отмена</a>
        </div>
      </div>
    </form>
    """
    return render_layout("/acts", "Акт", "Документы → Акты / накладные → Редактор", body, flash=flash_box(flash))


def page_payments(path: str, flash: str = "") -> str:
    rows = []
    for po in STATE.payment_orders:
        poid = po.get("id", "")
        amt = decimal_from_str(po.get("amount", 0))
        receiver = po.get("receiver", {}) or {}
        rows.append(f"""
        <tr>
          <td class="center">{h(po.get("number",""))}</td>
          <td class="center">{h(po.get("date",""))}</td>
          <td>{h(receiver.get("name",""))}</td>
          <td class="right">{h(fmt_money(amt)) if amt else ""}</td>
          <td>{h((po.get("purpose","") or "")[:80])}...</td>
          <td class="center" style="white-space:nowrap">
            <a class="btn small" href="/payments/edit?id={h(poid)}">Открыть</a>
            <a class="btn small" href="/action/payments/export?id={h(poid)}">Excel</a>
            <form method="POST" action="/action/payments/delete" style="display:inline;margin:0">
              <input type="hidden" name="id" value="{h(poid)}"/>
              <button class="btn small danger" type="submit">Удалить</button>
            </form>
          </td>
        </tr>
        """)
    body = f"""
    <div class="pagehead">
      <div>
        <h1>Платёжные поручения</h1>
        <div class="smallnote">Создание платёжек, экспорт Excel.</div>
      </div>
      <div class="toolbar">
        <a class="btn primary" href="/payments/new">+ Создать платёжку</a>
      </div>
    </div>
    <table class="table">
      <thead>
        <tr>
          <th style="width:90px" class="center">№</th>
          <th style="width:110px" class="center">Дата</th>
          <th style="width:260px">Получатель</th>
          <th style="width:160px" class="right">Сумма</th>
          <th>Назначение</th>
          <th style="width:220px" class="center">Действия</th>
        </tr>
      </thead>
      <tbody>
        {''.join(rows) if rows else '<tr><td colspan="6" class="muted">Платёжек нет. Создайте платёжку.</td></tr>'}
      </tbody>
    </table>
    """
    return render_layout("/payments", "Платёжки", "Документы → Платёжные поручения", body, flash=flash_box(flash))


def page_payment_editor(mode: str, qd: Dict[str, str], flash: str = "") -> str:
    if mode == "edit":
        poid = (qd.get("id") or "").strip()
        po = next((p for p in STATE.payment_orders if p.get("id") == poid), None)
        if not po:
            return render_layout("/payments", "Платёжка", "Документы → Платёжки", "<h1>Документ не найден</h1><a class='btn' href='/payments'>Назад</a>", flash=flash_box(flash))
        init = po.copy()
    else:
        init = {
            "id": new_id(),
            "number": "",
            "date": format_ddmmyyyy(date.today()),
            "amount": "0",
            "amount_words": "",
            "pay_type": "Электронно",
            "payer": {},
            "receiver": {},
            "purpose": "",
            "vid_op": "01",
            "ocher": "5",
        }
        our = STATE.get_our_company_card()
        if our:
            init["payer"] = party_from_counterparty(our)

    payer = init.get("payer", {}) or {}
    receiver = init.get("receiver", {}) or {}

    def party_block(prefix: str, title: str, data: Dict[str, Any]) -> str:
        names = counterparty_names()
        options = ['<option value=""></option>'] + [f'<option value="{h(n)}">{h(n)}</option>' for n in names]
        return f"""
        <div class="panel" style="margin-bottom:12px">
          <div style="display:flex;justify-content:space-between;align-items:center;gap:10px">
            <div><b>{h(title)}</b></div>
            <div style="display:flex;gap:8px;align-items:center">
              <select name="{prefix}_card" style="max-width:260px">{''.join(options)}</select>
              <button class="btn small" name="_apply_card" value="{prefix}" type="submit">Подставить</button>
            </div>
          </div>
          <div class="hr"></div>
          <div class="smallnote"><b>Наименование</b></div>
          <input name="{prefix}_name" value="{h(data.get('name',''))}" style="width:100%;padding:10px;border:1px solid #D7D7D7;border-radius:12px"/>
          <div style="display:grid;grid-template-columns:1fr 1fr;gap:12px;margin-top:10px">
            <div><div class="smallnote"><b>ИНН</b></div><input name="{prefix}_inn" value="{h(data.get('inn',''))}" style="width:100%;padding:10px;border:1px solid #D7D7D7;border-radius:12px"/></div>
            <div><div class="smallnote"><b>КПП</b></div><input name="{prefix}_kpp" value="{h(data.get('kpp',''))}" style="width:100%;padding:10px;border:1px solid #D7D7D7;border-radius:12px"/></div>
          </div>
          <div style="margin-top:10px"><div class="smallnote"><b>р/с</b></div><input name="{prefix}_account" value="{h(data.get('account',''))}" style="width:100%;padding:10px;border:1px solid #D7D7D7;border-radius:12px"/></div>
          <div style="margin-top:10px"><div class="smallnote"><b>Банк</b></div><input name="{prefix}_bank" value="{h(data.get('bank',''))}" style="width:100%;padding:10px;border:1px solid #D7D7D7;border-radius:12px"/></div>
          <div style="display:grid;grid-template-columns:1fr 1fr;gap:12px;margin-top:10px">
            <div><div class="smallnote"><b>БИК</b></div><input name="{prefix}_bik" value="{h(data.get('bik',''))}" style="width:100%;padding:10px;border:1px solid #D7D7D7;border-radius:12px"/></div>
            <div><div class="smallnote"><b>к/с</b></div><input name="{prefix}_corr" value="{h(data.get('corr',''))}" style="width:100%;padding:10px;border:1px solid #D7D7D7;border-radius:12px"/></div>
          </div>
        </div>
        """

    body = f"""
    <div class="pagehead">
      <div>
        <h1>{'Новая платёжка' if mode=='new' else 'Редактирование платёжки'}</h1>
      </div>
      <div class="toolbar">
        <a class="btn" href="/payments">← Назад</a>
      </div>
    </div>
    <form method="POST" action="/action/payments/save">
      <input type="hidden" name="id" value="{h(init.get("id",""))}"/>
      <input type="hidden" name="mode" value="{h(mode)}"/>
      <div class="panel" style="margin-bottom:12px">
        <div style="display:flex;gap:10px;flex-wrap:wrap;align-items:flex-end">
          <div style="max-width:140px"><div class="smallnote"><b>№</b></div><input name="number" value="{h(init.get("number",""))}" style="width:100%;padding:10px;border:1px solid #D7D7D7;border-radius:12px"/></div>
          <div style="max-width:160px"><div class="smallnote"><b>Дата</b></div><input name="date" value="{h(init.get("date",""))}" placeholder="ДД.ММ.ГГГГ" style="width:100%;padding:10px;border:1px solid #D7D7D7;border-radius:12px"/></div>
          <div style="max-width:200px"><div class="smallnote"><b>Сумма</b></div><input name="amount" value="{h(str(init.get("amount","0")))}" style="width:100%;padding:10px;border:1px solid #D7D7D7;border-radius:12px"/></div>
          <div style="max-width:180px"><div class="smallnote"><b>Вид платежа</b></div>
            <select name="pay_type" style="width:100%;padding:10px;border:1px solid #D7D7D7;border-radius:12px">
              <option value="Электронно" {"selected" if init.get("pay_type")=="Электронно" else ""}>Электронно</option>
              <option value="Почтой" {"selected" if init.get("pay_type")=="Почтой" else ""}>Почтой</option>
              <option value="Телеграфом" {"selected" if init.get("pay_type")=="Телеграфом" else ""}>Телеграфом</option>
            </select>
          </div>
          <div style="margin-left:auto;display:flex;gap:10px">
            <button class="btn primary" type="submit">Сохранить</button>
            <a class="btn" href="/payments">Отмена</a>
          </div>
        </div>
        <div class="hr"></div>
        <div><div class="smallnote"><b>Сумма прописью</b></div><input name="amount_words" value="{h(init.get("amount_words",""))}" style="width:100%;padding:10px;border:1px solid #D7D7D7;border-radius:12px"/></div>
        <div style="display:grid;grid-template-columns:1fr 1fr;gap:12px;margin-top:10px">
          <div><div class="smallnote"><b>Вид оп.</b></div><input name="vid_op" value="{h(init.get("vid_op","01"))}" style="width:100%;padding:10px;border:1px solid #D7D7D7;border-radius:12px"/></div>
          <div><div class="smallnote"><b>Очер. плат.</b></div><input name="ocher" value="{h(init.get("ocher","5"))}" style="width:100%;padding:10px;border:1px solid #D7D7D7;border-radius:12px"/></div>
        </div>
      </div>
      {party_block("payer", "Плательщик", payer)}
      {party_block("receiver", "Получатель", receiver)}
      <div class="panel">
        <div class="smallnote"><b>Назначение платежа</b></div>
        <textarea name="purpose" style="width:100%;min-height:100px;padding:10px;border:1px solid #D7D7D7;border-radius:12px">{h(init.get("purpose",""))}</textarea>
        <div style="margin-top:12px;display:flex;gap:10px;justify-content:flex-end">
          <button class="btn primary" type="submit">Сохранить</button>
          <a class="btn" href="/payments">Отмена</a>
        </div>
      </div>
    </form>
    """
    return render_layout("/payments", "Платёжка", "Документы → Платёжки → Редактор", body, flash=flash_box(flash))


def page_salary(path: str, flash: str = "") -> str:
    emp_rows = []
    for e in STATE.employees:
        eid = e.get("id", "")
        name_inn = e.get("name", "")
        if e.get("inn"):
            name_inn += f" <span class='muted'>ИНН: {h(e.get('inn',''))}</span>"
        salary = decimal_from_str(e.get("salary") or 0)
        advance = decimal_from_str(e.get("advance") or 0)
        main_part = decimal_from_str(e.get("main_part") or e.get("main") or 0)
        emp_rows.append(f"""
        <tr>
          <td>{name_inn}</td>
          <td class="right">{h(fmt_money(salary)) if salary else ""}</td>
          <td class="right">{h(fmt_money(advance)) if advance else ""}</td>
          <td class="right">{h(fmt_money(main_part)) if main_part else ""}</td>
          <td class="center" style="white-space:nowrap">
            <a class="btn small" href="/salary/emp-edit?id={h(eid)}">Изменить</a>
            <form method="POST" action="/action/salary/emp-delete" style="display:inline;margin:0">
              <input type="hidden" name="id" value="{h(eid)}"/>
              <button class="btn small danger" type="submit">Удалить</button>
            </form>
          </td>
        </tr>
        """)

    pay_rows = []
    for sp in STATE.salary_payments:
        emp = next((e for e in STATE.employees if e.get("id") == sp.get("employee_id")), None)
        emp_name = emp.get("name", "") if emp else "—"
        amt = decimal_from_str(sp.get("amount") or 0)
        pay_type = sp.get("pay_type", "")
        type_label = "Аванс" if pay_type == "advance" else "Основная" if pay_type == "main" else pay_type
        po_id = sp.get("payment_order_id", "")
        po_link = ""
        if po_id:
            po = next((p for p in STATE.payment_orders if p.get("id") == po_id), None)
            if po:
                po_num = po.get("number", "") or "б/н"
                po_link = f"<a href='/payments/edit?id={h(po_id)}'>{h(po_num)}</a>"
        pay_rows.append(f"""
        <tr>
          <td style="width:160px">{h(sp.get("created_at_str",""))}</td>
          <td class="center" style="width:110px">{h(sp.get("month",""))}</td>
          <td>{h(emp_name)}</td>
          <td style="width:140px">{h(type_label)}</td>
          <td class="right" style="width:140px">{h(fmt_money(amt)) if amt else ""}</td>
          <td class="center" style="width:110px">{po_link}</td>
          <td class="center" style="width:200px">
            <form method="POST" action="/action/salary/pay-delete" style="display:inline;margin:0">
              <input type="hidden" name="id" value="{h(sp.get('id',''))}"/>
              <button class="btn small danger" type="submit">Удалить</button>
            </form>
          </td>
        </tr>
        """)

    emp_options = "".join([f'<option value="{h(e.get("id",""))}">{h(e.get("name",""))}</option>' for e in STATE.employees])
    current_month = date.today().strftime("%m.%Y")

    body = f"""
    <div class="pagehead">
      <div>
        <h1>Зарплата</h1>
        <div class="smallnote">Сотрудники и выплаты. Автоматическое создание платёжных поручений.</div>
      </div>
      <div class="toolbar">
        <a class="btn primary" href="/salary/emp-new">+ Сотрудник</a>
      </div>
    </div>
    <div style="display:grid;grid-template-columns:1fr 1fr;gap:12px">
      <div>
        <div class="panel" style="margin-bottom:10px">
          <b>Сотрудники</b>
          <div class="smallnote">Настройте суммы аванса/основной части в карточке сотрудника.</div>
        </div>
        <table class="table">
          <thead>
            <tr>
              <th>ФИО / ИНН</th>
              <th class="right" style="width:140px">Оклад</th>
              <th class="right" style="width:140px">Аванс</th>
              <th class="right" style="width:160px">Основная</th>
              <th class="center" style="width:260px">Действия</th>
            </tr>
          </thead>
          <tbody>
            {''.join(emp_rows) if emp_rows else '<tr><td colspan="5" class="muted">Сотрудников нет.</td></tr>'}
          </tbody>
        </table>
        <div class="panel" style="margin-top:12px">
          <b>Создать выплату</b>
          <form method="POST" action="/action/salary/create-pay" style="display:flex;gap:10px;flex-wrap:wrap;align-items:flex-end;margin:0;margin-top:10px">
            <div style="flex:1;min-width:260px">
              <div class="smallnote"><b>Сотрудник</b></div>
              <select name="employee_id" style="width:100%;padding:10px;border:1px solid #D7D7D7;border-radius:12px" required>
                <option value=""></option>
                {emp_options}
              </select>
            </div>
            <div style="max-width:140px">
              <div class="smallnote"><b>Месяц</b></div>
              <input name="month" value="{h(current_month)}" style="width:100%;padding:10px;border:1px solid #D7D7D7;border-radius:12px" required/>
            </div>
            <div style="max-width:220px">
              <div class="smallnote"><b>Тип выплаты</b></div>
              <select name="type" style="width:100%;padding:10px;border:1px solid #D7D7D7;border-radius:12px">
                <option value="advance">Аванс</option>
                <option value="main">Основная часть</option>
              </select>
            </div>
            <button class="btn primary" type="submit">Создать выплату</button>
          </form>
        </div>
      </div>
      <div>
        <div class="panel" style="margin-bottom:10px">
          <b>Выплаты (журнал)</b>
          <div class="smallnote">Удаление выплаты удаляет и связанную платёжку.</div>
        </div>
        <table class="table">
          <thead>
            <tr>
              <th style="width:160px">Создано</th>
              <th class="center" style="width:110px">Месяц</th>
              <th>Сотрудник</th>
              <th style="width:140px">Тип</th>
              <th class="right" style="width:140px">Сумма</th>
              <th class="center" style="width:110px">П/п №</th>
              <th class="center" style="width:200px">Действия</th>
            </tr>
          </thead>
          <tbody>
            {''.join(pay_rows) if pay_rows else '<tr><td colspan="7" class="muted">Выплат нет.</td></tr>'}
          </tbody>
        </table>
      </div>
    </div>
    """
    return render_layout("/salary", "Зарплата", "Документы → Зарплата", body, flash=flash_box(flash))


def page_employee_editor(mode: str, qd: Dict[str, str], flash: str = "") -> str:
    if mode == "edit":
        eid = (qd.get("id") or "").strip()
        emp = STATE.get_employee_by_id(eid)
        if not emp:
            return render_layout("/salary", "Сотрудник", "Документы → Зарплата", "<h1>Сотрудник не найден</h1><a class='btn' href='/salary'>Назад</a>", flash=flash_box(flash))
        init = emp.copy()
    else:
        init = {"id": new_id(), "name": "", "inn": "", "passport": "", "passport_issued": "", "bank": "", "bik": "", "corr": "", "account": "", "salary": "", "advance": "", "main": ""}

    body = f"""
    <div class="pagehead">
      <div>
        <h1>{'Новый сотрудник' if mode=='new' else 'Редактирование сотрудника'}</h1>
        <div class="smallnote">ФИО обязательно. Если «Основная часть» пустая — будет посчитана как Оклад - Аванс.</div>
      </div>
      <div class="toolbar">
        <a class="btn" href="/salary">← Назад</a>
      </div>
    </div>
    <form method="POST" action="/action/salary/emp-save">
      <input type="hidden" name="id" value="{h(init.get("id",""))}"/>
      <input type="hidden" name="mode" value="{h(mode)}"/>
      <div class="panel">
        <div style="display:grid;grid-template-columns:1fr 1fr;gap:12px">
          <div>
            <div class="smallnote"><b>ФИО*</b></div>
            <input name="name" value="{h(init.get("name",""))}" style="width:100%;padding:10px;border:1px solid #D7D7D7;border-radius:12px" required/>
          </div>
          <div>
            <div class="smallnote"><b>ИНН (опц.)</b></div>
            <input name="inn" value="{h(init.get("inn") or "")}" style="width:100%;padding:10px;border:1px solid #D7D7D7;border-radius:12px"/>
          </div>
        </div>
        <div class="hr"></div>
        <b>Паспортные данные (опционально)</b>
        <div style="display:grid;grid-template-columns:1fr 1fr;gap:12px;margin-top:10px">
          <div>
            <div class="smallnote"><b>Паспорт</b></div>
            <input name="passport" value="{h(init.get("passport") or "")}" style="width:100%;padding:10px;border:1px solid #D7D7D7;border-radius:12px"/>
          </div>
          <div>
            <div class="smallnote"><b>Кем/когда выдан</b></div>
            <input name="passport_issued" value="{h(init.get("passport_issued") or "")}" style="width:100%;padding:10px;border:1px solid #D7D7D7;border-radius:12px"/>
          </div>
        </div>
        <div class="hr"></div>
        <b>Банковские реквизиты</b>
        <div style="margin-top:10px">
          <div class="smallnote"><b>Банк</b></div>
          <input name="bank" value="{h(init.get("bank") or "")}" style="width:100%;padding:10px;border:1px solid #D7D7D7;border-radius:12px"/>
        </div>
        <div style="display:grid;grid-template-columns:1fr 1fr;gap:12px;margin-top:10px">
          <div><div class="smallnote"><b>БИК</b></div><input name="bik" value="{h(init.get("bik") or "")}" style="width:100%;padding:10px;border:1px solid #D7D7D7;border-radius:12px"/></div>
          <div><div class="smallnote"><b>к/с</b></div><input name="corr" value="{h(init.get("corr") or "")}" style="width:100%;padding:10px;border:1px solid #D7D7D7;border-radius:12px"/></div>
        </div>
        <div style="margin-top:10px">
          <div class="smallnote"><b>Счёт №</b></div>
          <input name="account" value="{h(init.get("account") or "")}" style="width:100%;padding:10px;border:1px solid #D7D7D7;border-radius:12px"/>
        </div>
        <div class="hr"></div>
        <b>Условия оплаты</b>
        <div style="display:grid;grid-template-columns:1fr 1fr 1fr;gap:12px;margin-top:10px">
          <div><div class="smallnote"><b>Оклад</b></div><input name="salary" value="{h(str(init.get("salary") or ""))}" style="width:100%;padding:10px;border:1px solid #D7D7D7;border-radius:12px"/></div>
          <div><div class="smallnote"><b>Аванс</b></div><input name="advance" value="{h(str(init.get("advance") or ""))}" style="width:100%;padding:10px;border:1px solid #D7D7D7;border-radius:12px"/></div>
          <div><div class="smallnote"><b>Основная часть</b></div><input name="main" value="{h(str(init.get("main_part") or init.get("main") or ""))}" style="width:100%;padding:10px;border:1px solid #D7D7D7;border-radius:12px"/></div>
        </div>
        <div style="margin-top:12px;display:flex;gap:10px">
          <button class="btn primary" type="submit">Сохранить</button>
          <a class="btn" href="/salary">Отмена</a>
        </div>
      </div>
    </form>
    """
    return render_layout("/salary", "Сотрудник", "Документы → Зарплата → Сотрудник", body, flash=flash_box(flash))


def parse_upd_html(path: str) -> Dict[str, Any]:
    if BeautifulSoup is None:
        raise RuntimeError("Нужен beautifulsoup4: pip install beautifulsoup4")
    filename = os.path.basename(path)
    m_num = re.search(r"(\d+)", filename)
    doc_no = m_num.group(1) if m_num else ""
    m_date = re.search(r"(\d{8})", filename)
    yyyymmdd = m_date.group(1) if m_date else ""
    doc_date = f"{yyyymmdd[6:8]}.{yyyymmdd[4:6]}.{yyyymmdd[0:4]}" if len(yyyymmdd) == 8 else ""
    with open(path, "r", encoding="utf-8", errors="replace") as f:
        html = f.read()
    soup = BeautifulSoup(html, "html.parser")
    total = Decimal("0")
    for tbl in soup.find_all("table"):
        txt_ = tbl.get_text(separator=" ", strip=True)
        if "Всего к оплате" in txt_:
            nums = re.findall(r"[\d\s]+[,.][\d]+", txt_)
            if nums:
                total = decimal_from_str(nums[-1])
                break
    return {
        "id": new_id(),
        "doc_no": doc_no,
        "doc_date": doc_date,
        "counterparty": "",
        "inn": "",
        "amount": str(total),
        "vat": "0",
        "description": "",
        "source_file": filename,
    }


def page_upd(path: str, flash: str = "") -> str:
    rows = []
    for r in STATE.upd_rows:
        rid = r.get("id", "")
        amt = decimal_from_str(r.get("amount") or 0)
        rows.append(f"""
        <tr>
          <td class="center">{h(r.get("doc_no",""))}</td>
          <td class="center">{h(r.get("doc_date",""))}</td>
          <td class="right">{h(fmt_money(amt)) if amt else ""}</td>
          <td>{h(r.get("source_file",""))}</td>
          <td class="center">
            <form method="POST" action="/action/upd/delete" style="display:inline;margin:0">
              <input type="hidden" name="id" value="{h(rid)}"/>
              <button class="btn small danger" type="submit">Удалить</button>
            </form>
          </td>
        </tr>
        """)
    body = f"""
    <div class="pagehead">
      <div>
        <h1>УПД (HTML)</h1>
        <div class="smallnote">Загрузка HTML файлов УПД, парсинг сумм.</div>
      </div>
      <div class="toolbar">
        <form id="updUploadForm" method="POST" action="/action/upd/upload" enctype="multipart/form-data" style="margin:0">
          <input class="file-hidden" id="updFile" type="file" name="file" accept=".html,.htm" required/>
          <button class="btn primary" type="button" onclick="pickAndSubmit('updFile','updUploadForm')">Загрузить УПД</button>
        </form>
      </div>
    </div>
    <script>function pickAndSubmit(inputId, formId){{const inp = document.getElementById(inputId); const form = document.getElementById(formId); inp.onchange = ()=>{{ if(inp.files && inp.files.length){{ form.submit(); }} }}; inp.click();}}</script>
    <table class="table">
      <thead><tr><th style="width:90px" class="center">№</th><th style="width:110px" class="center">Дата</th><th style="width:160px" class="right">Сумма</th><th>Файл</th><th style="width:120px" class="center">Действия</th></tr></thead>
      <tbody>{''.join(rows) if rows else '<tr><td colspan="5" class="muted">УПД нет. Загрузите HTML файл.</td></tr>'}</tbody>
    </table>
    """
    return render_layout("/upd", "УПД", "Документы → УПД (HTML)", body, flash=flash_box(flash))


def page_realization(path: str, flash: str = "") -> str:
    rows = []
    for r in STATE.real_rows:
        amt = decimal_from_str(r.get("amount") or 0)
        rows.append(f"""
        <tr>
          <td class="center">{h(r.get("doc_no",""))}</td>
          <td class="center">{h(r.get("doc_date",""))}</td>
          <td>{h(r.get("counterparty",""))}</td>
          <td class="right">{h(fmt_money(amt)) if amt else ""}</td>
          <td>{h(r.get("description",""))}</td>
        </tr>
        """)
    body = f"""
    <div class="pagehead">
      <div>
        <h1>Передача на реализацию</h1>
        <div class="smallnote">Автоматически формируется на основе УПД.</div>
      </div>
    </div>
    <table class="table">
      <thead><tr><th style="width:90px" class="center">№</th><th style="width:110px" class="center">Дата</th><th>Контрагент</th><th style="width:160px" class="right">Сумма</th><th>Описание</th></tr></thead>
      <tbody>{''.join(rows) if rows else '<tr><td colspan="5" class="muted">Нет данных о реализации.</td></tr>'}</tbody>
    </table>
    """
    return render_layout("/realization", "Реализация", "Документы → Передача на реализацию", body, flash=flash_box(flash))


def calc_op_profit(start: Optional[date], end: Optional[date]) -> Tuple[str, List[Dict[str, Any]], Dict[str, Decimal]]:
    categories: Dict[str, Dict[str, Decimal]] = {}
    for r in STATE.bank_rows:
        d = parse_date_ddmmyyyy(r.get("date", ""))
        if not in_range(d, start, end):
            continue
        inc = Decimal(str(r.get("incoming") or 0))
        out = Decimal(str(r.get("outgoing") or 0))
        if out > 0 and bool(r.get("skip_outgoing", False)):
            continue
        cat = norm_category(r.get("category") or "")
        if cat not in categories:
            categories[cat] = {"income": Decimal("0"), "outgoing": Decimal("0")}
        if inc > 0:
            categories[cat]["income"] += inc
        if out > 0:
            categories[cat]["outgoing"] += out
    rows = []
    total_income = Decimal("0")
    total_outgoing = Decimal("0")
    for cat, vals in sorted(categories.items(), key=lambda x: x[1]["income"] - x[1]["outgoing"], reverse=True):
        net = vals["income"] - vals["outgoing"]
        total_income += vals["income"]
        total_outgoing += vals["outgoing"]
        rows.append({"category": cat, "income": vals["income"], "outgoing": vals["outgoing"], "net": net})
    title = f"{format_ddmmyyyy(start)} – {format_ddmmyyyy(end)}" if start and end else "За всё время"
    totals = {"income": total_income, "outgoing": total_outgoing, "net": total_income - total_outgoing}
    return title, rows, totals


def page_op_profit(path: str, qd: Dict[str, str], flash: str = "") -> str:
    mode = (qd.get("mode") or "all").strip()
    dfrom = (qd.get("from") or "").strip()
    dto = (qd.get("to") or "").strip()
    try:
        start, end, title = compute_date_range(mode, dfrom, dto)
    except Exception as e:
        start, end, title = None, None, "За всё время"
        flash = str(e)
    _, rows, totals = calc_op_profit(start, end)
    tr = []
    for r in rows:
        tr.append(f"""
        <tr>
          <td><a href="/reports/op-profit/details?category={h(urlencode({"c": r["category"]})[2:])}&mode={h(mode)}&from={h(dfrom)}&to={h(dto)}">{h(r["category"])}</a></td>
          <td class="right">{h(fmt_money(r["income"]))}</td>
          <td class="right">{h(fmt_money(r["outgoing"]))}</td>
          <td class="right"><b>{h(fmt_money(r["net"]))}</b></td>
        </tr>
        """)
    body = f"""
    <div class="pagehead">
      <div>
        <h1>Операционная прибыль</h1>
        <div class="smallnote">Группировка банковских операций по статьям.</div>
      </div>
    </div>
    <form method="GET" action="/reports/op-profit" class="panel" style="margin-bottom:12px">
      {_period_form_html("/reports/op-profit", mode, dfrom, dto)}
    </form>
    <div class="panel" style="margin-bottom:12px;display:flex;gap:12px;flex-wrap:wrap">
      <span class="badge2"><b>{h(title)}</b></span>
      <span class="badge2"><b>Поступления:</b>&nbsp;{h(fmt_money(totals["income"]))}</span>
      <span class="badge2"><b>Списания:</b>&nbsp;{h(fmt_money(totals["outgoing"]))}</span>
      <span class="badge2"><b>Прибыль:</b>&nbsp;{h(fmt_money(totals["net"]))}</span>
    </div>
    <table class="table">
      <thead><tr><th>Статья</th><th style="width:180px" class="right">Поступления</th><th style="width:180px" class="right">Списания</th><th style="width:180px" class="right">Итого</th></tr></thead>
      <tbody>{''.join(tr) if tr else '<tr><td colspan="4" class="muted">Нет данных.</td></tr>'}</tbody>
      <tfoot><tr><td>ИТОГО</td><td class="right">{h(fmt_money(totals["income"]))}</td><td class="right">{h(fmt_money(totals["outgoing"]))}</td><td class="right"><b>{h(fmt_money(totals["net"]))}</b></td></tr></tfoot>
    </table>
    """
    return render_layout("/reports/op-profit", "Операц. прибыль", "Отчёты → Операц. прибыль", body, flash=flash_box(flash))


def page_op_profit_details(path: str, qd: Dict[str, str], flash: str = "") -> str:
    cat_raw = parse_qs(qd.get("category", "")).get("c", [""])[0] if "category" in qd else ""
    cat = norm_category(cat_raw)
    mode = qd.get("mode", "all")
    dfrom = qd.get("from", "")
    dto = qd.get("to", "")
    try:
        start, end, _ = compute_date_range(mode, dfrom, dto)
    except:
        start, end = None, None
    ops = []
    for r in STATE.bank_rows:
        if norm_category(r.get("category") or "") != cat:
            continue
        d = parse_date_ddmmyyyy(r.get("date", ""))
        if not in_range(d, start, end):
            continue
        inc = Decimal(str(r.get("incoming") or 0))
        out = Decimal(str(r.get("outgoing") or 0))
        if out > 0 and bool(r.get("skip_outgoing", False)):
            continue
        ops.append(r)
    tr = []
    for r in ops:
        inc = Decimal(str(r.get("incoming") or 0))
        out = Decimal(str(r.get("outgoing") or 0))
        tr.append(f"""
        <tr>
          <td class="center">{h(r.get("date",""))}</td>
          <td class="right">{h(fmt_money(inc)) if inc else ""}</td>
          <td class="right">{h(fmt_money(out)) if out else ""}</td>
          <td>{h((r.get("purpose","") or "")[:60])}...</td>
          <td>{h(r.get("counterparty",""))}</td>
        </tr>
        """)
    body = f"""
    <div class="pagehead">
      <div><h1>Детализация: {h(cat)}</h1></div>
      <div class="toolbar"><a class="btn" href="/reports/op-profit?mode={h(mode)}&from={h(dfrom)}&to={h(dto)}">← Назад</a></div>
    </div>
    <table class="table">
      <thead><tr><th style="width:110px" class="center">Дата</th><th style="width:150px" class="right">Поступление</th><th style="width:150px" class="right">Списание</th><th>Назначение</th><th style="width:260px">Контрагент</th></tr></thead>
      <tbody>{''.join(tr) if tr else '<tr><td colspan="5" class="muted">Нет операций.</td></tr>'}</tbody>
    </table>
    """
    return render_layout("/reports/op-profit", "Детализация", "Отчёты → Операц. прибыль → Детализация", body, flash=flash_box(flash))


def calc_recon(counterparty_name: str, start: date, end: date) -> Tuple[str, List[Dict[str, Any]], Dict[str, Any]]:
    cp_norm = norm_text(counterparty_name)
    rows: List[Dict[str, Any]] = []
    debit_total = Decimal("0")
    credit_total = Decimal("0")
    for r in STATE.bank_rows:
        d = parse_date_ddmmyyyy(r.get("date", ""))
        if not in_range(d, start, end):
            continue
        cp = norm_text(r.get("counterparty", ""))
        if not cp or not (cp_norm in cp or cp in cp_norm):
            continue
        inc = Decimal(str(r.get("incoming") or 0))
        out = Decimal(str(r.get("outgoing") or 0))
        if out > 0 and bool(r.get("skip_outgoing", False)):
            continue
        if inc > 0:
            credit_total += inc
            rows.append({"date": r.get("date", ""), "source": "Банк", "doc": f"Поступление ({r.get('doctype','')})", "debit": "", "credit": float(inc), "comment": r.get("purpose", "")})
        if out > 0:
            debit_total += out
            rows.append({"date": r.get("date", ""), "source": "Банк", "doc": f"Списание ({r.get('doctype','')})", "debit": float(out), "credit": "", "comment": r.get("purpose", "")})
    rows.sort(key=lambda x: parse_date_ddmmyyyy(x.get("date", "")) or date.min)
    title = f"{counterparty_name}. Период: {format_ddmmyyyy(start)} – {format_ddmmyyyy(end)}"
    totals = {"debit": float(debit_total), "credit": float(credit_total)}
    return title, rows, totals


def page_recon(path: str, qd: Dict[str, str], flash: str = "") -> str:
    name = (qd.get("cp") or "").strip()
    dfrom = (qd.get("from") or "").strip()
    dto = (qd.get("to") or "").strip()
    if not dfrom or not dto:
        dates = [parse_date_ddmmyyyy(r.get("date", "")) for r in STATE.bank_rows]
        dates = [d for d in dates if d]
        if dates:
            dfrom = dfrom or format_ddmmyyyy(min(dates))
            dto = dto or format_ddmmyyyy(max(dates))
    rows: List[Dict[str, Any]] = []
    title = "—"
    totals = {"debit": 0.0, "credit": 0.0}
    if name and dfrom and dto:
        d1 = parse_date_ddmmyyyy(dfrom)
        d2 = parse_date_ddmmyyyy(dto)
        if not d1 or not d2:
            flash = "Неверные даты."
        else:
            title, rows, totals = calc_recon(name, d1, d2)
    cp_opts = ['<option value=""></option>'] + [f'<option value="{h(n)}" {"selected" if n==name else ""}>{h(n)}</option>' for n in counterparty_names()]
    tr = []
    for r in rows:
        debit = r.get("debit")
        credit = r.get("credit")
        tr.append(f"""
        <tr>
          <td class="center">{h(r.get("date",""))}</td>
          <td>{h(r.get("source",""))}</td>
          <td>{h(r.get("doc",""))}</td>
          <td class="right">{h(fmt_money(debit)) if isinstance(debit,(int,float,Decimal)) and debit != "" else ""}</td>
          <td class="right">{h(fmt_money(credit)) if isinstance(credit,(int,float,Decimal)) and credit != "" else ""}</td>
          <td>{h((r.get("comment","") or "")[:50])}...</td>
        </tr>
        """)
    body = f"""
    <div class="pagehead">
      <div><h1>Акт сверки</h1><div class="smallnote">Учитывает банковские операции.</div></div>
    </div>
    <form method="GET" action="/reports/recon" class="panel" style="margin-bottom:12px">
      <div style="display:flex;gap:12px;flex-wrap:wrap;align-items:flex-end">
        <div style="min-width:360px;flex:1"><div class="smallnote"><b>Контрагент</b></div><select name="cp" style="width:100%;padding:10px;border:1px solid #D7D7D7;border-radius:12px" required>{''.join(cp_opts)}</select></div>
        <div style="max-width:170px"><div class="smallnote"><b>с</b></div><input name="from" value="{h(dfrom)}" placeholder="ДД.ММ.ГГГГ" style="width:100%;padding:10px;border:1px solid #D7D7D7;border-radius:12px" required/></div>
        <div style="max-width:170px"><div class="smallnote"><b>по</b></div><input name="to" value="{h(dto)}" placeholder="ДД.ММ.ГГГГ" style="width:100%;padding:10px;border:1px solid #D7D7D7;border-radius:12px" required/></div>
        <button class="btn primary" type="submit">Сформировать</button>
      </div>
    </form>
    <div class="panel" style="margin-bottom:12px;display:flex;gap:12px;flex-wrap:wrap">
      <span class="badge2"><b>{h(title)}</b></span>
      <span class="badge2"><b>Дебет:</b>&nbsp;{h(fmt_money(totals["debit"]))}</span>
      <span class="badge2"><b>Кредит:</b>&nbsp;{h(fmt_money(totals["credit"]))}</span>
    </div>
    <table class="table">
      <thead><tr><th class="center" style="width:110px">Дата</th><th style="width:110px">Источник</th><th>Документ</th><th class="right" style="width:170px">Дебет</th><th class="right" style="width:170px">Кредит</th><th style="width:200px">Комментарий</th></tr></thead>
      <tbody>{''.join(tr) if tr else '<tr><td colspan="6" class="muted">Выберите контрагента и период.</td></tr>'}</tbody>
    </table>
    """
    return render_layout("/reports/recon", "Акт сверки", "Отчёты → Акт сверки", body, flash=flash_box(flash))


def calc_kudir(start: date, end: date) -> Tuple[str, List[Dict[str, Any]], Dict[str, Decimal]]:
    rows: List[Dict[str, Any]] = []
    income_total = Decimal("0")
    expense_total = Decimal("0")
    for r in STATE.bank_rows:
        d = parse_date_ddmmyyyy(r.get("date", ""))
        if not in_range(d, start, end):
            continue
        inc = Decimal(str(r.get("incoming") or 0))
        out = Decimal(str(r.get("outgoing") or 0))
        if out > 0 and bool(r.get("skip_outgoing", False)):
            continue
        if inc == 0 and out == 0:
            continue
        content = f"{r.get('counterparty','')}. {r.get('purpose','')}".strip()
        row = {"doc_ref": r.get("date", ""), "content": content, "income": "", "expense": ""}
        if inc > 0:
            row["income"] = float(money2(inc))
            income_total += inc
        if out > 0:
            row["expense"] = float(money2(out))
            expense_total += out
        rows.append(row)
    title = f"Период: {format_ddmmyyyy(start)} – {format_ddmmyyyy(end)}"
    totals = {"income": money2(income_total), "expense": money2(expense_total)}
    return title, rows, totals


def page_kudir(path: str, qd: Dict[str, str], flash: str = "") -> str:
    dfrom = (qd.get("from") or "").strip()
    dto = (qd.get("to") or "").strip()
    if not dfrom or not dto:
        dates = [parse_date_ddmmyyyy(r.get("date", "")) for r in STATE.bank_rows]
        dates = [d for d in dates if d]
        if dates:
            dfrom = dfrom or format_ddmmyyyy(min(dates))
            dto = dto or format_ddmmyyyy(max(dates))
    title = "—"
    rows: List[Dict[str, Any]] = []
    totals = {"income": Decimal("0"), "expense": Decimal("0")}
    d1 = parse_date_ddmmyyyy(dfrom) if dfrom else None
    d2 = parse_date_ddmmyyyy(dto) if dto else None
    if d1 and d2 and d2 >= d1:
        title, rows, totals = calc_kudir(d1, d2)
    elif dfrom or dto:
        flash = "Введите даты в формате ДД.ММ.ГГГГ."
    tr = []
    for r in rows[:600]:
        inc = r.get("income", "")
        exp = r.get("expense", "")
        tr.append(f"""
        <tr>
          <td style="width:160px">{h(r.get("doc_ref",""))}</td>
          <td>{h((r.get("content","") or "")[:80])}...</td>
          <td class="right" style="width:160px">{h(fmt_money(inc)) if isinstance(inc,(int,float,Decimal)) and inc != "" else ""}</td>
          <td class="right" style="width:160px">{h(fmt_money(exp)) if isinstance(exp,(int,float,Decimal)) and exp != "" else ""}</td>
        </tr>
        """)
    body = f"""
    <div class="pagehead">
      <div><h1>КУДиР (PDF)</h1><div class="smallnote">Предпросмотр и генерация PDF.</div></div>
      <div class="toolbar"><a class="btn primary" href="/action/reports/kudir-pdf?from={h(dfrom)}&to={h(dto)}">Скачать PDF</a></div>
    </div>
    <form method="GET" action="/reports/kudir" class="panel" style="margin-bottom:12px">
      <div style="display:flex;gap:12px;flex-wrap:wrap;align-items:flex-end">
        <div style="max-width:180px"><div class="smallnote"><b>с</b></div><input name="from" value="{h(dfrom)}" placeholder="ДД.ММ.ГГГГ" style="width:100%;padding:10px;border:1px solid #D7D7D7;border-radius:12px" required/></div>
        <div style="max-width:180px"><div class="smallnote"><b>по</b></div><input name="to" value="{h(dto)}" placeholder="ДД.ММ.ГГГГ" style="width:100%;padding:10px;border:1px solid #D7D7D7;border-radius:12px" required/></div>
        <button class="btn" type="submit">Предпросмотр</button>
        <span class="badge2"><b>Доход:</b>&nbsp;{h(fmt_money(totals["income"]))}</span>
        <span class="badge2"><b>Расход:</b>&nbsp;{h(fmt_money(totals["expense"]))}</span>
        <span class="badge2"><b>Разница:</b>&nbsp;{h(fmt_money(money2(totals["income"]-totals["expense"])))}</span>
      </div>
    </form>
    <table class="table">
      <thead><tr><th style="width:160px">Дата</th><th>Содержание</th><th class="right" style="width:160px">Доход</th><th class="right" style="width:160px">Расход</th></tr></thead>
      <tbody>{''.join(tr) if tr else '<tr><td colspan="4" class="muted">Выберите период.</td></tr>'}</tbody>
    </table>
    <div class="smallnote">Показаны первые 600 строк.</div>
    """
    return render_layout("/reports/kudir", "КУДиР", "Отчёты → КУДиР (PDF)", body, flash=flash_box(flash))


def _wrap_text(text: str, max_width: float, font_name: str, font_size: float) -> List[str]:
    """Разбивает текст на строки, чтобы он помещался в заданную ширину"""
    require_reportlab()
    from reportlab.pdfbase.pdfmetrics import stringWidth
    words = text.split()
    lines = []
    current_line = ""
    for word in words:
        test_line = (current_line + " " + word).strip() if current_line else word
        w = stringWidth(test_line, font_name, font_size)
        if w <= max_width and current_line:
            current_line = test_line
        elif w <= max_width:
            current_line = test_line
        else:
            if current_line:
                lines.append(current_line)
            if stringWidth(word, font_name, font_size) > max_width:
                while word:
                    for ci in range(len(word), 0, -1):
                        if stringWidth(word[:ci], font_name, font_size) <= max_width:
                            lines.append(word[:ci])
                            word = word[ci:]
                            break
                    else:
                        lines.append(word[:1])
                        word = word[1:]
                current_line = ""
            else:
                current_line = word
    if current_line:
        lines.append(current_line)
    return lines if lines else [""]

def build_kudir_pdf(out_path: str, org: Dict[str, Any], period_title: str, rows: List[Dict[str, Any]], totals: Dict[str, Decimal]):
    require_reportlab()
    font = register_ru_font()
    c = canvas.Canvas(out_path, pagesize=A4)
    W, H = A4

    col_num_x = 10*mm
    col_date_x = 22*mm
    col_content_x = 48*mm
    col_income_right = 170*mm
    col_expense_right = 197*mm
    content_max_width = 80*mm
    line_height = 4*mm
    font_size = 7.5

    c.setFont(font, 14)
    c.drawCentredString(W/2, H - 40*mm, "КНИГА УЧЕТА ДОХОДОВ И РАСХОДОВ")
    c.setFont(font, 11)
    c.drawCentredString(W/2, H - 50*mm, "при применении упрощенной системы налогообложения")
    c.drawString(15*mm, H - 65*mm, f"Налогоплательщик: {org.get('name', '')}")
    c.drawString(15*mm, H - 75*mm, f"ИНН: {org.get('inn', '')}  КПП: {org.get('kpp', '')}")
    c.drawString(15*mm, H - 85*mm, f"Период: {period_title}")

    y = H - 100*mm
    c.setFont(font, 9)
    c.setLineWidth(0.5)
    header_top = y + 5*mm
    header_bottom = y - 3*mm
    c.line(col_num_x - 2*mm, header_top, col_expense_right + 5*mm, header_top)
    c.line(col_num_x - 2*mm, header_bottom, col_expense_right + 5*mm, header_bottom)
    c.drawString(col_num_x, y, "№")
    c.drawString(col_date_x, y, "Дата")
    c.drawString(col_content_x, y, "Содержание операции")
    c.drawRightString(col_income_right, y, "Доход")
    c.drawRightString(col_expense_right, y, "Расход")
    y = header_bottom - 2*mm

    def _draw_table_header():
        nonlocal y
        c.setFont(font, 9)
        c.setLineWidth(0.5)
        ht = y + 5*mm
        hb = y - 3*mm
        c.line(col_num_x - 2*mm, ht, col_expense_right + 5*mm, ht)
        c.line(col_num_x - 2*mm, hb, col_expense_right + 5*mm, hb)
        c.drawString(col_num_x, y, "№")
        c.drawString(col_date_x, y, "Дата")
        c.drawString(col_content_x, y, "Содержание операции")
        c.drawRightString(col_income_right, y, "Доход")
        c.drawRightString(col_expense_right, y, "Расход")
        y = hb - 2*mm
        c.setFont(font, font_size)

    c.setFont(font, font_size)
    for i, r in enumerate(rows, 1):
        content = (r.get("content", "") or "")
        content_lines = _wrap_text(content, content_max_width, font, font_size)
        block_height = max(len(content_lines), 1) * line_height + 1*mm

        if y - block_height < 20*mm:
            c.showPage()
            y = H - 15*mm
            _draw_table_header()

        c.drawString(col_num_x, y, str(i))
        c.drawString(col_date_x, y, str(r.get("doc_ref", "")))
        for li, line_text in enumerate(content_lines):
            c.drawString(col_content_x, y - li * line_height, line_text)
        inc = r.get("income", "")
        if inc != "":
            c.drawRightString(col_income_right, y, fmt_money(inc))
        exp = r.get("expense", "")
        if exp != "":
            c.drawRightString(col_expense_right, y, fmt_money(exp))
        y -= block_height

    y -= 5*mm
    if y < 25*mm:
        c.showPage()
        c.setFont(font, 10)
        y = H - 20*mm
    c.setFont(font, 10)
    c.setLineWidth(0.5)
    c.line(col_num_x - 2*mm, y + 5*mm, col_expense_right + 5*mm, y + 5*mm)
    c.drawString(col_content_x, y, "ИТОГО:")
    c.drawRightString(col_income_right, y, fmt_money(totals.get("income", Decimal("0"))))
    c.drawRightString(col_expense_right, y, fmt_money(totals.get("expense", Decimal("0"))))
    c.line(col_num_x - 2*mm, y - 3*mm, col_expense_right + 5*mm, y - 3*mm)
    c.save()


def get_quarter_range(year: int, quarter: int):
    if quarter == 1:
        return date(year, 1, 1), date(year, 3, 31)
    elif quarter == 2:
        return date(year, 4, 1), date(year, 6, 30)
    elif quarter == 3:
        return date(year, 7, 1), date(year, 9, 30)
    else:
        return date(year, 10, 1), date(year, 12, 31)

def get_last_4_quarters():
    today = date.today()
    current_quarter = (today.month - 1) // 3 + 1
    current_year = today.year
    quarters = []
    for i in range(4):
        q = current_quarter - i
        y = current_year
        while q <= 0:
            q += 4
            y -= 1
        quarters.append((y, q))
    return list(reversed(quarters))

def extract_commission_from_purpose(purpose: str) -> Decimal:
    """Извлекает комиссию из назначения платежа (например: Ком-сия по операциям 140.30 р.)"""
    if not purpose:
        return Decimal("0")
    purpose_lower = purpose.lower()
    match = re.search(r'ком[и\-]?сия[^0-9]*(\d+[.,]\d{2})', purpose_lower)
    if match:
        amount_str = match.group(1).replace(",", ".")
        try:
            return Decimal(amount_str)
        except:
            return Decimal("0")
    return Decimal("0")

def calc_income_for_period(start: date, end: date) -> Decimal:
    income = Decimal("0")
    for r in STATE.bank_rows:
        d = parse_date_ddmmyyyy(r.get("date", ""))
        if not d or not (start <= d <= end):
            continue
        inc = Decimal(str(r.get("incoming") or 0))
        if inc > 0:
            income += inc
            category = (r.get("category") or "").lower()
            if "коворкинг" in category:
                commission = extract_commission_from_purpose(r.get("purpose", ""))
                income += commission
    for r in STATE.cash_rows:
        d = parse_date_ddmmyyyy(r.get("date", ""))
        if not d or not (start <= d <= end):
            continue
        amt = Decimal(str(r.get("amount") or 0))
        if amt > 0:
            income += amt
    return money2(income)

def calc_usn(start: Optional[date], end: Optional[date], tax_1_percent: bool = False, fixed_payment: bool = False, advances: bool = False) -> Dict[str, Any]:
    income = Decimal("0")
    for r in STATE.bank_rows:
        d = parse_date_ddmmyyyy(r.get("date", ""))
        if start and end and (not d or not (start <= d <= end)):
            continue
        inc = Decimal(str(r.get("incoming") or 0))
        if inc > 0:
            income += inc
            category = (r.get("category") or "").lower()
            if "коворкинг" in category:
                commission = extract_commission_from_purpose(r.get("purpose", ""))
                income += commission
    for r in STATE.cash_rows:
        d = parse_date_ddmmyyyy(r.get("date", ""))
        if start and end and (not d or not (start <= d <= end)):
            continue
        amt = Decimal(str(r.get("amount") or 0))
        if amt > 0:
            income += amt
    usn_6 = money2(income * Decimal("0.06")) if income > 0 else Decimal("0.00")
    tax_1 = Decimal("0.00")
    if tax_1_percent and income > Decimal("300000"):
        tax_1 = money2((income - Decimal("300000")) * Decimal("0.01"))
    fixed = Decimal("57390") if fixed_payment else Decimal("0.00")
    advances_list = []
    advances_total = Decimal("0.00")
    if advances:
        quarters = get_last_4_quarters()
        for y, q in quarters:
            q_start, q_end = get_quarter_range(y, q)
            q_income = calc_income_for_period(q_start, q_end)
            q_usn = money2(q_income * Decimal("0.06"))
            advances_list.append({"year": y, "quarter": q, "income": q_income, "usn": q_usn})
            advances_total += q_usn
    final_usn = money2(usn_6 - tax_1 - fixed - advances_total)
    if final_usn < 0:
        final_usn = Decimal("0.00")
    after = money2(income - final_usn)
    return {"income": money2(income), "usn_6": usn_6, "tax_1": tax_1, "fixed": fixed, "advances": advances_list, "advances_total": advances_total, "final_usn": final_usn, "after": after}


def page_usn(path: str, qd: Dict[str, str], flash: str = "") -> str:
    mode = (qd.get("mode") or "all").strip()
    dfrom = (qd.get("from") or "").strip()
    dto = (qd.get("to") or "").strip()
    tax_1_percent = (qd.get("tax_1_percent") or "") == "1"
    fixed_payment = (qd.get("fixed_payment") or "") == "1"
    advances = (qd.get("advances") or "") == "1"
    start = end = None
    title = "За всё время"
    res = {"income": Decimal("0"), "usn_6": Decimal("0"), "tax_1": Decimal("0"), "fixed": Decimal("0"), "advances": [], "advances_total": Decimal("0"), "final_usn": Decimal("0"), "after": Decimal("0")}
    try:
        start, end, title = compute_date_range(mode, dfrom, dto)
        res = calc_usn(start, end, tax_1_percent, fixed_payment, advances)
    except Exception as e:
        flash = str(e)
    chk_tax1 = "checked" if tax_1_percent else ""
    chk_fixed = "checked" if fixed_payment else ""
    chk_advances = "checked" if advances else ""
    advances_html = ""
    if advances and res.get("advances"):
        advances_rows = ""
        for adv in res["advances"]:
            q_name = f"{adv['year']} Q{adv['quarter']}"
            advances_rows += f"<div>{q_name}: доход {h(fmt_money(adv['income']))} → УСН {h(fmt_money(adv['usn']))}</div>"
        advances_html = f"""
        <div class="hr"></div>
        <div class="smallnote"><b>Авансы по кварталам:</b></div>
        {advances_rows}
        <div><b>Итого авансов: {h(fmt_money(res["advances_total"]))}</b></div>
        """
    body = f"""
    <div class="pagehead">
      <div><h1>УСН 6%</h1><div class="smallnote">Калькулятор налога по упрощённой системе.</div></div>
    </div>
    <form method="GET" action="/tax/usn" class="panel" style="margin-bottom:12px">
      <div class="smallnote"><b>Опции</b></div>
      <label style="display:flex;gap:8px;align-items:center;margin:0;padding:8px 0">
        <input type="checkbox" name="tax_1_percent" value="1" {chk_tax1}/>
        Налог 1% (если поступления &gt; 300 000, вычитаем 1% от превышения из УСН)
      </label>
      <label style="display:flex;gap:8px;align-items:center;margin:0;padding:8px 0">
        <input type="checkbox" name="fixed_payment" value="1" {chk_fixed}/>
        Фиксированная выплата (вычитаем 57 390 из УСН)
      </label>
      <label style="display:flex;gap:8px;align-items:center;margin:0;padding:8px 0">
        <input type="checkbox" name="advances" value="1" {chk_advances}/>
        Авансы (вычитаем УСН 6% за последние 4 квартала)
      </label>
      <div class="hr"></div>
      {_period_form_html("/tax/usn", mode, dfrom, dto)}
    </form>
    <div class="panel">
      <div class="badge2"><b>{h(title)}</b></div>
      <div class="hr"></div>
      <div style="display:grid;grid-template-columns: 1fr 1fr;gap:10px">
        <div>Поступления: <b>{h(fmt_money(res["income"]))}</b></div>
        <div>УСН 6%: <b>{h(fmt_money(res["usn_6"]))}</b></div>
        <div>Вычет 1%: <b>{h(fmt_money(res["tax_1"]))}</b></div>
        <div>Фикс. выплата: <b>{h(fmt_money(res["fixed"]))}</b></div>
        <div>Вычет авансов: <b>{h(fmt_money(res["advances_total"]))}</b></div>
        <div>Итого УСН к уплате: <b>{h(fmt_money(res["final_usn"]))}</b></div>
        <div>Остаток после УСН: <b>{h(fmt_money(res["after"]))}</b></div>
      </div>
      {advances_html}
    </div>
    """
    return render_layout("/tax/usn", "УСН", "Налоги → УСН", body, flash=flash_box(flash))


def application(environ, start_response):
    try:
        path = environ.get("PATH_INFO", "") or "/"
        method = environ.get("REQUEST_METHOD", "GET").upper()
        qd = qs(environ)

        if path == "/api/check-invoice" and method == "GET":
            import json as json_mod
            invoice_number = (qd.get("invoice", [""])[0] or "").strip()
            do_refresh = (qd.get("refresh", ["0"])[0] or "0").strip()

            if do_refresh == "1":
                try:
                    fetched = fetch_bank_statement_from_email()
                except Exception:
                    fetched = 0
            else:
                fetched = None

            if not invoice_number:
                body = json_mod.dumps({"error": "Параметр invoice обязателен", "example": "/api/check-invoice?invoice=00111&refresh=1"}, ensure_ascii=False)
                start_response("400 Bad Request", [("Content-Type", "application/json; charset=utf-8")])
                return [body.encode("utf-8")]

            search_pattern = f"оплата по счету"
            found_rows = []
            for r in STATE.bank_rows:
                purpose = (r.get("purpose") or "").lower()
                if search_pattern in purpose and invoice_number in purpose:
                    found_rows.append({
                        "date": r.get("date", ""),
                        "amount": str(r.get("incoming") or r.get("outgoing") or "0"),
                        "counterparty": r.get("counterparty", ""),
                        "purpose": r.get("purpose", ""),
                        "category": r.get("category", ""),
                    })

            result = {
                "invoice": invoice_number,
                "paid": len(found_rows) > 0,
                "payments_count": len(found_rows),
                "payments": found_rows,
            }
            if fetched is not None:
                result["statements_refreshed"] = True
                result["new_operations_loaded"] = fetched

            body = json_mod.dumps(result, ensure_ascii=False, indent=2)
            start_response("200 OK", [("Content-Type", "application/json; charset=utf-8")])
            return [body.encode("utf-8")]

        if path.startswith("/download/"):
            token = path.split("/download/", 1)[1]
            fpath = os.path.join(DOWNLOAD_DIR, token)
            ext = os.path.splitext(fpath)[1].lower()
            ctype = "application/octet-stream"
            if ext == ".csv":
                ctype = "text/csv"
            elif ext == ".xlsx":
                ctype = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            elif ext == ".pdf":
                ctype = "application/pdf"
            return serve_file_download(start_response, fpath, ctype)

        if path == "/action/save" and method == "POST":
            STATE.save()
            return redirect("/bank?m=" + urlencode({"m": "Состояние сохранено."})[2:], start_response)

        if path == "/action/bank/upload" and method == "POST":
            form = parse_post_form(environ)
            item = get_upload(form, "file")
            if item is None:
                return redirect("/bank?m=" + urlencode({"m": "Файл не выбран."})[2:], start_response)

            filename = getattr(item, "filename", "") or "bank.txt"
            data = item.file.read() if getattr(item, "file", None) is not None else b""
            if not data:
                return redirect("/bank?m=" + urlencode({"m": "Файл пустой или не прочитан."})[2:], start_response)

            tmp = os.path.join(DOWNLOAD_DIR, f"upload_{int(time.time())}_{safe_filename(filename)}")
            with open(tmp, "wb") as f:
                f.write(data)

            new_rows = parse_client_bank_file(tmp)
            
            cp_map = STATE.cp_category_map
            user_map = STATE.user_category_map
            
            for r in new_rows:
                r["category"] = detect_category(
                    r.get("counterparty", ""),
                    r.get("purpose", ""),
                    cp_map,
                    user_map
                )

            existing_fp = set(bank_row_fingerprint(x) for x in STATE.bank_rows)
            new_fp = [bank_row_fingerprint(x) for x in new_rows]
            
            added = 0
            for r, fp in zip(new_rows, new_fp):
                if fp not in existing_fp:
                    STATE.auto_upsert_counterparty_from_bank_row(r)
                    STATE.add_bank_row(r)
                    added += 1

            STATE.sanitize_names_and_inn()
            STATE.save()
            return redirect("/bank?m=" + urlencode({"m": f"Добавлено операций: {added}"})[2:], start_response)

        if path == "/action/bank/fetch-email" and method == "POST":
            added = fetch_bank_statement_from_email()
            status = EMAIL_LAST_DOWNLOAD.get("status", "")
            msg = f"Загружено с почты: {added} операций. {status}"
            return redirect("/bank?m=" + urlencode({"m": msg})[2:], start_response)

        if path == "/action/bank/set-skip" and method == "POST":
            form = parse_post_form(environ)
            rid = (form.get("id") or "").strip()
            val = (form.get("value") or "0").strip()
            rr = STATE.get_bank_row(rid)
            if rr and decimal_from_str(rr.get("outgoing") or 0) > 0:
                STATE.update_bank_row(rid, {"skip_outgoing": (val == "1")})
                STATE.save()
            return redirect("/bank", start_response)

        if path == "/action/bank/assign-category" and method == "POST":
            form = parse_post_form(environ)
            rid = (form.get("id") or "").strip()
            cat = norm_spaces((form.get("category") or "").strip())
            return_url = form.get("return_url", "/bank").strip() or "/bank"
            if not return_url.startswith("/"):
                return_url = "/bank"
            rr = STATE.get_bank_row(rid)
            if rr and cat:
                cp_name = (rr.get("counterparty") or "").strip()
                if cp_name:
                    STATE.add_user_category(norm_text(cp_name), cat)
                    STATE.recalc_bank_categories()
                    STATE.save()
                    sep = "&" if "?" in return_url else "?"
                    return redirect(return_url + sep + urlencode({"m": f"Контрагент подписан: {cp_name} → {cat}"}), start_response)
            return redirect(return_url + ("&" if "?" in return_url else "?") + urlencode({"m": "Не удалось подписать."}), start_response)

        if path == "/action/bank/resolve-disputed" and method == "POST":
            form = parse_post_form(environ)
            rid = (form.get("id") or "").strip()
            cat = norm_spaces((form.get("category") or "").strip())
            return_url = form.get("return_url", "/bank").strip() or "/bank"
            if not return_url.startswith("/"):
                return_url = "/bank"
            rr = STATE.get_bank_row(rid)
            if rr and cat:
                STATE.update_bank_row(rid, {"category": cat})
                STATE.save()
                sep = "&" if "?" in return_url else "?"
                return redirect(return_url + sep + urlencode({"m": f"Статья назначена: {cat}"}), start_response)
            return redirect(return_url + ("&" if "?" in return_url else "?") + urlencode({"m": "Не удалось назначить статью."}), start_response)

        if path == "/action/bank/remove-category" and method == "POST":
            form = parse_post_form(environ)
            cp_key = (form.get("cp_key") or "").strip()
            cat = norm_spaces((form.get("category") or "").strip())
            return_url = form.get("return_url", "/bank").strip() or "/bank"
            if not return_url.startswith("/"):
                return_url = "/bank"
            if cp_key and cat:
                STATE.remove_user_category(cp_key, cat)
                STATE.recalc_bank_categories()
                STATE.save()
                sep = "&" if "?" in return_url else "?"
                return redirect(return_url + sep + urlencode({"m": f"Статья удалена: {cat}"}), start_response)
            return redirect(return_url + ("&" if "?" in return_url else "?") + urlencode({"m": "Не удалось удалить статью."}), start_response)

        if path == "/action/bank/bulk-assign" and method == "POST":
            form = parse_post_form(environ)
            counterparties_str = (form.get("counterparties") or "").strip()
            cat = norm_spaces((form.get("category") or "").strip())
            return_url = form.get("return_url", "/bank").strip() or "/bank"
            if not return_url.startswith("/"):
                return_url = "/bank"
            if counterparties_str and cat:
                cp_names = [cp.strip() for cp in counterparties_str.split("|||") if cp.strip()]
                count = 0
                for cp_name in cp_names:
                    STATE.add_user_category(norm_text(cp_name), cat)
                    count += 1
                STATE.recalc_bank_categories()
                STATE.save()
                sep = "&" if "?" in return_url else "?"
                return redirect(return_url + sep + urlencode({"m": f"Статья «{cat}» назначена для {count} контрагентов"}), start_response)
            return redirect(return_url + ("&" if "?" in return_url else "?") + urlencode({"m": "Не удалось назначить статью."}), start_response)

        if path == "/action/bank/cpmap" and method == "POST":
            form = parse_post_form(environ)
            item = get_upload(form, "file")
            if item is None:
                return redirect("/bank?m=" + urlencode({"m": "Файл не выбран."})[2:], start_response)

            filename = getattr(item, "filename", "") or "map.xlsx"
            data = item.file.read() if getattr(item, "file", None) is not None else b""
            if not data:
                return redirect("/bank?m=" + urlencode({"m": "Файл пустой или не прочитан."})[2:], start_response)

            tmp = os.path.join(DOWNLOAD_DIR, f"cpmap_{int(time.time())}_{safe_filename(filename)}")
            with open(tmp, "wb") as f:
                f.write(data)

            mp = load_counterparty_category_map(tmp)
            STATE.cp_category_map = mp
            STATE.cp_map_source = tmp
            STATE.recalc_bank_categories()
            STATE.save()
            return redirect("/bank?m=" + urlencode({"m": f"Справочник загружен: {os.path.basename(filename)} ({len(mp)} строк)"})[2:], start_response)

        if path == "/action/bank/export-csv" and method == "GET":
            token = f"bank_{int(time.time())}.csv"
            fpath = os.path.join(DOWNLOAD_DIR, token)
            bank_rows = STATE.bank_rows
            with open(fpath, "w", newline="", encoding="cp1251") as f:
                w = csv.writer(f, delimiter=";")
                w.writerow(["Не учитывать", "Дата", "Месяц", "Поступление", "Списание", "Статья", "Назначение", "Контрагент", "Вид операции"])
                for r in bank_rows:
                    inc = Decimal(str(r.get("incoming") or 0))
                    out = Decimal(str(r.get("outgoing") or 0))
                    w.writerow([
                        "1" if r.get("skip_outgoing", False) else "",
                        r.get("date", ""), r.get("month", ""),
                        fmt_money(inc) if inc else "",
                        fmt_money(out) if out else "",
                        r.get("category", ""),
                        r.get("purpose", ""),
                        r.get("counterparty", ""),
                        r.get("doctype", ""),
                    ])
            return redirect("/download/" + token, start_response)

        if path == "/action/cash/add" and method == "POST":
            form = parse_post_form(environ)
            r = {
                "id": new_id(),
                "date": form.get("date", "").strip(),
                "nomenclature": form.get("nomenclature", "").strip(),
                "amount": str(decimal_from_str(form.get("amount", "0"))),
            }
            STATE.add_cash_row(r)
            STATE.save()
            return redirect("/cash?m=" + urlencode({"m": "Запись добавлена."})[2:], start_response)

        if path == "/action/cash/update" and method == "POST":
            form = parse_post_form(environ)
            rid = (form.get("id") or "").strip()
            if rid:
                updates = {
                    "date": form.get("date", "").strip(),
                    "nomenclature": form.get("nomenclature", "").strip(),
                    "amount": str(decimal_from_str(form.get("amount", "0"))),
                }
                STATE.update_cash_row(rid, updates)
                STATE.save()
            return redirect("/cash?m=" + urlencode({"m": "Запись обновлена."})[2:], start_response)

        if path == "/action/cash/delete" and method == "POST":
            form = parse_post_form(environ)
            rid = (form.get("id") or "").strip()
            if rid:
                STATE.delete_cash_row(rid)
                STATE.save()
            return redirect("/cash?m=" + urlencode({"m": "Запись удалена."})[2:], start_response)

        if path == "/action/marketplace/add" and method == "POST":
            form = parse_post_form(environ)
            platform = (form.get("platform") or "").strip()
            period_type = (form.get("period_type") or "quarter").strip()
            year = int(form.get("year") or date.today().year)
            quarter = int(form.get("quarter") or 1)
            month = int(form.get("month") or 1)
            input_mode = (form.get("input_mode") or "manual").strip()
            
            amount = Decimal("0")
            if input_mode == "manual":
                amount = decimal_from_str(form.get("amount") or "0")
            else:
                upload = get_upload(form, "file")
                if upload:
                    tmp_path = os.path.join(DOWNLOAD_DIR, f"mp_{new_id()}_{safe_filename(upload.filename)}")
                    with open(tmp_path, "wb") as f:
                        f.write(upload.file.read())
                    if platform.lower() == "ozon":
                        amount = parse_ozon_excel_total(tmp_path)
                    else:
                        amount = parse_wb_pdf_total(tmp_path)
                    try:
                        os.remove(tmp_path)
                    except Exception:
                        pass
            
            period_label = get_period_label(period_type, year, quarter, month)
            date_start, date_end = get_period_dates(period_type, year, quarter, month)
            
            conn = get_db_connection()
            cur = conn.cursor()
            cur.execute("""
                INSERT INTO marketplace_rows (id, platform, period_type, period_label, year, quarter, month, date_start, date_end, amount)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """, (new_id(), platform, period_type, period_label, year, quarter, month, date_start, date_end, amount))
            conn.commit()
            cur.close()
            return_db_connection(conn)
            
            msg = f"Добавлено поступление {platform}: {fmt_money(amount)}"
            return redirect("/marketplace?m=" + urlencode({"m": msg})[2:], start_response)

        if path == "/action/marketplace/delete" and method == "POST":
            form = parse_post_form(environ)
            rid = (form.get("id") or "").strip()
            if rid:
                conn = get_db_connection()
                cur = conn.cursor()
                cur.execute("DELETE FROM marketplace_rows WHERE id = %s", (rid,))
                conn.commit()
                cur.close()
                return_db_connection(conn)
            return redirect("/marketplace?m=" + urlencode({"m": "Запись удалена."})[2:], start_response)

        if path == "/action/cp/save" and method == "POST":
            form = parse_post_form(environ)
            cid = (form.get("id") or "").strip()
            is_our = form.get("is_our") == "1"
            c = {
                "kind": form.get("kind", "Юридическое лицо"),
                "name": form.get("name", "").strip(),
                "inn": form.get("inn", "").strip(),
                "kpp": form.get("kpp", "").strip(),
                "bank": form.get("bank", "").strip(),
                "bik": form.get("bik", "").strip(),
                "corr": form.get("corr", "").strip(),
                "account": form.get("account", "").strip(),
                "legal_address": form.get("legal_address", "").strip(),
                "phone": form.get("phone", "").strip(),
                "is_our_company": is_our,
                "full_name": form.get("full_name", "").strip(),
                "inspection_code": form.get("inspection_code", "").strip(),
                "oktmo": form.get("oktmo", "").strip(),
                "okato": form.get("okato", "").strip(),
                "signatory": form.get("signatory", "").strip(),
                "sfr_reg_number": form.get("sfr_reg_number", "").strip(),
                "pfr_reg_self": form.get("pfr_reg_self", "").strip(),
                "pfr_reg_employees": form.get("pfr_reg_employees", "").strip(),
                "pfr_terr_code": form.get("pfr_terr_code", "").strip(),
                "pfr_terr_organ": form.get("pfr_terr_organ", "").strip(),
                "payment_details": form.get("payment_details", "").strip(),
                "okpo": form.get("okpo", "").strip(),
                "okopf": form.get("okopf", "").strip(),
                "okfs": form.get("okfs", "").strip(),
                "okved1": form.get("okved1", "").strip(),
                "okved2": form.get("okved2", "").strip(),
                "okpo_rosstat": form.get("okpo_rosstat", "").strip(),
            }
            if not c["name"]:
                return redirect("/counterparties?m=" + urlencode({"m": "Наименование обязательно."})[2:], start_response)
            
            if cid:
                STATE.update_counterparty(cid, c)
                if is_our:
                    STATE.set_setting("our_company_id", cid)
                elif STATE.settings.get("our_company_id") == cid:
                    STATE.set_setting("our_company_id", "")
            else:
                c["id"] = new_id()
                STATE.add_counterparty(c)
                if is_our:
                    STATE.set_setting("our_company_id", c["id"])
            
            STATE.save()
            return redirect("/counterparties?m=" + urlencode({"m": "Контрагент сохранён."})[2:], start_response)

        if path == "/action/cp/delete" and method == "POST":
            form = parse_post_form(environ)
            cid = (form.get("id") or "").strip()
            if cid:
                STATE.delete_counterparty(cid)
                STATE.save()
            return redirect("/counterparties?m=" + urlencode({"m": "Контрагент удалён."})[2:], start_response)

        if path == "/health":
            start_response("200 OK", [("Content-Type", "text/plain")])
            return [b"OK"]

        if path == "/" or path == "":
            html = '<!DOCTYPE html><html><head><meta http-equiv="refresh" content="0;url=/bank"></head><body>Redirecting...</body></html>'
            return serve_text(start_response, html)

        if path == "/bank" and method == "GET":
            return serve_text(start_response, page_bank("/bank", qd=qd, flash=qd.get("m","") or ""))

        if path == "/bank/assign" and method == "GET":
            return serve_text(start_response, page_bank_assign("/bank", qd, flash=qd.get("m","") or ""))

        if path == "/cash" and method == "GET":
            return serve_text(start_response, page_cash("/cash", flash=qd.get("m","") or ""))

        if path == "/cash/new" and method == "GET":
            return serve_text(start_response, page_cash_form("/cash", {}, flash=qd.get("m","") or ""))

        if path == "/cash/edit" and method == "GET":
            return serve_text(start_response, page_cash_form("/cash", qd, flash=qd.get("m","") or ""))

        if path == "/marketplace" and method == "GET":
            return serve_text(start_response, page_marketplace("/marketplace", flash=qd.get("m","") or ""))

        if path == "/marketplace/add" and method == "GET":
            return serve_text(start_response, page_marketplace_add("/marketplace", qd, flash=qd.get("m","") or ""))

        if path == "/counterparties" and method == "GET":
            return serve_text(start_response, page_counterparties("/counterparties", flash=qd.get("m","") or ""))

        if path == "/counterparties/new" and method == "GET":
            return serve_text(start_response, page_counterparty_new(flash=qd.get("m","") or ""))

        if path == "/counterparties/edit" and method == "GET":
            return serve_text(start_response, page_counterparty_edit(qd, flash=qd.get("m","") or ""))

        if path == "/search" and method == "GET":
            return serve_text(start_response, page_search(qd, flash=qd.get("m","") or ""))

        if path == "/action/acts/save" and method == "POST":
            form = parse_post_form(environ)
            aid = (form.get("id") or "").strip() or new_id()
            doc_no = (form.get("doc_no") or "").strip()
            doc_date = (form.get("doc_date") or "").strip()
            direction = (form.get("direction") or "provide").strip()
            basis = (form.get("basis") or "").strip()
            vat_mode = (form.get("vat_mode") or "Без налога (НДС)").strip()
            
            def get_party(prefix: str) -> Dict[str, Any]:
                card_name = (form.get(f"{prefix}_card") or "").strip()
                if card_name:
                    c = STATE.find_counterparty_by_name(card_name)
                    if c:
                        return party_from_counterparty(c)
                return {
                    "name": (form.get(f"{prefix}_name") or "").strip(),
                    "inn": (form.get(f"{prefix}_inn") or "").strip(),
                    "kpp": (form.get(f"{prefix}_kpp") or "").strip(),
                    "account": (form.get(f"{prefix}_account") or "").strip(),
                    "bank": (form.get(f"{prefix}_bank") or "").strip(),
                    "bik": (form.get(f"{prefix}_bik") or "").strip(),
                    "corr": (form.get(f"{prefix}_corr") or "").strip(),
                    "address": (form.get(f"{prefix}_address") or "").strip(),
                }
            
            executor = get_party("executor")
            customer = get_party("customer")
            
            try:
                cnt = int(str(form.get("lines_count") or "0"))
            except:
                cnt = 0
            lines = []
            for i in range(cnt):
                name = (form.get(f"ln_name_{i}") or "").strip()
                if not name:
                    continue
                qty = decimal_from_str(form.get(f"ln_qty_{i}") or "0")
                unit = (form.get(f"ln_unit_{i}") or "шт").strip()
                price = decimal_from_str(form.get(f"ln_price_{i}") or "0")
                amount_raw = (form.get(f"ln_amount_{i}") or "").strip()
                amount = decimal_from_str(amount_raw) if amount_raw else money2(qty * price)
                lines.append({"name": name, "qty": str(qty), "unit": unit, "price": str(price), "amount": str(amount)})
            
            act = {
                "id": aid,
                "doc_no": doc_no,
                "doc_date": doc_date,
                "direction": direction,
                "executor": executor,
                "customer": customer,
                "basis": basis,
                "vat_mode": vat_mode,
                "lines": lines,
            }
            
            existing = STATE.get_act_by_id(aid)
            if existing:
                STATE.update_act(aid, act)
            else:
                STATE.add_act(act)
            if basis:
                STATE.add_basis(basis)
            STATE.save()
            return redirect("/acts?m=" + urlencode({"m": "Акт сохранён."})[2:], start_response)

        if path == "/action/acts/delete" and method == "POST":
            form = parse_post_form(environ)
            aid = (form.get("id") or "").strip()
            if aid:
                STATE.delete_act(aid)
                STATE.save()
            return redirect("/acts?m=" + urlencode({"m": "Акт удалён."})[2:], start_response)

        if path == "/action/acts/export" and method == "GET":
            aid = (qd.get("id") or "").strip()
            act = STATE.get_act_by_id(aid)
            if not act:
                return redirect("/acts?m=" + urlencode({"m": "Акт не найден."})[2:], start_response)
            token = f"act_{aid}_{int(time.time())}.xlsx"
            fpath = os.path.join(DOWNLOAD_DIR, token)
            build_act_excel(fpath, act)
            return redirect("/download/" + token, start_response)

        if path == "/action/payments/save" and method == "POST":
            form = parse_post_form(environ)
            poid = (form.get("id") or "").strip() or new_id()
            
            def get_party_po(prefix: str) -> Dict[str, Any]:
                card_name = (form.get(f"{prefix}_card") or "").strip()
                if card_name:
                    c = STATE.find_counterparty_by_name(card_name)
                    if c:
                        return party_from_counterparty(c)
                return {
                    "name": (form.get(f"{prefix}_name") or "").strip(),
                    "inn": (form.get(f"{prefix}_inn") or "").strip(),
                    "kpp": (form.get(f"{prefix}_kpp") or "").strip(),
                    "account": (form.get(f"{prefix}_account") or "").strip(),
                    "bank": (form.get(f"{prefix}_bank") or "").strip(),
                    "bik": (form.get(f"{prefix}_bik") or "").strip(),
                    "corr": (form.get(f"{prefix}_corr") or "").strip(),
                }
            
            po = {
                "id": poid,
                "number": (form.get("number") or "").strip(),
                "date": (form.get("date") or "").strip(),
                "amount": str(decimal_from_str(form.get("amount") or "0")),
                "amount_words": (form.get("amount_words") or "").strip(),
                "pay_type": (form.get("pay_type") or "Электронно").strip(),
                "payer": get_party_po("payer"),
                "receiver": get_party_po("receiver"),
                "purpose": (form.get("purpose") or "").strip(),
                "vid_op": (form.get("vid_op") or "01").strip(),
                "ocher": (form.get("ocher") or "5").strip(),
            }
            existing_po = next((p for p in STATE.payment_orders if p.get("id") == poid), None)
            if existing_po:
                STATE.delete_payment_order(poid)
            STATE.add_payment_order(po)
            STATE.save()
            return redirect("/payments?m=" + urlencode({"m": "Платёжка сохранена."})[2:], start_response)

        if path == "/action/payments/delete" and method == "POST":
            form = parse_post_form(environ)
            poid = (form.get("id") or "").strip()
            if poid:
                STATE.delete_payment_order(poid)
                STATE.save()
            return redirect("/payments?m=" + urlencode({"m": "Платёжка удалена."})[2:], start_response)

        if path == "/action/payments/export" and method == "GET":
            poid = (qd.get("id") or "").strip()
            po = next((p for p in STATE.payment_orders if p.get("id") == poid), None)
            if not po:
                return redirect("/payments?m=" + urlencode({"m": "Платёжка не найдена."})[2:], start_response)
            token = f"payment_{poid}_{int(time.time())}.xlsx"
            fpath = os.path.join(DOWNLOAD_DIR, token)
            build_payment_order_excel(fpath, po)
            return redirect("/download/" + token, start_response)

        if path == "/action/salary/emp-save" and method == "POST":
            form = parse_post_form(environ)
            eid = (form.get("id") or "").strip() or new_id()
            name = (form.get("name") or "").strip()
            if not name:
                return redirect("/salary?m=" + urlencode({"m": "ФИО обязательно."})[2:], start_response)
            
            sal = decimal_from_str(form.get("salary") or "0")
            adv = decimal_from_str(form.get("advance") or "0")
            main_val = form.get("main", "").strip()
            if not main_val and sal > 0:
                main_computed = sal - adv if sal >= adv else Decimal("0")
            else:
                main_computed = decimal_from_str(main_val)
            
            emp = {
                "id": eid,
                "name": name,
                "inn": (form.get("inn") or "").strip(),
                "passport": (form.get("passport") or "").strip(),
                "passport_issued": (form.get("passport_issued") or "").strip(),
                "bank": (form.get("bank") or "").strip(),
                "bik": (form.get("bik") or "").strip(),
                "corr": (form.get("corr") or "").strip(),
                "account": (form.get("account") or "").strip(),
                "salary": str(sal),
                "advance": str(adv),
                "main": str(main_computed),
            }
            existing_emp = STATE.get_employee_by_id(eid)
            if existing_emp:
                STATE.update_employee(eid, emp)
            else:
                STATE.add_employee(emp)
            STATE.save()
            return redirect("/salary?m=" + urlencode({"m": "Карточка сотрудника сохранена."})[2:], start_response)

        if path == "/action/salary/emp-delete" and method == "POST":
            form = parse_post_form(environ)
            eid = (form.get("id") or "").strip()
            if eid:
                if any(p.get("employee_id") == eid for p in STATE.salary_payments):
                    return redirect("/salary?m=" + urlencode({"m": "Есть выплаты этому сотруднику. Сначала удалите выплаты."})[2:], start_response)
                STATE.delete_employee(eid)
                STATE.save()
            return redirect("/salary?m=" + urlencode({"m": "Сотрудник удалён."})[2:], start_response)

        if path == "/action/salary/create-pay" and method == "POST":
            form = parse_post_form(environ)
            eid = (form.get("employee_id") or "").strip()
            month = (form.get("month") or "").strip()
            ptype = (form.get("type") or "advance").strip()
            
            if not eid:
                return redirect("/salary?m=" + urlencode({"m": "Выберите сотрудника."})[2:], start_response)
            emp = STATE.get_employee_by_id(eid)
            if not emp:
                return redirect("/salary?m=" + urlencode({"m": "Сотрудник не найден."})[2:], start_response)
            
            if not re.fullmatch(r"\d{2}\.\d{4}", month):
                return redirect("/salary?m=" + urlencode({"m": "Месяц должен быть в формате MM.YYYY"})[2:], start_response)
            
            for p in STATE.salary_payments:
                if p.get("employee_id") == eid and p.get("month") == month and p.get("pay_type") == ptype:
                    return redirect("/salary?m=" + urlencode({"m": "Нельзя выдать два раза один и тот же вид выплаты за месяц."})[2:], start_response)
            
            amt = decimal_from_str(emp.get("advance") or "0") if ptype == "advance" else decimal_from_str(emp.get("main_part") or emp.get("main") or "0")
            if amt <= 0:
                return redirect("/salary?m=" + urlencode({"m": "Сумма выплаты <= 0. Проверьте карточку сотрудника."})[2:], start_response)
            
            our = STATE.get_our_company_card()
            if not our:
                return redirect("/salary?m=" + urlencode({"m": "Не задана карточка нашей организации (Справочники → Контрагенты)."})[2:], start_response)
            
            po_id = new_id()
            po = {
                "id": po_id,
                "number": "",
                "date": format_ddmmyyyy(date.today()),
                "amount": str(money2(amt)),
                "amount_words": "",
                "pay_type": "Электронно",
                "payer": {
                    "name": our.get("name", ""),
                    "inn": our.get("inn", ""),
                    "kpp": our.get("kpp", ""),
                    "bank": our.get("bank", ""),
                    "bik": our.get("bik", ""),
                    "corr": our.get("corr", ""),
                    "account": our.get("account", ""),
                },
                "receiver": {
                    "name": emp.get("name", ""),
                    "inn": emp.get("inn", ""),
                    "kpp": "",
                    "bank": emp.get("bank", ""),
                    "bik": emp.get("bik", ""),
                    "corr": emp.get("corr", ""),
                    "account": emp.get("account", ""),
                },
                "purpose": f"Выплата заработной платы ({'аванс' if ptype=='advance' else 'основная часть'}) за {month}",
                "vid_op": "01",
                "ocher": "5",
                "source": "salary",
            }
            STATE.add_payment_order_with_source(po)
            
            pay = {
                "id": new_id(),
                "employee_id": eid,
                "month": month,
                "type": ptype,
                "amount": str(money2(amt)),
                "payment_order_id": po_id,
            }
            STATE.add_salary_payment(pay)
            STATE.save()
            return redirect("/salary?m=" + urlencode({"m": "Выплата создана. Платёжка добавлена в раздел «Платёжные поручения»."})[2:], start_response)

        if path == "/action/salary/pay-delete" and method == "POST":
            form = parse_post_form(environ)
            sid = (form.get("id") or "").strip()
            if sid:
                sp = next((x for x in STATE.salary_payments if x.get("id") == sid), None)
                if sp:
                    po_id = sp.get("payment_order_id")
                    if po_id:
                        STATE.delete_payment_order(po_id)
                    STATE.delete_salary_payment(sid)
                    STATE.save()
            return redirect("/salary?m=" + urlencode({"m": "Выплата и платёжка удалены."})[2:], start_response)

        if path == "/action/upd/upload" and method == "POST":
            form = parse_post_form(environ)
            item = get_upload(form, "file")
            if item is None:
                return redirect("/upd?m=" + urlencode({"m": "Файл не выбран."})[2:], start_response)
            filename = getattr(item, "filename", "") or "upd.html"
            data = item.file.read() if getattr(item, "file", None) else b""
            if not data:
                return redirect("/upd?m=" + urlencode({"m": "Файл пустой."})[2:], start_response)
            tmp = os.path.join(DOWNLOAD_DIR, f"upd_{int(time.time())}_{safe_filename(filename)}")
            with open(tmp, "wb") as f:
                f.write(data)
            try:
                upd_row = parse_upd_html(tmp)
                STATE.add_upd_row(upd_row)
                STATE.save()
                return redirect("/upd?m=" + urlencode({"m": f"УПД № {upd_row.get('doc_no','')} добавлен."})[2:], start_response)
            except Exception as e:
                return redirect("/upd?m=" + urlencode({"m": f"Ошибка парсинга: {str(e)}"})[2:], start_response)

        if path == "/action/upd/delete" and method == "POST":
            form = parse_post_form(environ)
            rid = (form.get("id") or "").strip()
            if rid:
                STATE.delete_upd_row(rid)
                STATE.save()
            return redirect("/upd?m=" + urlencode({"m": "УПД удалён."})[2:], start_response)

        if path == "/action/reports/kudir-pdf" and method == "GET":
            dfrom = (qd.get("from") or "").strip()
            dto = (qd.get("to") or "").strip()
            d1 = parse_date_ddmmyyyy(dfrom)
            d2 = parse_date_ddmmyyyy(dto)
            if not d1 or not d2 or d2 < d1:
                return redirect("/reports/kudir?m=" + urlencode({"m": "Неверный период."})[2:], start_response)
            title, rows, totals = calc_kudir(d1, d2)
            our = STATE.get_our_company_card() or {"name": OUR_COMPANY_DEFAULT_NAME, "inn": "", "kpp": ""}
            token = f"kudir_{int(time.time())}.pdf"
            fpath = os.path.join(DOWNLOAD_DIR, token)
            build_kudir_pdf(fpath, our, title, rows, totals)
            return redirect("/download/" + token, start_response)

        if path == "/acts" and method == "GET":
            return serve_text(start_response, page_acts("/acts", flash=qd.get("m","") or ""))
        if path == "/acts/new" and method == "GET":
            return serve_text(start_response, page_act_editor("new", qd, flash=qd.get("m","") or ""))
        if path == "/acts/edit" and method == "GET":
            return serve_text(start_response, page_act_editor("edit", qd, flash=qd.get("m","") or ""))
        if path == "/payments" and method == "GET":
            return serve_text(start_response, page_payments("/payments", flash=qd.get("m","") or ""))
        if path == "/payments/new" and method == "GET":
            return serve_text(start_response, page_payment_editor("new", qd, flash=qd.get("m","") or ""))
        if path == "/payments/edit" and method == "GET":
            return serve_text(start_response, page_payment_editor("edit", qd, flash=qd.get("m","") or ""))
        if path == "/salary" and method == "GET":
            return serve_text(start_response, page_salary("/salary", flash=qd.get("m","") or ""))
        if path == "/salary/emp-new" and method == "GET":
            return serve_text(start_response, page_employee_editor("new", qd, flash=qd.get("m","") or ""))
        if path == "/salary/emp-edit" and method == "GET":
            return serve_text(start_response, page_employee_editor("edit", qd, flash=qd.get("m","") or ""))
        if path == "/upd" and method == "GET":
            return serve_text(start_response, page_upd("/upd", flash=qd.get("m","") or ""))
        if path == "/realization" and method == "GET":
            return serve_text(start_response, page_realization("/realization", flash=qd.get("m","") or ""))
        if path == "/reports/op-profit" and method == "GET":
            return serve_text(start_response, page_op_profit("/reports/op-profit", qd, flash=qd.get("m","") or ""))
        if path == "/reports/op-profit/details" and method == "GET":
            return serve_text(start_response, page_op_profit_details("/reports/op-profit/details", qd, flash=qd.get("m","") or ""))
        if path == "/reports/recon" and method == "GET":
            return serve_text(start_response, page_recon("/reports/recon", qd, flash=qd.get("m","") or ""))
        if path == "/reports/kudir" and method == "GET":
            return serve_text(start_response, page_kudir("/reports/kudir", qd, flash=qd.get("m","") or ""))
        if path == "/tax/usn" and method == "GET":
            return serve_text(start_response, page_usn("/tax/usn", qd, flash=qd.get("m","") or ""))

        return not_found(start_response)

    except Exception as e:
        tb = traceback.format_exc()
        html = f"<h1>Ошибка</h1><pre>{h(str(e))}</pre><pre>{h(tb)}</pre>"
        return serve_text(start_response, html, status="500 Internal Server Error")


class QuietHandler(WSGIRequestHandler):
    def log_message(self, format, *args):
        pass


IMAP_SERVER = "imap.gmail.com"
EMAIL_LAST_DOWNLOAD = {"time": None, "status": "", "added": 0}

def _get_last_bank_date() -> Optional[date]:
    """Определяет дату последней загруженной банковской операции"""
    last_d = None
    for r in STATE.bank_rows:
        d = parse_date_ddmmyyyy(r.get("date", ""))
        if d and (last_d is None or d > last_d):
            last_d = d
    return last_d

def fetch_bank_statement_from_email():
    """Скачивает банковские выписки с почты, начиная от последней загруженной даты"""
    global EMAIL_LAST_DOWNLOAD
    email_account = os.environ.get("EMAIL_ACCOUNT", "")
    email_password = os.environ.get("EMAIL_PASSWORD", "")
    
    if not email_account or not email_password:
        EMAIL_LAST_DOWNLOAD = {"time": datetime.now(), "status": "Не заданы EMAIL_ACCOUNT/EMAIL_PASSWORD", "added": 0}
        return 0
    
    try:
        last_loaded_date = _get_last_bank_date()

        mail = imaplib.IMAP4_SSL(IMAP_SERVER, 993)
        mail.login(email_account, email_password)
        mail.select("INBOX")
        
        status, messages = mail.search(None, "ALL")
        if status != "OK":
            EMAIL_LAST_DOWNLOAD = {"time": datetime.now(), "status": "Ошибка поиска писем", "added": 0}
            return 0
        
        mail_ids = messages[0].split()
        statement_emails = []
        
        for mail_id in mail_ids:
            status, msg_data = mail.fetch(mail_id, "(RFC822)")
            if status != "OK":
                continue
            
            raw_email = msg_data[0][1]
            msg = email_lib.message_from_bytes(raw_email)
            
            subject_raw = msg.get("Subject", "")
            subject_decoded, encoding = decode_header(subject_raw)[0]
            if isinstance(subject_decoded, bytes):
                subject = subject_decoded.decode(encoding or "utf-8", errors="ignore")
            else:
                subject = subject_decoded
            
            if "выписка" not in subject.lower():
                continue
            
            date_tuple = email.utils.parsedate_tz(msg["Date"])
            if date_tuple:
                email_date = datetime.fromtimestamp(email.utils.mktime_tz(date_tuple))
                if last_loaded_date and email_date.date() <= last_loaded_date:
                    continue
                statement_emails.append((email_date, msg))
        
        if not statement_emails:
            EMAIL_LAST_DOWNLOAD = {"time": datetime.now(), "status": "Нет новых писем с выписками", "added": 0}
            mail.close()
            mail.logout()
            return 0
        
        statement_emails.sort(key=lambda x: x[0])
        if not last_loaded_date and len(statement_emails) > 30:
            statement_emails = statement_emails[-30:]

        total_added = 0
        processed_files = []
        existing_fp = set(bank_row_fingerprint(x) for x in STATE.bank_rows)
        
        for email_date, email_msg in statement_emails:
            txt_data = None
            txt_filename = None
            
            for part in email_msg.walk():
                if part.get_content_maintype() == "multipart":
                    continue
                if part.get("Content-Disposition") is None:
                    continue
                
                filename_raw = part.get_filename()
                if filename_raw:
                    fn_decoded, fn_encoding = decode_header(filename_raw)[0]
                    if isinstance(fn_decoded, bytes):
                        filename = fn_decoded.decode(fn_encoding or "utf-8", errors="ignore")
                    else:
                        filename = fn_decoded
                    
                    if filename.lower().endswith(".txt"):
                        txt_data = part.get_payload(decode=True)
                        txt_filename = filename
                        break
            
            if not txt_data:
                continue
            
            tmp = os.path.join(DOWNLOAD_DIR, f"email_{int(time.time())}_{safe_filename(txt_filename)}")
            with open(tmp, "wb") as f:
                f.write(txt_data)
            
            new_rows = parse_client_bank_file(tmp)
            
            cp_map = STATE.cp_category_map
            user_map = STATE.user_category_map
            
            for r in new_rows:
                r["category"] = detect_category(
                    r.get("counterparty", ""),
                    r.get("purpose", ""),
                    cp_map,
                    user_map
                )
            
            new_fp = [bank_row_fingerprint(x) for x in new_rows]
            
            added = 0
            for r, fp in zip(new_rows, new_fp):
                if fp not in existing_fp:
                    STATE.auto_upsert_counterparty_from_bank_row(r)
                    STATE.add_bank_row(r)
                    existing_fp.add(fp)
                    added += 1
            
            total_added += added
            if txt_filename:
                processed_files.append(txt_filename)
        
        if total_added > 0:
            STATE.sanitize_names_and_inn()
            STATE.save()
        
        files_info = ", ".join(processed_files[-3:]) if processed_files else "нет файлов"
        EMAIL_LAST_DOWNLOAD = {"time": datetime.now(), "status": f"OK ({len(processed_files)} файлов: {files_info})", "added": total_added}
        return total_added
        
    except Exception as e:
        EMAIL_LAST_DOWNLOAD = {"time": datetime.now(), "status": f"Ошибка: {str(e)[:50]}", "added": 0}
        return 0


def email_scheduler_thread():
    """Фоновый поток для автоматической загрузки выписок в 12:00"""
    last_run_date = None
    target_hour = 12
    while True:
        now = datetime.now()
        today = now.date()
        if now.hour >= target_hour and last_run_date != today:
            try:
                fetch_bank_statement_from_email()
                last_run_date = today
                print(f"[Планировщик] Выписка загружена в {now.strftime('%H:%M:%S')}")
            except Exception as e:
                print(f"[Планировщик] Ошибка: {e}")
                last_run_date = today
        time.sleep(60)


_db_initialized = False

def _init_db_background():
    global _db_initialized
    try:
        print("Инициализация базы данных...")
        init_database()
        _db_initialized = True
        print("База данных готова.")
    except Exception as e:
        print(f"Ошибка инициализации БД: {e}")

def run_server(host="0.0.0.0", port=5000):
    db_thread = threading.Thread(target=_init_db_background, daemon=True)
    db_thread.start()
    
    scheduler = threading.Thread(target=email_scheduler_thread, daemon=True)
    scheduler.start()
    print("Запущен планировщик загрузки выписок с почты (12:00 ежедневно)")
    
    print(f"Запуск сервера на http://{host}:{port}")
    server = make_server(host, port, application, handler_class=QuietHandler)
    server.serve_forever()


if __name__ == "__main__":
    run_server()
